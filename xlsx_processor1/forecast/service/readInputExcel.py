# import pandas as pd
# from openpyxl import load_workbook
# from multiprocessing import Pool, cpu_count
# from typing import Tuple, Dict
# import logging

# def read_single_sheet(args: Tuple[str, str]) -> Tuple[str, pd.DataFrame]:
#     """Read a single sheet from an Excel file."""
#     input_path, sheet_name = args
#     df = pd.read_excel(input_path, sheet_name=sheet_name, engine="openpyxl")
#     return sheet_name, df

# def read_input_excel(input_path: str) -> Tuple[Dict[str, pd.DataFrame], pd.DataFrame]:
#     """
#     Reads all sheets from the input Excel file in parallel and loads
#     a fixed return QA file.

#     Args:
#         input_path (str): Path to the main Excel file.

#     Returns:
#         Tuple[Dict[str, pd.DataFrame], pd.DataFrame]: 
#             - Dictionary of sheet name to DataFrame.
#             - DataFrame of return QA data.
#     """
#     # Load sheet names quickly using openpyxl
#     workbook = load_workbook(input_path, read_only=True)
#     sheet_names = workbook.sheetnames

#     # Prepare arguments for parallel processing
#     args = [(input_path, sheet) for sheet in sheet_names]
#     with Pool(processes=min(cpu_count(), len(sheet_names))) as pool:
#         results = pool.map(read_single_sheet, args)

#     # Create a dictionary of all sheets
#     sheets = {sheet: df for sheet, df in results}

#     # Load return quantity data from fixed path
#     return_qty_file_path = "forecast/service/Macys returns 2025.xlsx"
#     return_qty_df = pd.read_excel(return_qty_file_path, engine="openpyxl")

#     holidays_df = pd.read_excel(r"media/Images & Category List Replen Items.xlsx", sheet_name="Replen Items Images & Category")
#     print("holidays_df shape:", holidays_df.shape)
#     print("holidays_df columns:", holidays_df.columns.tolist())
#     print("Sheets in the Excel file:")
#     # Logging sheet info
#     logging.info("Sheets in the Excel file: %s", ", ".join(sheets.keys()))
#     logging.info("Successfully read the input Excel file and return quantity file.")

#     return sheets, return_qty_df, holidays_df

import pandas as pd
from openpyxl import load_workbook
from multiprocessing import Pool, cpu_count
from typing import Tuple, Dict
import logging
import io
from django.core.files.storage import default_storage

def read_single_sheet(args: Tuple[str, str]) -> Tuple[str, pd.DataFrame]:
    """Read a single sheet from a local Excel file."""
    input_path, sheet_name = args
    df = pd.read_excel(input_path, sheet_name=sheet_name, engine="openpyxl")
    return sheet_name, df

def read_single_sheet_from_buffer(args: Tuple[bytes, str]) -> Tuple[str, pd.DataFrame]:
    """Read a single sheet from S3 file content with parallel processing."""
    file_content, sheet_name = args
    file_buffer = io.BytesIO(file_content)
    df = pd.read_excel(file_buffer, sheet_name=sheet_name, engine="openpyxl")
    return sheet_name, df

def read_input_excel(input_source) -> Tuple[Dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    """Read Excel file with parallel processing."""
    
    if 'processed_files/' in str(input_source):
        # S3 file - parallel processing
        file_obj = default_storage.open(input_source, 'rb')
        file_content = file_obj.read()
        file_obj.close()
        
        # Get sheet names
        temp_buffer = io.BytesIO(file_content)
        workbook = load_workbook(temp_buffer, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        # Parallel processing
        args = [(file_content, sheet_name) for sheet_name in sheet_names]
        with Pool(processes=min(cpu_count(), len(sheet_names))) as pool:
            results = pool.map(read_single_sheet_from_buffer, args)
        
        sheets = {sheet: df for sheet, df in results}
        
    else:
        # Local file - parallel processing
        workbook = load_workbook(input_source, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        args = [(input_source, sheet) for sheet in sheet_names]
        with Pool(processes=min(cpu_count(), len(sheet_names))) as pool:
            results = pool.map(read_single_sheet, args)
        
        sheets = {sheet: df for sheet, df in results}

    # Load fixed files
    return_qty_df = pd.read_excel("forecast/service/Macys returns 2025.xlsx", engine="openpyxl")
    holidays_df = pd.read_excel(r"media/Images & Category List Replen Items.xlsx", sheet_name="Replen Items Images & Category")
    
    logging.info("Successfully read Excel file with parallel processing.")
    return sheets, return_qty_df, holidays_df