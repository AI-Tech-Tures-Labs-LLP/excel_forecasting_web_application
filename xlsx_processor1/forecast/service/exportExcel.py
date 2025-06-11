# exportExcel.pyVariableLoader
# Standard library imports
import os
from multiprocessing import Pool, cpu_count
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from django.conf import settings
from django.db import transaction
import re
import json
# Third-party imports
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (Alignment,Border,Font,PatternFill,GradientFill,Side)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime

# Local application imports
from .readInputExcel import read_input_excel
from forecast.service.adddatabase import save_macys_projection_receipts, save_monthly_forecasts, save_rolling_forecasts
from forecast.models import MonthlyForecast, ProductDetail, StoreForecast, ComForecast, OmniForecast, RetailInfo
from forecast.service.staticVariable import OMNI_RENAME_MAP, COM_RENAME_MAP, STORE_RENAME_MAP,MONTH_COLUMN_MAP,MONTH_DATA,MONTHLY_VALUES,ALL_VALUES
import calendar
from forecast.service.getretailinfo import get_previous_retail_week
from forecast.service.createDataframe import DataFrameBuilder
from forecast.service.sales_forecasting_algorithmpartwithcom_new_log import algorithm
from forecast.service.var import VariableLoader
import logging
from forecast.service.utils import generate_std_period


def process_data(input_path, file_path, month_from, month_to, percentage, input_tuple,  sheet_object, current_date):
    current_date = datetime(2025,5,8)
    logging.info(f"Input path: {input_path}, File path: {file_path}, Month from: {month_from}, Month to: {month_to}, Percentage: {percentage}, Input tuple: {input_tuple}, Current date: {current_date}")

    sheets, return_qty_df = read_input_excel(input_path)

    std_period = generate_std_period(month_from, month_to)
    current_month_sales_percentage = float(percentage)

    logging.info(f"Current month sales percentage: {current_month_sales_percentage} , Standard period: {std_period}")

    builder = DataFrameBuilder(sheets, return_qty_df)

    builder.build()

    df_outputs = builder.get_outputs()

    logging.info("Report Grouping DataFrame: %s", df_outputs['report_grouping_df'].head())


    # category = [
    #     ('Bridge Gem', '742'),
    #     ('Gold', '746'),
    #     ('Gold', '262&270'),
    #     ('Womens Silver', '260&404'),
    #     ('Precious', '264&268'),
    #     ('Fine Pearl', '265&271'),
    #     ('Semi', '272&733'),
    #     ('Diamond', '734&737&748'),
    #     ('Bridal', '739&267&263'),
    #     ("Men's", '768&771'),
    #     ("Pearl",'740')
    # ]

    category = input_tuple
    logging.info(f"Category: {category}")

    
    dynamic_categories = []
    for category_name, code in category:
        try:
            match_key = f"{category_name.upper()}{code}".upper()
            filtered = df_outputs['report_grouping_df'].loc[
                df_outputs['report_grouping_df']["Cross ref"].str.upper() == match_key, "Count of PID"
            ]
            
            if not filtered.empty:
                # If multiple values returned (e.g., [187, NaN]), get the first valid number
                if isinstance(filtered.values[0], (np.ndarray, list, pd.Series)):
                    num_products = next((x for x in filtered.values[0] if pd.notna(x)), None)
                else:
                    num_products = filtered.values[0] if pd.notna(filtered.values[0]) else None
            else:
                num_products = None

            dynamic_categories.append((category_name, code, num_products))
        except Exception as e:
            print(f"Error processing {category_name} {code}: {e}")
            dynamic_categories.append((category_name, code, None))


    (current_month,current_month_number,rolling_method, previous_week_number, year_of_previous_month,last_year_of_previous_month, last_month_of_previous_month_numeric,season, feb_weeks, mar_weeks, apr_weeks, may_weeks,jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks,nov_weeks, dec_weeks, jan_weeks) = static_data = get_previous_retail_week(current_date)

    logging.info(f"Retail info: {year_of_previous_month}, {last_year_of_previous_month}, {season}, {current_month}, {current_month_number}, {previous_week_number}, {last_month_of_previous_month_numeric}, {rolling_method}, {feb_weeks}, {mar_weeks}, {apr_weeks}, {may_weeks}, {jun_weeks}, {jul_weeks}, {aug_weeks}, {sep_weeks}, {oct_weeks}, {nov_weeks}, {dec_weeks}, {jan_weeks}")

    # Create or update RetailInfo instance
    with transaction.atomic():
        RetailInfo.objects.create(
            sheet=sheet_object,  # Add this if you have a SheetUpload instance to link
            year_of_previous_month=year_of_previous_month,
            last_year_of_previous_month=last_year_of_previous_month,
            season=season,
            # current_date=current_date,  # Add this if you have the value
            current_month=current_month,
            current_month_number=current_month_number,
            previous_week_number=previous_week_number,
            last_month_of_previous_month_numeric=last_month_of_previous_month_numeric,

            feb_weeks=feb_weeks,
            mar_weeks=mar_weeks,
            apr_weeks=apr_weeks,
            may_weeks=may_weeks,
            jun_weeks=jun_weeks,
            jul_weeks=jul_weeks,
            aug_weeks=aug_weeks,
            sep_weeks=sep_weeks,
            oct_weeks=oct_weeks,
            nov_weeks=nov_weeks,
            dec_weeks=dec_weeks,
            jan_weeks=jan_weeks
        )


    logging.info("RetailInfo saved successfully.")

    args_list = [
        (sheets, return_qty_df, df_outputs, category, code, num_products, static_data, file_path, std_period, current_month_sales_percentage, current_date, sheet_object)
        for category, code, num_products in dynamic_categories
    ]

    store, coms, omni = [], [], []

    with Pool(processes=cpu_count()) as pool:
        results = pool.map(process_category, args_list)

    for result in results:
        if result:
            s, c, o = result
            store.extend(s)
            coms.extend(c)
            omni.extend(o)

    df_store = pd.DataFrame(store)
    df_coms = pd.DataFrame(coms)
    df_omni = pd.DataFrame(omni)

    def parse_selected_months(value):
        if isinstance(value, (list, tuple)):
            # Already a list or tuple, return as is
            return value
        elif isinstance(value, str):
            try:
                # Try to evaluate it as a string representation of a list
                return eval(value)
            except:
                # If eval fails, try json.loads
                
                try:
                    return json.loads(value)
                except:
                    # If all parsing fails, return an empty list as fallback
                    return []
        else:
            # Not a string or list, return empty list
            return []
        
    print("DataFrames created successfully.")
    print()
    # Create dictionaries for each record
    store_instances = [
    {
        'pid':row['pid'],        
        'loss': row['loss'],
        'new_month_12_fc_index': row['month_12_fc_index_(loss)'],
        'new_trend': row['trend'],
        'is_inventory_maintained': bool(row.get('Inventory maintained', False)),
        'trend_index_difference': row['trend index difference'],
        'average_com_oh': row['average com_oh'],
        'fldc': row['FLDC'],
        'forecasting_method': row['forecasting_method'],
    }
    for _, row in df_store.iterrows()
]
    # Similarly for ComForecast
    com_instances = [
    {
        'pid':row['pid'],  
         'new_month_12_fc_index': row['month_12_fc_index_original'],
         'is_inventory_maintained_com_sales': bool(row.get('Inventory maintained', False)),
         'forecasting_method': row['forecasting_method'],
         'minimum_required_oh_for_com': row['minimum required oh for com'],
         'fldc': row['FLDC'],
         'vdf_added_quantity': row['VDF_added_qty'],
         'month_12_fc_index_for_com_sales': row['com_month_12_fc_index']
        #   trend_of_total_sales
        #   trend_of_com_sales_for_selected_month

    }
    for _, row in df_coms.iterrows()
]

    # And for OmniForecast
    omni_instances = [
    {
            'pid':row['pid'],  
           'com_month_12_fc_index': row['Com month_12_fc_index'],
           'com_trend': row['com trend'],
           'is_com_inventory_maintained': bool(row.get('Com Inventory maintained', False)),
           'trend_index_difference_com': row['trend index difference(com)'],
           'forecasting_method_for_com': row['forecasting_method(com)'],
           'minimum_required_oh_for_com': row['minimum required oh for com'],
           'com_fldc': row['Com FLDC'],
           'store_month_12_fc_index': row['store_month_12_fc_index'],
           'loss': row['loss'],
           'store_month_12_fc_index_loss': row['store_month_12_fc_index_(loss)'],
           'store_trend': row['store_trend'],
           'trend_index_difference_store': row['trend index difference(store)'],
           'is_store_inventory_maintained': bool(row.get('store Inventory maintained', False)),
           'forecasting_method_for_store': row['forecasting_method(store)'],
           'store_fldc': row['store FLDC']


        }
    for _, row in df_omni.iterrows()
]


    # For StoreForecast

    print("Starting StoreForecast data save/update...")
    with transaction.atomic():
        for instance in store_instances:
            StoreForecast.objects.update_or_create(
                sheet=sheet_object,
                pid=instance['pid'],
                defaults={
                    'loss': instance['loss'],
                    'new_month_12_fc_index': instance['new_month_12_fc_index'],
                    'new_trend': instance['new_trend'],
                    'is_inventory_maintained': instance['is_inventory_maintained'],
                    'trend_index_difference': instance['trend_index_difference'],
                    'average_com_oh': instance['average_com_oh'],
                    'fldc': instance['fldc'],
                    'forecasting_method': instance['forecasting_method'],
                }
            )
    print("StoreForecast data saved/updated successfully.")


    # For ComForecast
    print("Starting ComForecast data save/update...")
    with transaction.atomic():
        for instance in com_instances:
            ComForecast.objects.update_or_create(
                sheet = sheet_object,
                pid=instance['pid'],
                defaults={
                    'new_month_12_fc_index': instance['new_month_12_fc_index'],
                    'is_inventory_maintained_com_sales': instance['is_inventory_maintained_com_sales'],
                    'forecasting_method': instance['forecasting_method'],
                    'minimum_required_oh_for_com': instance['minimum_required_oh_for_com'],
                    'fldc': instance['fldc'],
                    'vdf_added_quantity': instance['vdf_added_quantity'],
                    'month_12_fc_index_for_com_sales': instance['month_12_fc_index_for_com_sales'],
                }
            )
    print("ComForecast data saved/updated successfully.")


   # For OmniForecast - Updated with all the fields matching the model definition
    print("Starting OmniForecast data save/update...")
    with transaction.atomic():
        for instance in omni_instances:
            OmniForecast.objects.update_or_create(
                sheet=sheet_object,
                pid=instance['pid'],
                defaults={
                    'com_month_12_fc_index': instance['com_month_12_fc_index'],
                    'com_trend': instance['com_trend'],
                    'is_com_inventory_maintained': instance['is_com_inventory_maintained'],
                    'trend_index_difference_com': instance['trend_index_difference_com'],
                    'forecasting_method_for_com': instance['forecasting_method_for_com'],
                    'minimum_required_oh_for_com': instance['minimum_required_oh_for_com'],
                    'com_fldc': instance['com_fldc'],
                    'store_month_12_fc_index': instance['store_month_12_fc_index'],
                    'loss': instance['loss'],
                    'store_month_12_fc_index_loss': instance['store_month_12_fc_index_loss'],
                    'store_trend': instance['store_trend'],
                    'trend_index_difference_store': instance['trend_index_difference_store'],
                    'is_store_inventory_maintained': instance['is_store_inventory_maintained'],
                    'forecasting_method_for_store': instance['forecasting_method_for_store'],
                    'store_fldc': instance['store_fldc']
                }
            )
    print("OmniForecast data saved/updated successfully.")




    # Write to different sheets in one Excel file
    summary_filename = f"forecast_summary_{sheet_object.id}.xlsx"
    output_file = os.path.join(settings.MEDIA_ROOT, "summary", summary_filename)
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    df_store_renamed = df_store.rename(columns=STORE_RENAME_MAP)
    df_coms_renamed = df_coms.rename(columns=COM_RENAME_MAP)
    df_omni_renamed = df_omni.rename(columns=OMNI_RENAME_MAP)

    df_store_filtered = df_store_renamed[[col for col in STORE_RENAME_MAP.values()]]
    df_coms_filtered = df_coms_renamed[[col for col in COM_RENAME_MAP.values()]]
    df_omni_filtered = df_omni_renamed[[col for col in OMNI_RENAME_MAP.values()]]

    with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        for sheet_name, df in [
            ("store", df_store_filtered),
            ("coms", df_coms_filtered),
            ("omni", df_omni_filtered)
        ]:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, header=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Write header manually
            for col_num, column_title in enumerate(df.columns):
                worksheet.write(0, col_num, column_title)

            # Add Excel table
            worksheet.add_table(
                0, 0, len(df), len(df.columns) - 1,
                {
                    "name": f"{sheet_name}_table",
                    "columns": [{"header": col} for col in df.columns],
                    "style": "Table Style Medium 9"
                }
            )

    # Save the summary file path to the SheetUpload model
    relative_summary_path = f"summary/{summary_filename}"
    sheet_object.summary.name = relative_summary_path
    sheet_object.save()
    
    print("Data written to Excel file successfully.")    


def process_category(args):    

    sheets, return_qty_df, df_outputs, category, code, num_products, static_data, file_path, std_period, current_month_sales_percentage, current_date, sheet_object = args

    logging.info(f"[DEBUG] category: {category}, code: {code}, num_products: {num_products}")




    (current_month,current_month_number,rolling_method, previous_week_number, year_of_previous_month,last_year_of_previous_month, last_month_of_previous_month_numeric,season, feb_weeks, mar_weeks, apr_weeks, may_weeks,jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks,nov_weeks, dec_weeks, jan_weeks) = static_data
    
    
    H_VALUES = [
    'ROLLING 12M FC', 'Index', 'FC by Index', 'FC by Trend', 'Recommended FC', 'Planned FC',
    'Planned Shipments', 'Planned EOH (Cal)', 'Gross Projection (Nav)', 'Macys Proj Receipts',
    'Planned Sell thru %', f"TOTAL {year_of_previous_month}",'Total Sales Units', 'Store Sales Units', 'Com Sales Units',
    'COM % to TTL (Sales)', 'TOTAL EOM OH', 'Store EOM OH', 'COM EOM OH',
    'COM % to TTL (EOH)', 'Omni Sales $', 'COM Sales $', 'Omni AUR/% Diff Own',
    'Omni Sell Thru %', 'Store SellThru %', 'Omni Turn', 'Store turn',
    'TY Store Sales U vs LY', 'TY COM sales U vs LY', 'TY Store EOH vs LY',
    'Omni OO Units', 'COM OO Units', 'Omni Receipts',    f"TOTAL {last_year_of_previous_month}",
    "Total Sales Units",
    "Store Sales Units",
    "Com Sales Units",
    "COM % to TTL (Sales)",
    "TOTAL EOM OH",
    "Store EOM OH",
    "COM EOM OH",
    "COM % to TTL (EOH)",
    "Omni Receipts",
    "Omni Sell Thru %",
    "Store SellThru %",
    "Omni Turn",
    "Store Turn",
    "Omni Sales $",
    "COM Sales $",
    "Omni AUR/% Diff Own"
]
    store = []
    coms = []
    omni = []

    logging.info(f'num_products: {num_products}')

        # Update the data variable with category-specific values
    data = [
        ["", "TY", year_of_previous_month, "LY", last_year_of_previous_month, "Season", season, "Current Year", "Month", current_month, current_month_number, "Week", previous_week_number, "", "MAY-SEP", "", "", "Last Completed Month", last_month_of_previous_month_numeric, "", "Use EOM Actual?", rolling_method],
        ["", "Count of Items", "", 8215, "", "", "", "Last SP / FA Months", "Month", "Jul", "", "Jan", 12, "Sorted by:", "Dept Grouping >Class ID", "", "", ""],
        ["", "", "", "", "", "", "", "# of Wks in Mth", feb_weeks, mar_weeks, apr_weeks, may_weeks, jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks, nov_weeks, dec_weeks, jan_weeks],
        ["", category.upper(), code, "", "Avg Sales 1st & last Mth", 8, 11, "Month #", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, "", "", ""]
    ]

    output_file_name = f"{category}{code}.xlsx"
    # output_dir = r"D:\Yamini\Excel automation\original_excel_automation\Refactored_Samuel\April_week_4"
    output_dir = file_path
    output_file = os.path.join(output_dir, output_file_name)
    # Initialize workbook and create sheets
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category
    ws_index = wb.create_sheet(title="Index")
    ws_month = wb.create_sheet(title="Month")
    ws_dropdown = wb.create_sheet(title="DropdownData")


    # Loop through products to generate PIDs

    # Step 4: Populate Worksheet with Data
    for row_num, row_data in enumerate(data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Step 4: Freeze Top 4 Rows
    ws.freeze_panes = ws["A5"]
    # Define options for dropdowns
    print('yamini',num_products)

    
    ws['C2'] = num_products
    ws['D2'] = "=C2*51+4"
    ws['K1'] = "=VLOOKUP(J1,Month!A:B,2,0)"
    ws['K2'] = "=VLOOKUP(J2,Month!A:B,2,0)"
    ws['M2'] = "=VLOOKUP(L2,Month!A:B,2,0)"
    # Define additional dropdown options

    # Define your dropdown options (38 items)
    dropdown_options = [
        "BT", "Citrine", "Cross", "CZ", "Dia", "Ear", "EMER", "Garnet", "Gem",
        "GEM EAR", "Gold Chain", "GOLD EAR", "Amy", "Anklet", "Aqua", "Bridal",
        "Heart", "Heavy Gold Chain", "Jade", "KIDS", "Locket", "Mens Gold Bracelet",
        "Mens Misc", "Mens Silver chain", "Mom", "MOP", "Neck", "Onyx", "Opal",
        "Pearl", "Peridot", "Religious", "Ring", "Ruby", "Saph", "Womens Silver Chain",
        "Wrist", "Grand Total"
    ]

    # Write the dropdown options to the "DropdownData" sheet in column A
    for i, option in enumerate(dropdown_options, start=1):
        ws_dropdown.cell(row=i, column=1, value=option)

    # Define the named range for the dropdown options
    # Make sure the range covers exactly the 38 items (A1 to A38)
    named_range = DefinedName(name="DropdownOptions", attr_text="DropdownData!$A$1:$A$38")
    wb.defined_names.add(named_range)

    # Create a data validation that references the named range
    dropdown = DataValidation(type="list", formula1="DropdownOptions", allow_blank=True)
    dropdown.prompt = "Please select an option"
    dropdown.promptTitle = "Dropdown List"



    for loop in range(num_products):
        start_row = 5 + (51 * loop)  # Adjust the range as needed for your loop
        ws.add_data_validation(dropdown)
        dropdown.add(ws[f"F{start_row + 1}"])  # Apply to column F as an example

    forecast_method_options = ["FC by Index", "FC by Trend", "Average", "Current Year", "Last Year"]


    season_option = ["FA", "SP"]
    month_option = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
    year_option = ["Current MTH", "YTD", "SPRING", "FALL", "LY FALL"]


    # Function to add a dropdown to a specific cell
    def add_dropdown(ws, cell, options):
        # Create a data validation object with the list of options
        dropdown = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
        dropdown.prompt = "Please select an option"
        dropdown.promptTitle = "Dropdown List"
        # Apply the data validation dropdown to the specified cell
        ws.add_data_validation(dropdown)
        dropdown.add(ws[cell])
    # Add dropdowns to specified cells
    add_dropdown(ws, "G1", season_option)     # G1 for Season
    add_dropdown(ws, "V1", year_option)       # V1 for Year options

    # Add dropdowns to multiple cells for Months
    for cell in ["J1", "J2", "L2"]:
        add_dropdown(ws, cell, month_option)

    def apply_round_format(ws, cell_ranges, decimal_places):
        for cell_range in cell_ranges:
            # Check if the cell range is a single cell
            if ":" in cell_range:
                # This is a range, so we can iterate through it
                for row in ws[cell_range]:
                    for cell in row:
                        # Check if cell contains a formula by verifying if the value starts with "="
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            # Wrap formula in ROUND with the specified decimal places
                            cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"
            else:
                # This is a single cell reference
                cell = ws[cell_range]
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Wrap formula in ROUND with the specified decimal places
                    cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"

    def apply_format(ws, cell_ranges, number_format):
        for cell_range in cell_ranges:
            # Check if the cell range is a single cell or a range
            if ":" in cell_range:
                # This is a range, so we can iterate through it
                for row in ws[cell_range]:
                    for cell in row:
                        cell.number_format = number_format
            else:
                # This is a single cell
                cell = ws[cell_range]
                cell.number_format = number_format
    # Open the source workbook

    for col_num, col_name in enumerate(df_outputs['index_df'].columns, start=1):
        ws_index.cell(row=3, column=col_num, value=col_name)

    # Step 2: Write data starting from row 4
    for row_num, row_data in enumerate(df_outputs['index_df'].values, start=4):
        for col_num, value in enumerate(row_data, start=1):
            ws_index.cell(row=row_num, column=col_num, value=value)

    # Step 3: Apply Percentage Formatting to B4:P41 (i.e., Feb to Dec + Jan)
    for row in ws_index.iter_rows(min_row=4, max_row=41, min_col=2, max_col=16):
        for cell in row:
            if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                cell.number_format = '0.00%'  # Format as percentage

    # Step 6: Apply Filters to A3:P3
    ws_index.auto_filter.ref = "A3:P3"




    # Write the data to the "Month" sheet
    for row_num, (month, number) in enumerate(MONTH_DATA, start=1):
        ws_month.cell(row=row_num, column=1, value=month)  # Column A
        ws_month.cell(row=row_num, column=2, value=number)  # Column B

    border_color = "D9D9D9"  # Gridline color
    gridline1 = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )

    # Create PatternFill objects for the colors
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    light_yellow = PatternFill(start_color="FFEB9C" , end_color="FFEB9C" , fill_type="solid")
    light_pink = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Apply the fills to specific cells
    ws['C1'].fill = light_gray
    ws['E1'].fill = light_gray
    ws['G1'].fill = light_gray
    ws['J1'].fill = light_gray
    ws['K1'].fill = light_gray
    ws['M1'].fill = light_gray
    ws['S1'].fill = light_gray
    ws['J2'].fill = light_gray
    ws['L2'].fill = light_gray

    ws['K2'].fill = light_yellow
    ws['O1'].fill = light_yellow
    ws['M2'].fill = light_yellow

    ws['B2'].fill = yellow_fill
    ws['C2'].fill = yellow_fill
    ws['V1'].fill = light_pink

    #no boader
    # Define white border style
    white_border = Border(
        left=Side(style='thin', color="FFFFFF"),
        right=Side(style='thin', color="FFFFFF"),
        top=Side(style='thin', color="FFFFFF"),
        bottom=Side(style='thin', color="FFFFFF")
    )

    # Apply no border to the range B1:W2
    for row in range(1, 3):  # Rows 1 to 2
        for col in range(2, 24):  # Columns B (2) to W (23)
            ws.cell(row=row, column=col).border = white_border

    # Remove borders for specific cells to show gridlines
    # Define the border style
    border = Border(
        left=Side(style='thin', color='7F7F7F'),  # Thin black border on the left
        right=Side(style='thin', color='7F7F7F'),  # Thin black border on the right
        top=Side(style='thin', color='7F7F7F'),  # Thin black border on the top
        bottom=Side(style='thin', color='7F7F7F')  # Thin black border on the bottom
    )
    #no_boader list
    no_boader_list=['C1','E1','G1','J1','K1','L1','M1','O1','S1','V1','J2','K2','L2','M2']
    for i in no_boader_list:
        ws[i].border = border

    #font color
    red_font_italic= Font(color="FF0000", italic=True) 
    yellow_font = Font(color="9C5700") 
    blue_bold_font = Font(color="0563C1",bold=True) 
    ws['B4'].font = blue_bold_font
    ws['C4'].font = blue_bold_font
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['K2'].font = yellow_font
    ws['O1'].font = yellow_font
    ws['M2'].font = yellow_font

    ws['H1'].font = red_font_italic
    ws['H2'].font = red_font_italic
    ws['O2'].font = red_font_italic
    # Define the alignment (right-aligned)
    right_alignment = Alignment(horizontal='right')
    ws['H1'].alignment = right_alignment
    ws['H2'].alignment = right_alignment
    # Apply no border to the cells in the no_boader list
    bold_font_list=['C1','E1','G1','J1','K1','M1','S1','V1','L2','J2','C2','E4','F4','G4','H4',]
    for i in bold_font_list:
        ws[i].font = Font(bold=True)
    for row in ws["H3:W4"]:  # Use the dynamically calculated range
        for cell in row:
            cell.border = gridline1 

    ws.column_dimensions['c'].width = 20
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['H'].width = 25

    for loop in range(num_products):
        g_value = loop + 1
        cross_ref = f"{g_value}{category.upper()}{code}"  # Ensure no spaces in category
        matching_row = df_outputs['planning_df'].loc[df_outputs['planning_df']['Cross ref'].str.upper() == cross_ref]
        product_id = matching_row['PID'].iloc[0]
        macys_receipts_matching_row=df_outputs['macys_receipts'].loc[df_outputs['macys_receipts']['PID'].str.upper() == product_id]
        # Find the matching rows
        logging.info(f"product_id: {product_id}")
        loader = VariableLoader(cross_ref,matching_row,macys_receipts_matching_row,df_outputs['index_df'],df_outputs['all_data'],df_outputs['mcom_data'],std_period,year_of_previous_month,last_year_of_previous_month)
        logging.info(f"Processing PID: {product_id}")
        logging.info(f"Loader object created successfully.")
        logging.info(f"df_outputs['master_sheet']: {df_outputs['master_sheet']}")
        logging.info(f"df_outputs['vendor_sheet']: {df_outputs['vendor_sheet']}")
        logging.info(f"df_outputs['birthstone_sheet']: {df_outputs['birthstone_sheet']}")
        logging.info(f"df_outputs['return_qty_df']: {df_outputs['return_qty_df']}")
        vendor = df_outputs['master_sheet'].loc[df_outputs['master_sheet']['PID'] == product_id, 'Vendor Name'].values[0] if not df_outputs['master_sheet'].loc[df_outputs['master_sheet']['PID'] == product_id, 'Vendor Name'].empty else None
        logging.info(f"Vendor: {vendor}")
        return_qty_df_row = df_outputs['return_qty_df'][df_outputs['return_qty_df']['PID'] == product_id]
        logging.info(f"Return Quantity DataFrame row: {return_qty_df_row}")
        master_sheet_row = df_outputs['master_sheet'].loc[df_outputs['master_sheet']['PID'] == product_id]
        print("Master sheet row:", master_sheet_row)
        current_month,pid_type,std_trend,STD_index_value ,month_12_fc_index,forecasting_method,planned_shp,planned_fc,pid_omni_status,store,coms,omni,fc_by_index, fc_by_trend, recommended_fc, planned_oh, planned_sell_thru,total_added_quantity = algorithm(vendor,master_sheet_row, df_outputs['vendor_sheet'],df_outputs['birthstone_sheet'], return_qty_df_row, loader, category, store, coms, omni, code,current_month_sales_percentage,std_period, current_date,static_data )
        print("################################################################3",total_added_quantity)

        def safe_int(value):
            """Convert value to int or return None if conversion fails"""
            if pd.isna(value):
                return None
            try:
                return int(value)
            except (ValueError, TypeError):
                return None
                
        def safe_float(value):
            """Convert value to float or return None if conversion fails"""
            if pd.isna(value):
                return None
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
                
        def safe_str(value, max_length=None):
            """Convert value to string, respecting max_length if provided"""
            if pd.isna(value):
                return None
            try:
                result = str(value).strip()
                if max_length and len(result) > max_length:
                    result = result[:max_length]
                return result
            except:
                return None
        
        def parse_date(date_value):
            """
            Parses a date value (string or pandas Timestamp) into a datetime.date object.
            Handles multiple formats and NaT (Not a Time) values.
            """
            if pd.isna(date_value) or date_value in ["NaT", None, ""]:
                return None  # Handle missing or NaT values

            # If the value is already a pandas Timestamp, convert it to date
            if isinstance(date_value, pd.Timestamp):
                return date_value.date()

            # Ensure the value is a string for parsing
            date_str = str(date_value).strip()

            try:
                # Try to parse "M/D/YYYY" or "MM/DD/YYYY" format
                return datetime.strptime(date_str.split()[0], "%m/%d/%Y").date()
            except ValueError:
                try:
                    # Try to parse "M/D/YYYY H:MM:SS AM/PM" format
                    return datetime.strptime(date_str.split(' ', 1)[0], "%m/%d/%Y").date()
                except ValueError:
                    try:
                        # Try parsing "YYYY-MM-DD HH:MM:SS" format (e.g., "2022-02-26 00:00:00")
                        return datetime.strptime(date_str.split()[0], "%Y-%m-%d").date()
                    except ValueError:
                        # Try various other possible formats
                        date_formats = ["%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"]
                        for fmt in date_formats:
                            try:
                                return datetime.strptime(date_str.split()[0], fmt).date()
                            except ValueError:
                                continue
                        
                        # If all parsing attempts fail, print a warning and return None
                        print(f"Could not parse date: {date_value}")
                        return None
                    except Exception as e:
                        print(f"Error parsing date {date_value}: {e}")
                        return None

        months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        current_year = datetime.now().year 
        website_link = f"http://www.macys.com/shop/product/{loader.product_description}?ID={loader.marketing_id}"

        def extract_product_type(pid_type):
            match = re.match(r"(store|com|omni)", str(pid_type).lower())
            if match:
                return match.group(1)
            return None
        with transaction.atomic():

            ProductDetail.objects.update_or_create(
                sheet=sheet_object,
                product_id=loader.product_id,
                defaults={
                    "product_description": safe_str(loader.product_description),
                    "product_type": extract_product_type(pid_type),
                    "safe_non_safe": safe_str(loader.safe_non_safe),
                    "item_code": safe_str(loader.item_code),
                    "rlj": safe_str(loader.rlj),
                    "mkst": safe_str(loader.mkst),
                    "current_door_count": safe_int(loader.current_door_count),
                    "last_store_count": safe_int(loader.last_store_count),
                    "door_count_updated": parse_date(loader.door_count_updated),
                    "store_model": safe_int(loader.store_model),
                    "com_model": safe_int(loader.com_model),
                    "holiday_build_fc": safe_int(loader.holiday_build_fc),
                    "macys_onhand": safe_int(loader.macys_onhand),
                    "oo_units": safe_int(loader.oo_units),
                    "in_transit": safe_int(loader.nav_oo),
                    "month_to_date_shipment": safe_int(loader.month_to_date_shipment),
                    "last_week_shipment": safe_int(loader.last_week_shipment),
                    "planned_weeks_of_stock": safe_int(loader.planned_weeks_of_stock),
                    "weeks_of_projection": safe_int(loader.weeks_of_projection),
                    "last_4_weeks_shipment": safe_int(loader.last_4_weeks_shipment),
                    "vendor_name": safe_str(loader.vendor_name),
                    "min_order": safe_int(loader.min_order),
                    "rl_total": safe_int(loader.rl_total),
                    "net_projection": safe_int(loader.net_projection),
                    "unalloc_order": safe_int(loader.unalloc_order),
                    "ma_bin": safe_int(loader.ma_bin),
                    "fldc": safe_int(loader.fldc),
                    "wip_quantity": safe_int(loader.wip_quantity),
                    "md_status": safe_str(loader.md_status),
                    "replenishment_flag": safe_str(loader.replenishment_flag),
                    "mcom_replenishment": safe_str(loader.mcom_replenishment),
                    "pool_stock": safe_int(loader.pool_stock),
                    "first_receipt_date": parse_date(loader.first_receipt_date),
                    "last_receipt_date": parse_date(loader.last_receipt_date),
                    "item_age": safe_int(loader.item_age),
                    "first_live_date": parse_date(loader.first_live_date),
                    "this_year_last_cost": safe_float(loader.this_year_last_cost),
                    "macys_owned_retail": safe_float(loader.macys_owned_retail),
                    "awr_first_ticket_retail": safe_float(loader.awr_first_ticket_retail),
                    "metal_lock": safe_float(loader.metal_lock),
                    "mfg_policy": safe_str(loader.mfg_policy),
                    "kpi_data_updated": safe_str(loader.kpi_data_updated),
                    "kpi_door_count": safe_int(loader.kpi_door_count),
                    "out_of_stock_location": safe_int(loader.out_of_stock_location),
                    "suspended_location_count": safe_int(loader.suspended_location_count),
                    "live_site": safe_str(loader.live_site),
                    "masterstyle_description": safe_float(loader.masterstyle_description),
                    "masterstyle_id": safe_str(loader.masterstyle_id),
                    "department_id": safe_int(loader.department_id),
                    "department_description": safe_str(loader.department_description),
                    "subclass_id": safe_int(loader.subclass_id),
                    "subclass_description": safe_str(loader.subclass_description),
                    "webid_description": safe_str(loader.webid_description),
                    "v2c": safe_str(loader.v2c),
                    "marketing_id": safe_str(loader.marketing_id),
                    "std_store_return": safe_float(loader.std_store_return),
                    "last_projection_review_date": parse_date(loader.last_projection_review_date),
                    "macy_spring_projection_note": safe_str(loader.macy_spring_projection_note),
                    "planner_response": safe_str(loader.planner_response),
                    "website": website_link,

                    "rolling_method" : rolling_method,
                    "std_trend_original" : std_trend,
                    "std_index_value_original" : STD_index_value,
                    "current_fc_index": safe_str(loader.current_fc_index),                    
                    "month_12_fc_index_original" : month_12_fc_index,
                    "forecasting_method_original" : forecasting_method,

                    "algorithm_generated_final_quantity" : total_added_quantity if total_added_quantity > 0 else 0,
                    "category": f"{category}{code}",
                    "user_updated_final_quantity": total_added_quantity if total_added_quantity > 0 else 0,
                }
            )


        productmain = ProductDetail.objects.get(product_id=str(loader.product_id),sheet=sheet_object)

        
        save_macys_projection_receipts(productmain, loader.matched_row, current_year, sheet_object)
        print(f"Macys projection receipts saved for product {loader.product_id}")
        
        save_monthly_forecasts(productmain, sheet_object, current_year, months, loader.ty_total_sales_units, loader.ly_total_sales_units, loader.ly_total_eom_oh, loader.ty_total_eom_oh, loader.ty_omni_receipts, loader.ly_omni_receipts, loader.ty_com_sales_units, loader.ly_com_sales_units, loader.ty_com_eom_oh, loader.ly_com_eom_oh, loader.ty_omni_sales_usd, loader.ly_omni_sales_usd, loader.ty_com_sales_usd, loader.ly_com_sales_usd, loader.ty_omni_oo_units, loader.ty_com_oo_units, loader.ly_store_sales_units, loader.ly_store_eom_oh, loader.ly_com_to_ttl_sales_pct, loader.ly_com_to_ttl_eoh_pct, loader.ly_omni_sell_thru_pct, loader.ly_store_sell_thru_pct, loader.ly_omni_turn, loader.ly_store_turn, loader.ly_omni_aur_diff_own, loader.ty_store_sales_units, loader.ty_store_eom_oh, loader.ty_com_to_ttl_sales_pct, loader.ty_com_to_ttl_eoh_pct, loader.ty_omni_aur_diff_own, loader.ty_omni_sell_thru_pct, loader.ty_store_sell_thru_pct, loader.ty_omni_turn, loader.ty_store_turn, loader.ty_store_sales_vs_ly, loader.ty_com_sales_vs_ly, loader.ty_store_eoh_vs_ly)
        print(f"Monthly forecasts saved for product {loader.product_id}")
        forecast_data = { 
            'index':loader.index_value,
            'fc_by_index':fc_by_index,
            'fc_by_trend': fc_by_trend,
            'recommended_fc': recommended_fc,
            'planned_fc': planned_fc,
            'planned_shipments': planned_shp,
            'planned_eoh_cal': planned_oh,
            'gross_projection_nav' : loader.gross_projection_nav,
            'macys_proj_receipts':loader.macys_proj_receipts,
            'planned_sell_thru_pct': planned_sell_thru
        }
        save_rolling_forecasts(productmain, sheet_object, current_year, forecast_data)
        print(f"Product {loader.product_id} saved successfully")



























        dynamic_formulas = {
            f"G{start_row + 34}": g_value,
            f"F{start_row + 1}": f"=C{start_row + 1}",
            f"C{start_row}": loader.product_id,
            f"D{start_row}": loader.rlj,
            f"F{start_row}":loader.mkst,
            f"O1":std_period[0]+"-"+std_period[-1],
            f"C{start_row + 1}":loader.current_fc_index,
            f"C{start_row + 2}": sum(loader.std_ty_total_sales_units_list),
            f"D{start_row + 2}": sum(loader.std_ly_total_sales_units_list),
            f"C{start_row + 3}": f"=IFERROR(ROUND((C{start_row + 2}-D{start_row + 2})/D{start_row + 2},2),0)",
            f"F{start_row + 3}": std_trend,
            f"E{start_row + 3}": "Chg Trend",
            f"D{start_row + 1}": "Change Index",
            f"E{start_row + 2}": STD_index_value,
            f"F{start_row + 2}": month_12_fc_index,  # Added IFERROR to prevent divide-by-zero issues
            f"F{start_row + 4}": forecasting_method,
            f"C{start_row + 4}": loader.item_status,
            f"C{start_row + 5}": loader.current_door_count,
            f"D{start_row + 5}": loader.last_store_count,
            f"E{start_row + 5}": f"=ROUND(C{start_row + 5}-D{start_row + 5},2)",
            f"F{start_row + 5}": loader.door_count_updated,
            f"C{start_row + 6}": loader.store_model,
            f"D{start_row + 6}": loader.com_model,
            f"E{start_row + 6}": f"=ROUND(C{start_row + 6}+D{start_row + 6},2)",
            f"G{start_row + 7}": loader.holiday_build_fc,
            f"C{start_row + 8}": loader.macys_onhand,
            f"D{start_row + 8}": loader.oo_units,
            f"E{start_row + 8}": f"=IF(D{start_row + 8}-E{start_row + 14}<0,0,D{start_row + 8}-E{start_row + 14})",
            f"E{start_row + 14}": loader.nav_oo,
            f"F{start_row + 8}": loader.month_to_date_shipment,
            f"G{start_row + 8}":loader.last_week_shipment,
            f"C{start_row + 9}": loader.planned_weeks_of_stock,
            f"D{start_row + 9}":loader.weeks_of_projection,
            f"F{start_row + 9}": loader.last_4_weeks_shipment,
            f"C{start_row + 12}": "=0",
            f"C{start_row + 13}": loader.vendor_name,
            f"G{start_row + 13}": loader.min_order,
            f"C{start_row + 14}": loader.rl_total,
            f"D{start_row + 14}": loader.net_projection,
            f"F{start_row + 14}": loader.unalloc_order,
            f"G{start_row + 14}": f"=IF(C{start_row + 5}>G{start_row + 5},(C{start_row + 8}+D{start_row + 8})-C{start_row + 5},IF(OR(E{start_row + 6}>C{start_row + 5},E{start_row + 6}>G{start_row + 5}),(C{start_row + 8}+D{start_row + 8})-E{start_row + 6},(C{start_row + 8}+D{start_row + 8})-G{start_row + 5}))",
            f"C{start_row + 15}": loader.ma_bin,
            f"D{start_row + 15}": loader.fldc,
            f"E{start_row + 15}": loader.wip_quantity,
            f"C{start_row + 16}": f"=C{start_row + 14}+F{start_row + 14}-C{start_row + 15}-E{start_row + 15}",
            f"D{start_row + 16}": f"=IF(AND(F{start_row + 14}>=C{start_row + 16},C{start_row + 16}>0),\"Demand due to Unalloc Sales Orders\",IF(C{start_row + 16}<0,\"Excess OH Units\",IF(AND(C{start_row + 16}>0,F{start_row + 14}<C{start_row + 16}),\"Demand\",\"\")))",
            f"C{start_row + 17}": loader.md_status,
            f"D{start_row + 17}":loader.replenishment_flag,
            f"E{start_row + 17}": loader.mcom_replenishment,
            f"F{start_row + 17}": loader.pool_stock,
            f"C{start_row + 18}": loader.first_receipt_date,
            f"D{start_row + 18}": loader.last_receipt_date,
            f"E{start_row + 18}": loader.item_age,
            f"F{start_row + 18}": f"=IFERROR((NOW()-VALUE(C{start_row + 18}))/30,0)",
            f"C{start_row + 19}": loader.this_year_last_cost,
            f"D{start_row + 19}": loader.macys_owned_retail,
            f"E{start_row + 19}": loader.awr_first_ticket_retail,
            f"F{start_row + 19}": f"=(D{start_row + 19}-C{start_row + 19})/D{start_row + 19}",
            f"C{start_row + 20}": loader.metal_lock,
            f"D{start_row + 20}":loader.mfg_policy,
            f"C{start_row + 21}": loader.kpi_data_updated,
            f"C{start_row + 22}": loader.kpi_door_count,
            f"C{start_row + 23}": f"=C{start_row + 22}-C{start_row + 5}",
            f"C{start_row + 24}": loader.out_of_stock_location,
            f"C{start_row + 25}": loader.suspended_location_count,
            f"D{start_row + 21}": f"=SUBSTITUTE(D{start_row},\"/\",\"\")",
            f"B{start_row + 26}": f"=HYPERLINK(\"http://www.macys.com/shop/product/\"&C{start_row + 32}&\"?ID=\"&E{start_row + 31},\"Click to View online\")",
            f"C{start_row + 27}": loader.department_id,
            f"D{start_row + 27}": loader.department_description,
            f"C{start_row + 28}":loader.subclass_id,
            f"D{start_row + 28}": loader.subclass_description,
            f"C{start_row + 29}": loader.masterstyle_id,
            f"D{start_row + 29}":loader.masterstyle_description,
            f"C{start_row + 30}":loader.product_description,
            f"C{start_row + 31}": loader.first_live_date,
            f"D{start_row + 31}": f"{loader.live_site}/{loader.v2c}",
            f"E{start_row + 31}": loader.marketing_id,
            f"F{start_row + 31}": loader.std_store_return,
            f"C{start_row + 32}":loader.webid_description,
            f"C{start_row + 33}":loader.last_projection_review_date,
            f"D{start_row + 35}":loader.macy_spring_projection_note,
            f"B{start_row + 39}": "\"Past Review Comments\"",
            f"B{start_row + 40}": loader.planner_response,
            f"I{start_row + 1}": loader.index_value['FEB'],
            f"J{start_row + 1}": loader.index_value['MAR'],
            f"K{start_row + 1}": loader.index_value['APR'],
            f"L{start_row + 1}": loader.index_value['MAY'],
            f"M{start_row + 1}": loader.index_value['JUN'],
            f"N{start_row + 1}": loader.index_value['JUL'],
            f"O{start_row + 1}": loader.index_value['AUG'],
            f"P{start_row + 1}": loader.index_value['SEP'],
            f"Q{start_row + 1}": loader.index_value['OCT'],
            f"R{start_row + 1}": loader.index_value['NOV'],
            f"S{start_row + 1}": loader.index_value['DEC'],
            f"T{start_row + 1}": loader.index_value['JAN'],
            f"U{start_row + 1}": f"=SUM(I{start_row + 1}:T{start_row + 1})",
            f"V{start_row + 1}": f"=IFERROR(SUM(I{start_row + 1}:N{start_row + 1}),0)",
            f"W{start_row + 1}": f"=IFERROR(SUM(O{start_row + 1}:T{start_row + 1}),0)",
            f"I{start_row + 2}": f"=ROUND($F{start_row + 2}*I{start_row + 1},0)",
            f"J{start_row + 2}": f"=ROUND($F{start_row + 2}*J{start_row + 1},0)",
            f"K{start_row + 2}": f"=ROUND($F{start_row + 2}*K{start_row + 1},0)",
            f"L{start_row + 2}": f"=ROUND($F{start_row + 2}*L{start_row + 1},0)",
            f"M{start_row + 2}": f"=ROUND($F{start_row + 2}*M{start_row + 1},0)",
            f"N{start_row + 2}": f"=ROUND($F{start_row + 2}*N{start_row + 1},0)",
            f"O{start_row + 2}": f"=ROUND($F{start_row + 2}*O{start_row + 1},0)",
            f"P{start_row + 2}": f"=ROUND($F{start_row + 2}*P{start_row + 1},0)",
            f"Q{start_row + 2}": f"=ROUND($F{start_row + 2}*Q{start_row + 1},0)",
            f"R{start_row + 2}": f"=ROUND($F{start_row + 2}*R{start_row + 1},0)",
            f"S{start_row + 2}": f"=ROUND($F{start_row + 2}*S{start_row + 1},0)",
            f"T{start_row + 2}": f"=ROUND($F{start_row + 2}*T{start_row + 1},0)",
            f"U{start_row + 2}": f"=SUM(I{start_row + 2}:T{start_row + 2})",
            f"V{start_row + 2}": f"=IFERROR(SUM(I{start_row + 2}:N{start_row + 2}),0)",
            f"W{start_row + 2}": f"=IFERROR(SUM(O{start_row + 2}:T{start_row + 2}),0)",

            f"I{start_row + 11}": "FEB",
            f"J{start_row + 11}": "MAR",
            f"K{start_row + 11}": "APR",
            f"L{start_row + 11}": "MAY",
            f"M{start_row + 11}": "JUN",
            f"N{start_row + 11}": "JUL",
            f"O{start_row + 11}": "AUG",
            f"P{start_row + 11}": "SEP",
            f"Q{start_row + 11}": "OCT",
            f"R{start_row + 11}": "NOV",
            f"S{start_row + 11}": "DEC",
            f"T{start_row + 11}": "JAN",
            f"U{start_row + 11}": "ANNUAL",
            f"V{start_row + 11}": "SPRING",
            f"W{start_row + 11}": "FALL",

            f"I{start_row + 33}": "FEB",
            f"J{start_row + 33}": "MAR",
            f"K{start_row + 33}": "APR",
            f"L{start_row + 33}": "MAY",
            f"M{start_row + 33}": "JUN",
            f"N{start_row + 33}": "JUL",
            f"O{start_row + 33}": "AUG",
            f"P{start_row + 33}": "SEP",
            f"Q{start_row + 33}": "OCT",
            f"R{start_row + 33}": "NOV",
            f"S{start_row + 33}": "DEC",
            f"T{start_row + 33}": "JAN",
            f"U{start_row + 33}": "ANNUAL",
            f"V{start_row + 33}": "SPRING",
            f"W{start_row + 33}": "FALL",

            f"F{start_row + 7}": f"=C{start_row + 14}+E{start_row + 8}-E{start_row + 7}",
            f"F{start_row + 10}": f"=MAX(0,C{start_row + 24}-F{start_row + 9})",
            f"D{start_row + 12}": f"=C{start_row + 12}*C{start_row + 19}",
            f"I{start_row + 12}": loader.ty_total_sales_units['FEB'],
            f"J{start_row + 12}":loader.ty_total_sales_units['MAR'],
            f"K{start_row + 12}": loader.ty_total_sales_units['APR'],
            f"L{start_row + 12}":loader.ty_total_sales_units['MAY'],
            f"M{start_row + 12}":loader.ty_total_sales_units['JUN'],
            f"N{start_row + 12}": loader.ty_total_sales_units['JUL'],
            f"O{start_row + 12}":loader.ty_total_sales_units['AUG'],
            f"P{start_row + 12}": loader.ty_total_sales_units['SEP'],
            f"Q{start_row + 12}": loader.ty_total_sales_units['OCT'],
            f"R{start_row + 12}":loader.ty_total_sales_units['NOV'],
            f"S{start_row + 12}": loader.ty_total_sales_units['DEC'],
            f"T{start_row + 12}": loader.ty_total_sales_units['JAN'],
            f"U{start_row + 12}": f"=IFERROR(SUM(I{start_row + 12}:T{start_row + 12}),0)",
            f"V{start_row + 12}": f"=IFERROR(SUM(I{start_row + 12}:N{start_row + 12}),0)",
            f"W{start_row + 12}": f"=IFERROR(SUM(O{start_row + 12}:T{start_row + 12}),0)" ,
            f"I{start_row + 34}": loader.ly_total_sales_units['FEB'],
            f"J{start_row + 34}":loader.ly_total_sales_units['MAR'],
            f"K{start_row + 34}":loader.ly_total_sales_units['APR'],
            f"L{start_row + 34}":loader.ly_total_sales_units['MAY'],
            f"M{start_row + 34}":loader.ly_total_sales_units['JUN'],
            f"N{start_row + 34}":loader.ly_total_sales_units['JUL'],
            f"O{start_row + 34}":loader.ly_total_sales_units['AUG'],
            f"P{start_row + 34}":loader.ly_total_sales_units['SEP'],
            f"Q{start_row + 34}":loader.ly_total_sales_units['OCT'],
            f"R{start_row + 34}":loader.ly_total_sales_units['NOV'],
            f"S{start_row + 34}":loader.ly_total_sales_units['DEC'],
            f"T{start_row + 34}":loader.ly_total_sales_units['JAN'],
            f"U{start_row + 34}": f"=SUM(I{start_row + 34}:T{start_row + 34})",
            f"V{start_row + 34}": f"=SUM(I{start_row + 34}:N{start_row + 34})",
            f"W{start_row + 34}": f"=SUM(O{start_row + 34}:T{start_row + 34})",
            f"I{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(I{start_row + 12}+I{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,I$4<7),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,I$4<7),ROUND(I{start_row + 12}+I{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,I$4>6),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,I$4<7),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,I$4>6),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0)))))))",

            f"J{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(J{start_row + 12}+J{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,J$4<7),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,J$4<7),ROUND(J{start_row + 12}+J{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,J$4>6),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,J$4<7),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,J$4>6),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0)))))))",

            f"L{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(L{start_row + 12}+L{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,L$4<7),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,L$4<7),ROUND(L{start_row + 12}+L{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,L$4>6),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,L$4<7),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,L$4>6),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0)))))))",

            f"M{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(M{start_row + 12}+M{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,M$4<7),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,M$4<7),ROUND(M{start_row + 12}+M{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,M$4>6),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,M$4<7),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,M$4>6),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0)))))))",

            f"N{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(N{start_row + 12}+N{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,N$4<7),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,N$4<7),ROUND(N{start_row + 12}+N{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,N$4>6),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,N$4<7),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,N$4>6),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0)))))))",

            f"O{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(O{start_row + 12}+O{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,O$4<7),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,O$4<7),ROUND(O{start_row + 12}+O{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,O$4>6),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,O$4<7),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,O$4>6),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0)))))))",

            f"P{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(P{start_row + 12}+P{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,P$4<7),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,P$4<7),ROUND(P{start_row + 12}+P{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,P$4>6),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,P$4<7),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,P$4>6),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0)))))))",

            f"Q{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(Q{start_row + 12}+Q{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,Q$4<7),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,Q$4<7),ROUND(Q{start_row + 12}+Q{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,Q$4>6),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,Q$4<7),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,Q$4>6),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0)))))))",

            f"R{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(R{start_row + 12}+R{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,R$4<7),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,R$4<7),ROUND(R{start_row + 12}+R{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,R$4>6),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,R$4<7),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,R$4>6),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0)))))))",

            f"S{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(S{start_row + 12}+S{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,S$4<7),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,S$4<7),ROUND(S{start_row + 12}+S{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,S$4>6),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,S$4<7),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,S$4>6),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0)))))))",

            f"T{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(T{start_row + 12}+T{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,T$4<7),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,T$4<7),ROUND(T{start_row + 12}+T{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,T$4>6),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,T$4<7),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,T$4>6),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0)))))))",
            f"K{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(K{start_row + 12}+K{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,K$4<7),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,K$4<7),ROUND(K{start_row + 12}+K{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,K$4>6),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,K$4<7),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,K$4>6),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0)))))))",

            # Continue similarly for columns L through T
            f"U{start_row + 3}": f"=SUM(I{start_row + 3}:T{start_row + 3})",
            f"V{start_row + 3}": f"=IFERROR(SUM(I{start_row + 3}:N{start_row + 3}),0)",
            f"W{start_row + 3}": f"=IFERROR(SUM(O{start_row + 3}:T{start_row + 3}),0)",
            
            f"I{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},I{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},I{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",I{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",I{start_row + 34},"
                                f"ROUND(AVERAGE(I{start_row + 2}:I{start_row + 3}),0)))))",

            f"J{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},J{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},J{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",J{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",J{start_row + 34},"
                                f"ROUND(AVERAGE(J{start_row + 2}:J{start_row + 3}),0)))))",

            f"K{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},K{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},K{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",K{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",K{start_row + 34},"
                                f"ROUND(AVERAGE(K{start_row + 2}:K{start_row + 3}),0)))))",

            f"L{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},L{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},L{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",L{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",L{start_row + 34},"
                                f"ROUND(AVERAGE(L{start_row + 2}:L{start_row + 3}),0)))))",

            f"M{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},M{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},M{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",M{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",M{start_row + 34},"
                                f"ROUND(AVERAGE(M{start_row + 2}:M{start_row + 3}),0)))))",

            f"N{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},N{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},N{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",N{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",N{start_row + 34},"
                                f"ROUND(AVERAGE(N{start_row + 2}:N{start_row + 3}),0)))))",

            f"O{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},O{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},O{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",O{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",O{start_row + 34},"
                                f"ROUND(AVERAGE(O{start_row + 2}:O{start_row + 3}),0)))))",

            f"P{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},P{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},P{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",P{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",P{start_row + 34},"
                                f"ROUND(AVERAGE(P{start_row + 2}:P{start_row + 3}),0)))))",

            f"Q{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},Q{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},Q{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",Q{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",Q{start_row + 34},"
                                f"ROUND(AVERAGE(Q{start_row + 2}:Q{start_row + 3}),0)))))",

            f"R{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},R{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},R{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",R{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",R{start_row + 34},"
                                f"ROUND(AVERAGE(R{start_row + 2}:R{start_row + 3}),0)))))",

            f"S{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},S{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},S{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",S{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",S{start_row + 34},"
                                f"ROUND(AVERAGE(S{start_row + 2}:S{start_row + 3}),0)))))",

            f"T{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},T{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},T{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",T{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",T{start_row + 34},"
                                f"ROUND(AVERAGE(T{start_row + 2}:T{start_row + 3}),0)))))",

            f"U{start_row + 4}": f"=SUM(I{start_row + 4}:T{start_row + 4})",
            f"V{start_row + 4}": f"=IFERROR(SUM(I{start_row + 4}:N{start_row + 4}),0)",
            f"W{start_row + 4}": f"=IFERROR(SUM(O{start_row + 4}:T{start_row + 4}),0)",

            f"I{start_row + 38}": loader.ly_total_eom_oh['FEB'],
            f"J{start_row + 38}": loader.ly_total_eom_oh['MAR'],
            f"K{start_row + 38}": loader.ly_total_eom_oh['APR'],
            f"L{start_row + 38}":loader.ly_total_eom_oh['MAY'],
            f"M{start_row + 38}":loader.ly_total_eom_oh['JUN'],
            f"N{start_row + 38}": loader.ly_total_eom_oh['JUL'],
            f"O{start_row + 38}": loader.ly_total_eom_oh['AUG'],
            f"P{start_row + 38}":loader.ly_total_eom_oh['SEP'],
            f"Q{start_row + 38}":loader.ly_total_eom_oh['OCT'],
            f"R{start_row + 38}": loader.ly_total_eom_oh['NOV'],
            f"S{start_row + 38}":loader.ly_total_eom_oh['DEC'],
            f"T{start_row + 38}": loader.ly_total_eom_oh['JAN'],
            f"U{start_row + 38}": f"=AVERAGEIFS(I{start_row + 38}:T{start_row + 38}, $I$4:$T$4, \"<=\"&$T$4)",
            f"V{start_row + 38}": f"=AVERAGEIFS(I{start_row + 38}:N{start_row + 38}, $I$4:$N$4, \"<=\"&$N$4)",
            f"W{start_row + 38}": f"=AVERAGEIFS(O{start_row + 38}:T{start_row + 38}, $O$4:$T$4, \"<=\"&$T$4)",
            f"I{start_row + 5}": f"=IF(AND($V$1=\"YTD\",I$4<$K$1),I{start_row + 12},IF(AND($V$1=\"YTD\",I$4=$K$1),I{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",I$4=$K$1),I{start_row + 12},IF(AND($V$1=\"SPRING\",I$4<7),I{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",I$4>6),I{start_row + 12},IF(AND($V$1=\"LY FALL\",I$4>6),I{start_row + 34},I{start_row + 4}))))))",

            f"J{start_row + 5}": f"=IF(AND($V$1=\"YTD\",J$4<$K$1),J{start_row + 12},IF(AND($V$1=\"YTD\",J$4=$K$1),J{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",J$4=$K$1),J{start_row + 12},IF(AND($V$1=\"SPRING\",J$4<7),J{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",J$4>6),J{start_row + 12},IF(AND($V$1=\"LY FALL\",J$4>6),J{start_row + 34},J{start_row + 4}))))))",

            f"K{start_row + 5}": f"=IF(AND($V$1=\"YTD\",K$4<$K$1),K{start_row + 12},IF(AND($V$1=\"YTD\",K$4=$K$1),K{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",K$4=$K$1),K{start_row + 12},IF(AND($V$1=\"SPRING\",K$4<7),K{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",K$4>6),K{start_row + 12},IF(AND($V$1=\"LY FALL\",K$4>6),K{start_row + 34},K{start_row + 4}))))))",

            f"L{start_row + 5}": f"=IF(AND($V$1=\"YTD\",L$4<$K$1),L{start_row + 12},IF(AND($V$1=\"YTD\",L$4=$K$1),L{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",L$4=$K$1),L{start_row + 12},IF(AND($V$1=\"SPRING\",L$4<7),L{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",L$4>6),L{start_row + 12},IF(AND($V$1=\"LY FALL\",L$4>6),L{start_row + 34},L{start_row + 4}))))))",

            f"M{start_row + 5}": f"=IF(AND($V$1=\"YTD\",M$4<$K$1),M{start_row + 12},IF(AND($V$1=\"YTD\",M$4=$K$1),M{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",M$4=$K$1),M{start_row + 12},IF(AND($V$1=\"SPRING\",M$4<7),M{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",M$4>6),M{start_row + 12},IF(AND($V$1=\"LY FALL\",M$4>6),M{start_row + 34},M{start_row + 4}))))))",

            f"N{start_row + 5}": f"=IF(AND($V$1=\"YTD\",N$4<$K$1),N{start_row + 12},IF(AND($V$1=\"YTD\",N$4=$K$1),N{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",N$4=$K$1),N{start_row + 12},IF(AND($V$1=\"SPRING\",N$4<7),N{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",N$4>6),N{start_row + 12},IF(AND($V$1=\"LY FALL\",N$4>6),N{start_row + 34},N{start_row + 4}))))))",

            f"O{start_row + 5}": f"=IF(AND($V$1=\"YTD\",O$4<$K$1),O{start_row + 12},IF(AND($V$1=\"YTD\",O$4=$K$1),O{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",O$4=$K$1),O{start_row + 12},IF(AND($V$1=\"SPRING\",O$4<7),O{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",O$4>6),O{start_row + 12},IF(AND($V$1=\"LY FALL\",O$4>6),O{start_row + 34},O{start_row + 4}))))))",
            f"P{start_row + 5}": f"=IF(AND($V$1=\"YTD\",P$4<$K$1),P{start_row + 12},IF(AND($V$1=\"YTD\",P$4=$K$1),P{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",P$4=$K$1),P{start_row + 12},IF(AND($V$1=\"SPRING\",P$4<7),P{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",P$4>6),P{start_row + 12},IF(AND($V$1=\"LY FALL\",P$4>6),P{start_row + 34},P{start_row + 4}))))))",

            f"Q{start_row + 5}": f"=IF(AND($V$1=\"YTD\",Q$4<$K$1),Q{start_row + 12},IF(AND($V$1=\"YTD\",Q$4=$K$1),Q{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",Q$4=$K$1),Q{start_row + 12},IF(AND($V$1=\"SPRING\",Q$4<7),Q{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",Q$4>6),Q{start_row + 12},IF(AND($V$1=\"LY FALL\",Q$4>6),Q{start_row + 34},Q{start_row + 4}))))))",
            
            f"R{start_row + 5}": f"=IF(AND($V$1=\"YTD\",R$4<$K$1),R{start_row + 12},IF(AND($V$1=\"YTD\",R$4=$K$1),R{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",R$4=$K$1),R{start_row + 12},IF(AND($V$1=\"SPRING\",R$4<7),R{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",R$4>6),R{start_row + 12},IF(AND($V$1=\"LY FALL\",R$4>6),R{start_row + 34},R{start_row + 4}))))))",
            
            f"S{start_row + 5}": f"=IF(AND($V$1=\"YTD\",S$4<$K$1),S{start_row + 12},IF(AND($V$1=\"YTD\",S$4=$K$1),S{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",S$4=$K$1),S{start_row + 12},IF(AND($V$1=\"SPRING\",S$4<7),S{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",S$4>6),S{start_row + 12},IF(AND($V$1=\"LY FALL\",S$4>6),S{start_row + 34},S{start_row + 4}))))))",
            
            f"T{start_row + 5}": f"=IF(AND($V$1=\"YTD\",T$4<$K$1),T{start_row + 12},IF(AND($V$1=\"YTD\",T$4=$K$1),T{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",T$4=$K$1),T{start_row + 12},IF(AND($V$1=\"SPRING\",T$4<7),T{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",T$4>6),T{start_row + 12},IF(AND($V$1=\"LY FALL\",T$4>6),T{start_row + 34},T{start_row + 4}))))))",
            
            f"U{start_row + 5}": f"=SUM(I{start_row + 5}:T{start_row + 5})",
            
            f"V{start_row + 5}": f"=IFERROR(SUM(I{start_row + 5}:N{start_row + 5}),0)",
            
            f"W{start_row + 5}": f"=IFERROR(SUM(O{start_row + 5}:T{start_row + 5}),0)",

            f"U{start_row + 6}": f"=SUM(I{start_row + 6}:T{start_row + 6})",
            f"V{start_row + 6}": f"=IFERROR(SUM(I{start_row + 6}:N{start_row + 6}),0)",
            f"W{start_row + 6}": f"=IFERROR(SUM(O{start_row + 6}:T{start_row + 6}),0)",
            f"I{start_row + 16}": loader.ty_total_eom_oh['FEB'],
            f"J{start_row + 16}": loader.ty_total_eom_oh['MAR'],
            f"K{start_row + 16}":loader.ty_total_eom_oh['APR'],
            f"L{start_row + 16}": loader.ty_total_eom_oh['MAY'],
            f"M{start_row + 16}": loader.ty_total_eom_oh['JUN'],
            f"N{start_row + 16}":loader.ty_total_eom_oh['JUL'],
            f"O{start_row + 16}": loader.ty_total_eom_oh['AUG'],
            f"P{start_row + 16}":loader.ty_total_eom_oh['SEP'],
            f"Q{start_row + 16}":loader.ty_total_eom_oh['OCT'],
            f"R{start_row + 16}": loader.ty_total_eom_oh['NOV'],
            f"S{start_row + 16}": loader.ty_total_eom_oh['DEC'],
            f"T{start_row + 16}": loader.ty_total_eom_oh['JAN'],

            f"U{start_row + 16}": f"=AVERAGEIFS(I{start_row + 16}:T{start_row + 16}, $I$4:$T$4, \"<=\"&$K$1)",
            
            f"V{start_row + 16}": f"=AVERAGEIFS(I{start_row + 16}:N{start_row + 16}, $I$4:$N$4, \"<=\"&$K$2)",
            
            f"W{start_row + 16}": f"=AVERAGEIFS(O{start_row + 16}:T{start_row + 16}, $O$4:$T$4, \"<=\"&$M$2)",

            f"I{start_row + 32}":loader.ty_omni_receipts['FEB'],
            
            f"J{start_row + 32}": loader.ty_omni_receipts['MAR'],

            f"K{start_row + 32}":loader.ty_omni_receipts['APR'],
            f"L{start_row + 32}":loader.ty_omni_receipts['MAY'],
            f"M{start_row + 32}":loader.ty_omni_receipts['JUN'],
            f"N{start_row + 32}":loader.ty_omni_receipts['JUL'],
            f"O{start_row + 32}":loader.ty_omni_receipts['AUG'],
            f"P{start_row + 32}":loader.ty_omni_receipts['SEP'],

            f"Q{start_row + 32}":loader.ty_omni_receipts['OCT'],

            f"R{start_row + 32}":loader.ty_omni_receipts['NOV'],

            f"S{start_row + 32}":loader.ty_omni_receipts['DEC'],

            f"T{start_row + 32}":loader.ty_omni_receipts['JAN'],

            f"U{start_row + 32}": f"=IFERROR(SUM(I{start_row + 32}:T{start_row + 32}),0)",

            f"V{start_row + 32}": f"=IFERROR(SUM(I{start_row + 32}:N{start_row + 32}),0)",

            f"W{start_row + 32}": f"=IFERROR(SUM(O{start_row + 32}:T{start_row + 32}),0)",
            f"I{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=I$4,I$4=1),$T{start_row + 38}+I{start_row + 6}+I{start_row + 32}-I{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=I$4),I{start_row + 16}+I{start_row + 6}-(I{start_row + 5}-I{start_row + 12}),IF(AND($V$1=\"Current Mth\",I$4=1,$K$1>1),$T{start_row + 7}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"YTD\",I$4<$K$1),I{start_row + 16},IF(AND($V$1=\"Spring\",I$4<7),I{start_row + 16},IF(AND($V$1=\"Fall\",I$4>6,I$4<$K$1),I{start_row + 16},IF(AND($V$1=\"Fall\",I$4=1,$K$1>1),$T{start_row + 7}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"Fall\",I$4>6,I$4=$K$1),H{start_row + 16}+I{start_row + 32}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"LY Fall\",I$4>6),I{start_row + 38},IF(AND(I$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+I{start_row + 6}+I{start_row + 6}-I{start_row + 5},IF(AND(I$4=$K$1,I$4>1),H{start_row + 16}+I{start_row + 32}+I{start_row + 6}-I{start_row + 5},H{start_row + 7}+I{start_row + 6}-I{start_row + 5})))))))))))",

            f"J{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=J$4,J$4=1),$T{start_row + 38}+J{start_row + 6}+J{start_row + 32}-J{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=J$4),J{start_row + 16}+J{start_row + 6}-(J{start_row + 5}-J{start_row + 12}),IF(AND($V$1=\"Current Mth\",J$4=1,$K$1>1),$T{start_row + 7}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"YTD\",J$4<$K$1),J{start_row + 16},IF(AND($V$1=\"Spring\",J$4<7),J{start_row + 16},IF(AND($V$1=\"Fall\",J$4>6,J$4<$K$1),J{start_row + 16},IF(AND($V$1=\"Fall\",J$4=1,$K$1>1),$T{start_row + 7}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"Fall\",J$4>6,J$4=$K$1),I{start_row + 16}+J{start_row + 32}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"LY Fall\",J$4>6),J{start_row + 38},IF(AND(J$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+J{start_row + 6}+J{start_row + 6}-J{start_row + 5},IF(AND(J$4=$K$1,J$4>1),I{start_row + 16}+J{start_row + 32}+J{start_row + 6}-J{start_row + 5},I{start_row + 7}+J{start_row + 6}-J{start_row + 5})))))))))))",

            f"K{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=K$4,K$4=1),$T{start_row + 38}+K{start_row + 6}+K{start_row + 32}-K{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=K$4),K{start_row + 16}+K{start_row + 6}-(K{start_row + 5}-K{start_row + 12}),IF(AND($V$1=\"Current Mth\",K$4=1,$K$1>1),$T{start_row + 7}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"YTD\",K$4<$K$1),K{start_row + 16},IF(AND($V$1=\"Spring\",K$4<7),K{start_row + 16},IF(AND($V$1=\"Fall\",K$4>6,K$4<$K$1),K{start_row + 16},IF(AND($V$1=\"Fall\",K$4=1,$K$1>1),$T{start_row + 7}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"Fall\",K$4>6,K$4=$K$1),J{start_row + 16}+K{start_row + 32}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"LY Fall\",K$4>6),K{start_row + 38},IF(AND(K$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+K{start_row + 6}+K{start_row + 6}-K{start_row + 5},IF(AND(K$4=$K$1,K$4>1),J{start_row + 16}+K{start_row + 32}+K{start_row + 6}-K{start_row + 5},J{start_row + 7}+K{start_row + 6}-K{start_row + 5})))))))))))",

            f"L{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=L$4,L$4=1),$T{start_row + 38}+L{start_row + 6}+L{start_row + 32}-L{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=L$4),L{start_row + 16}+L{start_row + 6}-(L{start_row + 5}-L{start_row + 12}),IF(AND($V$1=\"Current Mth\",L$4=1,$K$1>1),$T{start_row + 7}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"YTD\",L$4<$K$1),L{start_row + 16},IF(AND($V$1=\"Spring\",L$4<7),L{start_row + 16},IF(AND($V$1=\"Fall\",L$4>6,L$4<$K$1),L{start_row + 16},IF(AND($V$1=\"Fall\",L$4=1,$K$1>1),$T{start_row + 7}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"Fall\",L$4>6,L$4=$K$1),K{start_row + 16}+L{start_row + 32}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"LY Fall\",L$4>6),L{start_row + 38},IF(AND(L$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+L{start_row + 6}+L{start_row + 6}-L{start_row + 5},IF(AND(L$4=$K$1,L$4>1),K{start_row + 16}+L{start_row + 32}+L{start_row + 6}-L{start_row + 5},K{start_row + 7}+L{start_row + 6}-L{start_row + 5})))))))))))",

            f"M{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=M$4,M$4=1),$T{start_row + 38}+M{start_row + 6}+M{start_row + 32}-M{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=M$4),M{start_row + 16}+M{start_row + 6}-(M{start_row + 5}-M{start_row + 12}),IF(AND($V$1=\"Current Mth\",M$4=1,$K$1>1),$T{start_row + 7}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"YTD\",M$4<$K$1),M{start_row + 16},IF(AND($V$1=\"Spring\",M$4<7),M{start_row + 16},IF(AND($V$1=\"Fall\",M$4>6,M$4<$K$1),M{start_row + 16},IF(AND($V$1=\"Fall\",M$4=1,$K$1>1),$T{start_row + 7}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"Fall\",M$4>6,M$4=$K$1),L{start_row + 16}+M{start_row + 32}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"LY Fall\",M$4>6),M{start_row + 38},IF(AND(M$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+M{start_row + 6}+M{start_row + 6}-M{start_row + 5},IF(AND(M$4=$K$1,M$4>1),L{start_row + 16}+M{start_row + 32}+M{start_row + 6}-M{start_row + 5},L{start_row + 7}+M{start_row + 6}-M{start_row + 5})))))))))))",

            f"O{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=O$4,O$4=1),$T{start_row + 38}+O{start_row + 6}+O{start_row + 32}-O{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=O$4),O{start_row + 16}+O{start_row + 6}-(O{start_row + 5}-O{start_row + 12}),IF(AND($V$1=\"Current Mth\",O$4=1,$K$1>1),$T{start_row + 7}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"YTD\",O$4<$K$1),O{start_row + 16},IF(AND($V$1=\"Spring\",O$4<7),O{start_row + 16},IF(AND($V$1=\"Fall\",O$4>6,O$4<$K$1),O{start_row + 16},IF(AND($V$1=\"Fall\",O$4=1,$K$1>1),$T{start_row + 7}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"Fall\",O$4>6,O$4=$K$1),N{start_row + 16}+O{start_row + 32}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"LY Fall\",O$4>6),O{start_row + 38},IF(AND(O$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+O{start_row + 6}+O{start_row + 6}-O{start_row + 5},IF(AND(O$4=$K$1,O$4>1),N{start_row + 16}+O{start_row + 32}+O{start_row + 6}-O{start_row + 5},N{start_row + 7}+O{start_row + 6}-O{start_row + 5})))))))))))",

            f"P{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=P$4,P$4=1),$T{start_row + 38}+P{start_row + 6}+P{start_row + 32}-P{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=P$4),P{start_row + 16}+P{start_row + 6}-(P{start_row + 5}-P{start_row + 12}),IF(AND($V$1=\"Current Mth\",P$4=1,$K$1>1),$T{start_row + 7}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"YTD\",P$4<$K$1),P{start_row + 16},IF(AND($V$1=\"Spring\",P$4<7),P{start_row + 16},IF(AND($V$1=\"Fall\",P$4>6,P$4<$K$1),P{start_row + 16},IF(AND($V$1=\"Fall\",P$4=1,$K$1>1),$T{start_row + 7}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"Fall\",P$4>6,P$4=$K$1),O{start_row + 16}+P{start_row + 32}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"LY Fall\",P$4>6),P{start_row + 38},IF(AND(P$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+P{start_row + 6}+P{start_row + 6}-P{start_row + 5},IF(AND(P$4=$K$1,P$4>1),O{start_row + 16}+P{start_row + 32}+P{start_row + 6}-P{start_row + 5},O{start_row + 7}+P{start_row + 6}-P{start_row + 5})))))))))))",

            f"Q{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=Q$4,Q$4=1),$T{start_row + 38}+Q{start_row + 6}+Q{start_row + 32}-Q{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=Q$4),Q{start_row + 16}+Q{start_row + 6}-(Q{start_row + 5}-Q{start_row + 12}),IF(AND($V$1=\"Current Mth\",Q$4=1,$K$1>1),$T{start_row + 7}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"YTD\",Q$4<$K$1),Q{start_row + 16},IF(AND($V$1=\"Spring\",Q$4<7),Q{start_row + 16},IF(AND($V$1=\"Fall\",Q$4>6,Q$4<$K$1),Q{start_row + 16},IF(AND($V$1=\"Fall\",Q$4=1,$K$1>1),$T{start_row + 7}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"Fall\",Q$4>6,Q$4=$K$1),P{start_row + 16}+Q{start_row + 32}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"LY Fall\",Q$4>6),Q{start_row + 38},IF(AND(Q$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+Q{start_row + 6}+Q{start_row + 6}-Q{start_row + 5},IF(AND(Q$4=$K$1,Q$4>1),P{start_row + 16}+Q{start_row + 32}+Q{start_row + 6}-Q{start_row + 5},P{start_row + 7}+Q{start_row + 6}-Q{start_row + 5})))))))))))",

            f"N{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=N$4,N$4=1),$T{start_row + 38}+N{start_row + 6}+N{start_row + 32}-N{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=N$4),N{start_row + 16}+N{start_row + 6}-(N{start_row + 5}-N{start_row + 12}),IF(AND($V$1=\"Current Mth\",N$4=1,$K$1>1),$T{start_row + 7}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"YTD\",N$4<$K$1),N{start_row + 16},IF(AND($V$1=\"Spring\",N$4<7),N{start_row + 16},IF(AND($V$1=\"Fall\",N$4>6,N$4<$K$1),N{start_row + 16},IF(AND($V$1=\"Fall\",N$4=1,$K$1>1),$T{start_row + 7}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"Fall\",N$4>6,N$4=$K$1),M{start_row + 16}+N{start_row + 32}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"LY Fall\",N$4>6),N{start_row + 38},IF(AND(N$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+N{start_row + 6}+N{start_row + 6}-N{start_row + 5},IF(AND(N$4=$K$1,N$4>1),M{start_row + 16}+N{start_row + 32}+N{start_row + 6}-N{start_row + 5},M{start_row + 7}+N{start_row + 6}-N{start_row + 5})))))))))))",

            f"R{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=R$4,R$4=1),$T{start_row + 38}+R{start_row + 6}+R{start_row + 32}-R{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=R$4),R{start_row + 16}+R{start_row + 6}-(R{start_row + 5}-R{start_row + 12}),IF(AND($V$1=\"Current Mth\",R$4=1,$K$1>1),$T{start_row + 7}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"YTD\",R$4<$K$1),R{start_row + 16},IF(AND($V$1=\"Spring\",R$4<7),R{start_row + 16},IF(AND($V$1=\"Fall\",R$4>6,R$4<$K$1),R{start_row + 16},IF(AND($V$1=\"Fall\",R$4=1,$K$1>1),$T{start_row + 7}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"Fall\",R$4>6,R$4=$K$1),Q{start_row + 16}+R{start_row + 32}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"LY Fall\",R$4>6),R{start_row + 38},IF(AND(R$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+R{start_row + 6}+R{start_row + 6}-R{start_row + 5},IF(AND(R$4=$K$1,R$4>1),Q{start_row + 16}+R{start_row + 32}+R{start_row + 6}-R{start_row + 5},Q{start_row + 7}+R{start_row + 6}-R{start_row + 5})))))))))))",

            f"S{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=S$4,S$4=1),$T{start_row + 38}+S{start_row + 6}+S{start_row + 32}-S{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=S$4),S{start_row + 16}+S{start_row + 6}-(S{start_row + 5}-S{start_row + 12}),IF(AND($V$1=\"Current Mth\",S$4=1,$K$1>1),$T{start_row + 7}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"YTD\",S$4<$K$1),S{start_row + 16},IF(AND($V$1=\"Spring\",S$4<7),S{start_row + 16},IF(AND($V$1=\"Fall\",S$4>6,S$4<$K$1),S{start_row + 16},IF(AND($V$1=\"Fall\",S$4=1,$K$1>1),$T{start_row + 7}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"Fall\",S$4>6,S$4=$K$1),R{start_row + 16}+S{start_row + 32}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"LY Fall\",S$4>6),S{start_row + 38},IF(AND(S$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+S{start_row + 6}+S{start_row + 6}-S{start_row + 5},IF(AND(S$4=$K$1,S$4>1),R{start_row + 16}+S{start_row + 32}+S{start_row + 6}-S{start_row + 5},R{start_row + 7}+S{start_row + 6}-S{start_row + 5})))))))))))",

            f"T{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=T$4,T$4=1),$T{start_row + 38}+T{start_row + 6}+T{start_row + 32}-T{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=T$4),T{start_row + 16}+T{start_row + 6}-(T{start_row + 5}-T{start_row + 12}),IF(AND($V$1=\"Current Mth\",T$4=1,$K$1>1),$T{start_row + 7}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"YTD\",T$4<$K$1),T{start_row + 16},IF(AND($V$1=\"Spring\",T$4<7),T{start_row + 16},IF(AND($V$1=\"Fall\",T$4>6,T$4<$K$1),T{start_row + 16},IF(AND($V$1=\"Fall\",T$4=1,$K$1>1),$T{start_row + 7}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"Fall\",T$4>6,T$4=$K$1),S{start_row + 16}+T{start_row + 32}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"LY Fall\",T$4>6),T{start_row + 38},IF(AND(T$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+T{start_row + 6}+T{start_row + 6}-T{start_row + 5},IF(AND(T$4=$K$1,T$4>1),S{start_row + 16}+T{start_row + 32}+T{start_row + 6}-T{start_row + 5},S{start_row + 7}+T{start_row + 6}-T{start_row + 5})))))))))))",

            f"U{start_row + 7}": f"=AVERAGEIFS(I{start_row + 7}:T{start_row + 7},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 7}": f"=AVERAGEIFS(I{start_row + 7}:N{start_row + 7},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 7}": f"=AVERAGEIFS(O{start_row + 7}:T{start_row + 7},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 8}": loader.nav_feb,
            f"J{start_row + 8}": loader.nav_mar,
            f"K{start_row + 8}": loader.nav_apr,
            f"L{start_row + 8}": loader.nav_may,
            f"M{start_row + 8}": loader.nav_jun,
            f"N{start_row + 8}": loader.nav_jul,
            f"O{start_row + 8}": loader.nav_aug,
            f"P{start_row + 8}": loader.nav_sep,
            f"Q{start_row + 8}": loader.nav_oct,
            f"R{start_row + 8}": loader.nav_nov,
            f"S{start_row + 8}": loader.nav_dec,
            f"T{start_row + 8}": loader.nav_jan,
            f"I{start_row + 6}": planned_shp["FEB"],
            f"J{start_row + 6}": planned_shp["MAR"],
            f"K{start_row + 6}": planned_shp["APR"],
            f"L{start_row + 6}": planned_shp["MAY"],
            f"M{start_row + 6}": planned_shp["JUN"],
            f"N{start_row + 6}": planned_shp["JUL"],
            f"O{start_row + 6}": planned_shp["AUG"],
            f"P{start_row + 6}":planned_shp["SEP"],
            f"Q{start_row + 6}":planned_shp["OCT"],
            f"R{start_row + 6}":planned_shp["NOV"],
            f"S{start_row + 6}": planned_shp["DEC"],
            f"T{start_row + 6}": planned_shp["JAN"],
            f"U{start_row + 8}": f"=SUM(I{start_row + 8}:T{start_row + 8})",
            f"V{start_row + 8}": f"=IFERROR(SUM(I{start_row + 8}:N{start_row + 8}),0)",
            f"W{start_row + 8}": f"=IFERROR(SUM(O{start_row + 8}:T{start_row + 8}),0)",

            f"I{start_row + 9}": loader.macys_proj_receipts_feb,
            f"J{start_row + 9}": loader.macys_proj_receipts_mar,
            f"K{start_row + 9}": loader.macys_proj_receipts_apr,
            f"L{start_row + 9}": loader.macys_proj_receipts_may,
            f"M{start_row + 9}": loader.macys_proj_receipts_jun,
            f"N{start_row + 9}": loader.macys_proj_receipts_jul,
            f"O{start_row + 9}": loader.macys_proj_receipts_aug,
            f"P{start_row + 9}": loader.macys_proj_receipts_sep,
            f"Q{start_row + 9}": loader.macys_proj_receipts_oct,
            f"R{start_row + 9}": loader.macys_proj_receipts_nov,
            f"S{start_row + 9}": loader.macys_proj_receipts_dec,
            f"T{start_row + 9}": loader.macys_proj_receipts_jan,
            f"U{start_row + 9}": f"=SUM(I{start_row + 9}:T{start_row + 9})",
            f"V{start_row + 9}": f"=IFERROR(SUM(I{start_row + 9}:N{start_row + 9}),0)",
            f"W{start_row + 9}": f"=IFERROR(SUM(O{start_row + 9}:T{start_row + 9}),0)",
            f"I{start_row + 10}": f"=I{start_row + 5}/(I{start_row + 5}+I{start_row + 7})",
            f"J{start_row + 10}": f"=J{start_row + 5}/(J{start_row + 5}+J{start_row + 7})",
            f"K{start_row + 10}": f"=K{start_row + 5}/(K{start_row + 5}+K{start_row + 7})",
            f"L{start_row + 10}": f"=L{start_row + 5}/(L{start_row + 5}+L{start_row + 7})",
            f"M{start_row + 10}": f"=M{start_row + 5}/(M{start_row + 5}+M{start_row + 7})",
            f"N{start_row + 10}": f"=N{start_row + 5}/(N{start_row + 5}+N{start_row + 7})",
            f"O{start_row + 10}": f"=O{start_row + 5}/(O{start_row + 5}+O{start_row + 7})",
            f"P{start_row + 10}": f"=P{start_row + 5}/(P{start_row + 5}+P{start_row + 7})",
            f"Q{start_row + 10}": f"=Q{start_row + 5}/(Q{start_row + 5}+Q{start_row + 7})",
            f"R{start_row + 10}": f"=R{start_row + 5}/(R{start_row + 5}+R{start_row + 7})",
            f"S{start_row + 10}": f"=S{start_row + 5}/(S{start_row + 5}+S{start_row + 7})",
            f"T{start_row + 10}": f"=T{start_row + 5}/(T{start_row + 5}+T{start_row + 7})",
            f"U{start_row + 10}": f"=U{start_row + 5}/(U{start_row + 5}+U{start_row + 7})",
            f"V{start_row + 10}": f"=V{start_row + 5}/(V{start_row + 5}+V{start_row + 7})",
            f"W{start_row + 10}": f"=W{start_row + 5}/(W{start_row + 5}+W{start_row + 7})",
            f"I{start_row + 14}": loader.ty_com_sales_units['FEB'],

            f"J{start_row + 14}":loader.ty_com_sales_units['MAR'],

            f"K{start_row + 14}":loader.ty_com_sales_units['APR'],

            f"L{start_row + 14}":loader.ty_com_sales_units['MAY'],

            f"M{start_row + 14}": loader.ty_com_sales_units['JUN'],

            f"N{start_row + 14}": loader.ty_com_sales_units['JUL'],

            f"O{start_row + 14}": loader.ty_com_sales_units['AUG'],

            f"P{start_row + 14}":loader.ty_com_sales_units['SEP'],

            f"Q{start_row + 14}": loader.ty_com_sales_units['OCT'],

            f"R{start_row + 14}":loader.ty_com_sales_units['NOV'],

            f"S{start_row + 14}": loader.ty_com_sales_units['DEC'],

            f"T{start_row + 14}":loader.ty_com_sales_units['JAN'],

            f"U{start_row + 14}": f"=IFERROR(SUM(I{start_row + 14}:T{start_row + 14}),0)",
            f"V{start_row + 14}": f"=IFERROR(SUM(I{start_row + 14}:N{start_row + 14}),0)",
            f"W{start_row + 14}": f"=IFERROR(SUM(O{start_row + 14}:T{start_row + 14}),0)",
            f"I{start_row + 13}": f"=I{start_row + 12}-I{start_row + 14}",
            f"J{start_row + 13}": f"=J{start_row + 12}-J{start_row + 14}",
            f"K{start_row + 13}": f"=K{start_row + 12}-K{start_row + 14}",
            f"L{start_row + 13}": f"=L{start_row + 12}-L{start_row + 14}",
            f"M{start_row + 13}": f"=M{start_row + 12}-M{start_row + 14}",
            f"N{start_row + 13}": f"=N{start_row + 12}-N{start_row + 14}",
            f"O{start_row + 13}": f"=O{start_row + 12}-O{start_row + 14}",
            f"P{start_row + 13}": f"=P{start_row + 12}-P{start_row + 14}",
            f"Q{start_row + 13}": f"=Q{start_row + 12}-Q{start_row + 14}",
            f"R{start_row + 13}": f"=R{start_row + 12}-R{start_row + 14}",
            f"S{start_row + 13}": f"=S{start_row + 12}-S{start_row + 14}",
            f"T{start_row + 13}": f"=T{start_row + 12}-T{start_row + 14}",

            f"U{start_row + 13}": f"=IFERROR(SUM(I{start_row + 13}:T{start_row + 13}),0)",
            f"V{start_row + 13}": f"=IFERROR(SUM(I{start_row + 13}:N{start_row + 13}),0)",
            f"W{start_row + 13}": f"=IFERROR(SUM(O{start_row + 13}:T{start_row + 13}),0)",

            f"I{start_row + 15}": f"=IFERROR(I{start_row + 14}/I{start_row + 12},0)",
            f"J{start_row + 15}": f"=IFERROR(J{start_row + 14}/J{start_row + 12},0)",
            f"K{start_row + 15}": f"=IFERROR(K{start_row + 14}/K{start_row + 12},0)",
            f"L{start_row + 15}": f"=IFERROR(L{start_row + 14}/L{start_row + 12},0)",
            f"M{start_row + 15}": f"=IFERROR(M{start_row + 14}/M{start_row + 12},0)",
            f"N{start_row + 15}": f"=IFERROR(N{start_row + 14}/N{start_row + 12},0)",
            f"O{start_row + 15}": f"=IFERROR(O{start_row + 14}/O{start_row + 12},0)",
            f"P{start_row + 15}": f"=IFERROR(P{start_row + 14}/P{start_row + 12},0)",
            f"Q{start_row + 15}": f"=IFERROR(Q{start_row + 14}/Q{start_row + 12},0)",
            f"R{start_row + 15}": f"=IFERROR(R{start_row + 14}/R{start_row + 12},0)",
            f"S{start_row + 15}": f"=IFERROR(S{start_row + 14}/S{start_row + 12},0)",
            f"T{start_row + 15}": f"=IFERROR(T{start_row + 14}/T{start_row + 12},0)",
            f"U{start_row + 15}": f"=IFERROR(U{start_row + 14}/U{start_row + 12},0)",
            f"V{start_row + 15}": f"=IFERROR(V{start_row + 14}/V{start_row + 12},0)",
            f"W{start_row + 15}": f"=IFERROR(W{start_row + 14}/W{start_row + 12},0)",
            f"I{start_row + 18}": loader.ty_com_eom_oh['FEB'],

            f"J{start_row + 18}": loader.ty_com_eom_oh['MAR'],

            f"K{start_row + 18}":loader.ty_com_eom_oh['APR'],

            f"L{start_row + 18}":loader.ty_com_eom_oh['MAY'],

            f"M{start_row + 18}": loader.ty_com_eom_oh['JUN'],

            f"N{start_row + 18}": loader.ty_com_eom_oh['JUL'],

            f"O{start_row + 18}":loader.ty_com_eom_oh['AUG'],

            f"P{start_row + 18}":loader.ty_com_eom_oh['SEP'],

            f"Q{start_row + 18}":loader.ty_com_eom_oh['OCT'],

            f"R{start_row + 18}": loader.ty_com_eom_oh['NOV'],

            f"S{start_row + 18}": loader.ty_com_eom_oh['DEC'],

            f"T{start_row + 18}":loader.ty_com_eom_oh['JAN'],

            f"U{start_row + 18}": f"=AVERAGEIFS(I{start_row + 18}:T{start_row + 18},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 18}": f"=AVERAGEIFS(I{start_row + 18}:N{start_row + 18},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 18}": f"=AVERAGEIFS(O{start_row + 18}:T{start_row + 18},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 17}": f"=I{start_row + 16}-I{start_row + 18}",
            f"J{start_row + 17}": f"=J{start_row + 16}-J{start_row + 18}",
            f"K{start_row + 17}": f"=K{start_row + 16}-K{start_row + 18}",
            f"L{start_row + 17}": f"=L{start_row + 16}-L{start_row + 18}",
            f"M{start_row + 17}": f"=M{start_row + 16}-M{start_row + 18}",
            f"N{start_row + 17}": f"=N{start_row + 16}-N{start_row + 18}",
            f"O{start_row + 17}": f"=O{start_row + 16}-O{start_row + 18}",
            f"P{start_row + 17}": f"=P{start_row + 16}-P{start_row + 18}",
            f"Q{start_row + 17}": f"=Q{start_row + 16}-Q{start_row + 18}",
            f"R{start_row + 17}": f"=R{start_row + 16}-R{start_row + 18}",
            f"S{start_row + 17}": f"=S{start_row + 16}-S{start_row + 18}",
            f"T{start_row + 17}": f"=T{start_row + 16}-T{start_row + 18}",

            f"U{start_row + 17}": f"=AVERAGEIFS(I{start_row + 17}:T{start_row + 17},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 17}": f"=AVERAGEIFS(I{start_row + 17}:N{start_row + 17},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 17}": f"=AVERAGEIFS(O{start_row + 17}:T{start_row + 17},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 19}": f"=IFERROR(I{start_row + 18}/I{start_row + 16},0)",
            f"J{start_row + 19}": f"=IFERROR(J{start_row + 18}/J{start_row + 16},0)",
            f"K{start_row + 19}": f"=IFERROR(K{start_row + 18}/K{start_row + 16},0)",
            f"L{start_row + 19}": f"=IFERROR(L{start_row + 18}/L{start_row + 16},0)",
            f"M{start_row + 19}": f"=IFERROR(M{start_row + 18}/M{start_row + 16},0)",
            f"N{start_row + 19}": f"=IFERROR(N{start_row + 18}/N{start_row + 16},0)",
            f"O{start_row + 19}": f"=IFERROR(O{start_row + 18}/O{start_row + 16},0)",
            f"P{start_row + 19}": f"=IFERROR(P{start_row + 18}/P{start_row + 16},0)",
            f"Q{start_row + 19}": f"=IFERROR(Q{start_row + 18}/Q{start_row + 16},0)",
            f"R{start_row + 19}": f"=IFERROR(R{start_row + 18}/R{start_row + 16},0)",
            f"S{start_row + 19}": f"=IFERROR(S{start_row + 18}/S{start_row + 16},0)",
            f"T{start_row + 19}": f"=IFERROR(T{start_row + 18}/T{start_row + 16},0)",
            f"U{start_row + 19}": f"=IFERROR(U{start_row + 18}/U{start_row + 16},0)",
            f"V{start_row + 19}": f"=IFERROR(V{start_row + 18}/V{start_row + 16},0)",
            f"W{start_row + 19}": f"=IFERROR(W{start_row + 18}/W{start_row + 16},0)",
            f"I{start_row + 20}":loader.ty_omni_sales_usd['FEB'] ,

            f"J{start_row + 20}": loader.ty_omni_sales_usd['MAR'],

            f"K{start_row + 20}": loader.ty_omni_sales_usd['APR'],

            f"L{start_row + 20}":loader.ty_omni_sales_usd['MAY'],

            f"M{start_row + 20}": loader.ty_omni_sales_usd['JUN'],

            f"N{start_row + 20}": loader.ty_omni_sales_usd['JUL'],

            f"O{start_row + 20}":loader.ty_omni_sales_usd['AUG'],

            f"P{start_row + 20}": loader.ty_omni_sales_usd['SEP'],

            f"Q{start_row + 20}": loader.ty_omni_sales_usd['OCT'],

            f"R{start_row + 20}": loader.ty_omni_sales_usd['NOV'],

            f"S{start_row + 20}":loader.ty_omni_sales_usd['DEC'],

            f"T{start_row + 20}": loader.ty_omni_sales_usd['JAN'],

            # Aggregates
            f"U{start_row + 20}": f"=IFERROR(SUM(I{start_row + 20}:T{start_row + 20}),0)",
            f"V{start_row + 20}": f"=IFERROR(SUM(I{start_row + 20}:N{start_row + 20}),0)",
            f"W{start_row + 20}": f"=IFERROR(SUM(O{start_row + 20}:T{start_row + 20}),0)",
            f"I{start_row + 21}": loader.ty_com_sales_usd['FEB'],

            f"J{start_row + 21}": loader.ty_com_sales_usd['MAR'],

            f"K{start_row + 21}": loader.ty_com_sales_usd['APR'],

            f"L{start_row + 21}":loader.ty_com_sales_usd['MAY'],

            f"M{start_row + 21}":loader.ty_com_sales_usd['JUN'],

            f"N{start_row + 21}": loader.ty_com_sales_usd['JUL'],

            f"O{start_row + 21}":loader.ty_com_sales_usd['AUG'],

            f"P{start_row + 21}":loader.ty_com_sales_usd['SEP'],

            f"Q{start_row + 21}": loader.ty_com_sales_usd['OCT'],

            f"R{start_row + 21}": loader.ty_com_sales_usd['NOV'],

            f"S{start_row + 21}": loader.ty_com_sales_usd['DEC'],

            f"T{start_row + 21}":loader.ty_com_sales_usd['JAN'],

            f"U{start_row + 21}": f"=IFERROR(SUM(I{start_row + 21}:T{start_row + 21}),0)",
            f"V{start_row + 21}": f"=IFERROR(SUM(I{start_row + 21}:N{start_row + 21}),0)",
            f"W{start_row + 21}": f"=IFERROR(SUM(O{start_row + 21}:T{start_row + 21}),0)",

            f"I{start_row + 22}": f"=IFERROR(TEXT(IFERROR(I{start_row + 20}/I{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((I{start_row + 20}/I{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"J{start_row + 22}": f"=IFERROR(TEXT(IFERROR(J{start_row + 20}/J{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((J{start_row + 20}/J{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"K{start_row + 22}": f"=IFERROR(TEXT(IFERROR(K{start_row + 20}/K{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((K{start_row + 20}/K{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"L{start_row + 22}": f"=IFERROR(TEXT(IFERROR(L{start_row + 20}/L{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((L{start_row + 20}/L{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"M{start_row + 22}": f"=IFERROR(TEXT(IFERROR(M{start_row + 20}/M{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((M{start_row + 20}/M{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"N{start_row + 22}": f"=IFERROR(TEXT(IFERROR(N{start_row + 20}/N{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((N{start_row + 20}/N{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"O{start_row + 22}": f"=IFERROR(TEXT(IFERROR(O{start_row + 20}/O{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((O{start_row + 20}/O{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"P{start_row + 22}": f"=IFERROR(TEXT(IFERROR(P{start_row + 20}/P{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((P{start_row + 20}/P{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"Q{start_row + 22}": f"=IFERROR(TEXT(IFERROR(Q{start_row + 20}/Q{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((Q{start_row + 20}/Q{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"R{start_row + 22}": f"=IFERROR(TEXT(IFERROR(R{start_row + 20}/R{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((R{start_row + 20}/R{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"S{start_row + 22}": f"=IFERROR(TEXT(IFERROR(S{start_row + 20}/S{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((S{start_row + 20}/S{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"T{start_row + 22}": f"=IFERROR(TEXT(IFERROR(T{start_row + 20}/T{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((T{start_row + 20}/T{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"U{start_row + 22}": f"=IFERROR(TEXT(IFERROR(U{start_row + 20}/U{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((U{start_row + 20}/U{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"V{start_row + 22}": f"=IFERROR(TEXT(IFERROR(V{start_row + 20}/V{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((V{start_row + 20}/V{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"W{start_row + 22}": f"=IFERROR(TEXT(IFERROR(W{start_row + 20}/W{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((W{start_row + 20}/W{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"I{start_row + 23}": f"=IFERROR(IFERROR(I{start_row + 12}/(I{start_row + 12}+I{start_row + 16}),0),0)",
            f"J{start_row + 23}": f"=IFERROR(IFERROR(J{start_row + 12}/(J{start_row + 12}+J{start_row + 16}),0),0)",
            f"K{start_row + 23}": f"=IFERROR(IFERROR(K{start_row + 12}/(K{start_row + 12}+K{start_row + 16}),0),0)",
            f"L{start_row + 23}": f"=IFERROR(IFERROR(L{start_row + 12}/(L{start_row + 12}+L{start_row + 16}),0),0)",
            f"M{start_row + 23}": f"=IFERROR(IFERROR(M{start_row + 12}/(M{start_row + 12}+M{start_row + 16}),0),0)",
            f"N{start_row + 23}": f"=IFERROR(IFERROR(N{start_row + 12}/(N{start_row + 12}+N{start_row + 16}),0),0)",
            f"O{start_row + 23}": f"=IFERROR(IFERROR(O{start_row + 12}/(O{start_row + 12}+O{start_row + 16}),0),0)",
            f"P{start_row + 23}": f"=IFERROR(IFERROR(P{start_row + 12}/(P{start_row + 12}+P{start_row + 16}),0),0)",
            f"Q{start_row + 23}": f"=IFERROR(IFERROR(Q{start_row + 12}/(Q{start_row + 12}+Q{start_row + 16}),0),0)",
            f"R{start_row + 23}": f"=IFERROR(IFERROR(R{start_row + 12}/(R{start_row + 12}+R{start_row + 16}),0),0)",
            f"S{start_row + 23}": f"=IFERROR(IFERROR(S{start_row + 12}/(S{start_row + 12}+S{start_row + 16}),0),0)",
            f"T{start_row + 23}": f"=IFERROR(IFERROR(T{start_row + 12}/(T{start_row + 12}+T{start_row + 16}),0),0)",

            f"U{start_row + 23}": f"=IFERROR(IFERROR(U{start_row + 12}/(U{start_row + 12}+U{start_row + 16}-U{start_row + 8}),0),0)",
            f"V{start_row + 23}": f"=IFERROR(IFERROR(V{start_row + 12}/(V{start_row + 12}+V{start_row + 16}-V{start_row + 8}),0),0)",
            f"W{start_row + 23}": f"=IFERROR(IFERROR(W{start_row + 12}/(W{start_row + 12}+W{start_row + 16}-W{start_row + 8}),0),0)",
            f"I{start_row + 24}": f"=IFERROR((I{start_row + 12}-I{start_row + 14})/((I{start_row + 12}-I{start_row + 14})+(I{start_row + 16}-I{start_row + 18})),0)",
            f"J{start_row + 24}": f"=IFERROR((J{start_row + 12}-J{start_row + 14})/((J{start_row + 12}-J{start_row + 14})+(J{start_row + 16}-J{start_row + 18})),0)",
            f"K{start_row + 24}": f"=IFERROR((K{start_row + 12}-K{start_row + 14})/((K{start_row + 12}-K{start_row + 14})+(K{start_row + 16}-K{start_row + 18})),0)",
            f"L{start_row + 24}": f"=IFERROR((L{start_row + 12}-L{start_row + 14})/((L{start_row + 12}-L{start_row + 14})+(L{start_row + 16}-L{start_row + 18})),0)",
            f"M{start_row + 24}": f"=IFERROR((M{start_row + 12}-M{start_row + 14})/((M{start_row + 12}-M{start_row + 14})+(M{start_row + 16}-M{start_row + 18})),0)",
            f"N{start_row + 24}": f"=IFERROR((N{start_row + 12}-N{start_row + 14})/((N{start_row + 12}-N{start_row + 14})+(N{start_row + 16}-N{start_row + 18})),0)",
            f"O{start_row + 24}": f"=IFERROR((O{start_row + 12}-O{start_row + 14})/((O{start_row + 12}-O{start_row + 14})+(O{start_row + 16}-O{start_row + 18})),0)",
            f"P{start_row + 24}": f"=IFERROR((P{start_row + 12}-P{start_row + 14})/((P{start_row + 12}-P{start_row + 14})+(P{start_row + 16}-P{start_row + 18})),0)",
            f"Q{start_row + 24}": f"=IFERROR((Q{start_row + 12}-Q{start_row + 14})/((Q{start_row + 12}-Q{start_row + 14})+(Q{start_row + 16}-Q{start_row + 18})),0)",
            f"R{start_row + 24}": f"=IFERROR((R{start_row + 12}-R{start_row + 14})/((R{start_row + 12}-R{start_row + 14})+(R{start_row + 16}-R{start_row + 18})),0)",
            f"S{start_row + 24}": f"=IFERROR((S{start_row + 12}-S{start_row + 14})/((S{start_row + 12}-S{start_row + 14})+(S{start_row + 16}-S{start_row + 18})),0)",
            f"T{start_row + 24}": f"=IFERROR((T{start_row + 12}-T{start_row + 14})/((T{start_row + 12}-T{start_row + 14})+(T{start_row + 16}-T{start_row + 18})),0)",

            f"U{start_row + 24}": f"=IFERROR((U{start_row + 12}-U{start_row + 14})/((U{start_row + 12}-U{start_row + 14})+(U{start_row + 16}-U{start_row + 18}-U{start_row + 8})),0)",
            f"V{start_row + 24}": f"=IFERROR((V{start_row + 12}-V{start_row + 14})/((V{start_row + 12}-V{start_row + 14})+(V{start_row + 16}-V{start_row + 18}-V{start_row + 8})),0)",
            f"W{start_row + 24}": f"=IFERROR((W{start_row + 12}-W{start_row + 14})/((W{start_row + 12}-W{start_row + 14})+(W{start_row + 16}-W{start_row + 18})),0)",
            f"I{start_row + 25}": f"=IFERROR(IFERROR(I{start_row + 12}/I{start_row + 16},0),0)",
            f"J{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:J{start_row + 12})/AVERAGE(I{start_row + 16}:J{start_row + 16}),0),0)",
            f"K{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:K{start_row + 12})/AVERAGE(I{start_row + 16}:K{start_row + 16}),0),0)",
            f"L{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:L{start_row + 12})/AVERAGE(I{start_row + 16}:L{start_row + 16}),0),0)",
            f"M{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:M{start_row + 12})/AVERAGE(I{start_row + 16}:M{start_row + 16}),0),0)",
            f"N{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:N{start_row + 12})/AVERAGE(I{start_row + 16}:N{start_row + 16}),0),0)",
            f"O{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:O{start_row + 12})/AVERAGE(I{start_row + 16}:O{start_row + 16}),0),0)",
            f"P{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:P{start_row + 12})/AVERAGE(I{start_row + 16}:P{start_row + 16}),0),0)",
            f"Q{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:Q{start_row + 12})/AVERAGE(I{start_row + 16}:Q{start_row + 16}),0),0)",
            f"R{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:R{start_row + 12})/AVERAGE(I{start_row + 16}:R{start_row + 16}),0),0)",
            f"S{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:S{start_row + 12})/AVERAGE(I{start_row + 16}:S{start_row + 16}),0),0)",
            f"T{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:T{start_row + 12})/AVERAGE(I{start_row + 16}:T{start_row + 16}),0),0)",
            f"U{start_row + 25}": f"=IFERROR(U{start_row + 12}/U{start_row + 16},0)",
            f"V{start_row + 25}": f"=IFERROR(V{start_row + 12}/V{start_row + 16},0)",
            f"W{start_row + 25}": f"=IFERROR(W{start_row + 12}/W{start_row + 16},0)",
            f"I{start_row + 26}": f"=IFERROR(I{start_row + 13}/I{start_row + 17},0)",
            f"J{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:J{start_row + 13})/AVERAGE(I{start_row + 17}:J{start_row + 17}),0)",
            f"K{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:K{start_row + 13})/AVERAGE(I{start_row + 17}:K{start_row + 17}),0)",
            f"L{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:L{start_row + 13})/AVERAGE(I{start_row + 17}:L{start_row + 17}),0)",
            f"M{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:M{start_row + 13})/AVERAGE(I{start_row + 17}:M{start_row + 17}),0)",
            f"N{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:N{start_row + 13})/AVERAGE(I{start_row + 17}:N{start_row + 17}),0)",
            f"O{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:O{start_row + 13})/AVERAGE(I{start_row + 17}:O{start_row + 17}),0)",
            f"P{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:P{start_row + 13})/AVERAGE(I{start_row + 17}:P{start_row + 17}),0)",
            f"Q{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:Q{start_row + 13})/AVERAGE(I{start_row + 17}:Q{start_row + 17}),0)",
            f"R{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:R{start_row + 13})/AVERAGE(I{start_row + 17}:R{start_row + 17}),0)",
            f"S{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:S{start_row + 13})/AVERAGE(I{start_row + 17}:S{start_row + 17}),0)",
            f"T{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:T{start_row + 13})/AVERAGE(I{start_row + 17}:T{start_row + 17}),0)",
            f"U{start_row + 26}": f"=IFERROR(U{start_row + 13}/U{start_row + 17},0)",
            f"V{start_row + 26}": f"=IFERROR(V{start_row + 13}/V{start_row + 17},0)",
            f"W{start_row + 26}": f"=IFERROR(W{start_row + 13}/W{start_row + 17},0)",
            f"I{start_row + 36}": loader.ly_com_sales_units['FEB'],
            f"J{start_row + 36}":loader.ly_com_sales_units['MAR'] ,
            f"K{start_row + 36}":loader.ly_com_sales_units['APR'] ,
            f"L{start_row + 36}": loader.ly_com_sales_units['MAY'],
            f"M{start_row + 36}":loader.ly_com_sales_units['JUN'] ,
            f"N{start_row + 36}":loader.ly_com_sales_units['JUL'] ,
            f"O{start_row + 36}":loader.ly_com_sales_units['AUG'] ,
            f"P{start_row + 36}":loader.ly_com_sales_units['SEP'],
            f"Q{start_row + 36}":loader.ly_com_sales_units['OCT'] ,
            f"R{start_row + 36}":loader.ly_com_sales_units['NOV'],
            f"S{start_row + 36}":loader.ly_com_sales_units['DEC'] ,
            f"T{start_row + 36}":loader.ly_com_sales_units['JAN'] ,

            f"U{start_row + 36}": f"=SUM(I{start_row + 36}:T{start_row + 36})",
            f"V{start_row + 36}": f"=SUM(I{start_row + 36}:N{start_row + 36})",
            f"W{start_row + 36}": f"=SUM(O{start_row + 36}:T{start_row + 36})",
            f"I{start_row + 35}": f"=I{start_row + 34}-I{start_row + 36}",
            f"J{start_row + 35}": f"=J{start_row + 34}-J{start_row + 36}",
            f"K{start_row + 35}": f"=K{start_row + 34}-K{start_row + 36}",
            f"L{start_row + 35}": f"=L{start_row + 34}-L{start_row + 36}",
            f"M{start_row + 35}": f"=M{start_row + 34}-M{start_row + 36}",
            f"N{start_row + 35}": f"=N{start_row + 34}-N{start_row + 36}",
            f"O{start_row + 35}": f"=O{start_row + 34}-O{start_row + 36}",
            f"P{start_row + 35}": f"=P{start_row + 34}-P{start_row + 36}",
            f"Q{start_row + 35}": f"=Q{start_row + 34}-Q{start_row + 36}",
            f"R{start_row + 35}": f"=R{start_row + 34}-R{start_row + 36}",
            f"S{start_row + 35}": f"=S{start_row + 34}-S{start_row + 36}",
            f"T{start_row + 35}": f"=T{start_row + 34}-T{start_row + 36}",

            f"U{start_row + 35}": f"=IFERROR(SUM(I{start_row + 35}:T{start_row + 35}),0)",
            f"V{start_row + 35}": f"=IFERROR(SUM(I{start_row + 35}:N{start_row + 35}),0)",
            f"W{start_row + 35}": f"=IFERROR(SUM(O{start_row + 35}:T{start_row + 35}),0)",
            f"I{start_row + 27}": f"=IFERROR(IF(I{start_row + 13}+I{start_row + 35}=0,0,(I{start_row + 13}-I{start_row + 35})/I{start_row + 35}),1)",
            f"J{start_row + 27}": f"=IFERROR(IF(J{start_row + 13}+J{start_row + 35}=0,0,(J{start_row + 13}-J{start_row + 35})/J{start_row + 35}),1)",
            f"K{start_row + 27}": f"=IFERROR(IF(K{start_row + 13}+K{start_row + 35}=0,0,(K{start_row + 13}-K{start_row + 35})/K{start_row + 35}),1)",
            f"L{start_row + 27}": f"=IFERROR(IF(L{start_row + 13}+L{start_row + 35}=0,0,(L{start_row + 13}-L{start_row + 35})/L{start_row + 35}),1)",
            f"M{start_row + 27}": f"=IFERROR(IF(M{start_row + 13}+M{start_row + 35}=0,0,(M{start_row + 13}-M{start_row + 35})/M{start_row + 35}),1)",
            f"N{start_row + 27}": f"=IFERROR(IF(N{start_row + 13}+N{start_row + 35}=0,0,(N{start_row + 13}-N{start_row + 35})/N{start_row + 35}),1)",
            f"O{start_row + 27}": f"=IFERROR(IF(O{start_row + 13}+O{start_row + 35}=0,0,(O{start_row + 13}-O{start_row + 35})/O{start_row + 35}),1)",
            f"P{start_row + 27}": f"=IFERROR(IF(P{start_row + 13}+P{start_row + 35}=0,0,(P{start_row + 13}-P{start_row + 35})/P{start_row + 35}),1)",
            f"Q{start_row + 27}": f"=IFERROR(IF(Q{start_row + 13}+Q{start_row + 35}=0,0,(Q{start_row + 13}-Q{start_row + 35})/Q{start_row + 35}),1)",
            f"R{start_row + 27}": f"=IFERROR(IF(R{start_row + 13}+R{start_row + 35}=0,0,(R{start_row + 13}-R{start_row + 35})/R{start_row + 35}),1)",
            f"S{start_row + 27}": f"=IFERROR(IF(S{start_row + 13}+S{start_row + 35}=0,0,(S{start_row + 13}-S{start_row + 35})/S{start_row + 35}),1)",
            f"T{start_row + 27}": f"=IFERROR(IF(T{start_row + 13}+T{start_row + 35}=0,0,(T{start_row + 13}-T{start_row + 35})/T{start_row + 35}),1)",
            f"U{start_row + 27}": f"=IFERROR(IF(U{start_row + 13}+U{start_row + 35}=0,0,(U{start_row + 13}-U{start_row + 35})/U{start_row + 35}),1)",
            f"V{start_row + 27}": f"=IFERROR(IF(V{start_row + 13}+V{start_row + 35}=0,0,(V{start_row + 13}-V{start_row + 35})/V{start_row + 35}),1)",
            f"W{start_row + 27}": f"=IFERROR(IF(W{start_row + 13}+W{start_row + 35}=0,0,(W{start_row + 13}-W{start_row + 35})/W{start_row + 35}),1)",
            f"I{start_row + 28}": f"=IFERROR(IF(I{start_row + 36}+I{start_row + 14}=0,0,(I{start_row + 14}-I{start_row + 36})/I{start_row + 36}),1)",
            f"J{start_row + 28}": f"=IFERROR(IF(J{start_row + 36}+J{start_row + 14}=0,0,(J{start_row + 14}-J{start_row + 36})/J{start_row + 36}),1)",
            f"K{start_row + 28}": f"=IFERROR(IF(K{start_row + 36}+K{start_row + 14}=0,0,(K{start_row + 14}-K{start_row + 36})/K{start_row + 36}),1)",
            f"L{start_row + 28}": f"=IFERROR(IF(L{start_row + 36}+L{start_row + 14}=0,0,(L{start_row + 14}-L{start_row + 36})/L{start_row + 36}),1)",
            f"M{start_row + 28}": f"=IFERROR(IF(M{start_row + 36}+M{start_row + 14}=0,0,(M{start_row + 14}-M{start_row + 36})/M{start_row + 36}),1)",
            f"N{start_row + 28}": f"=IFERROR(IF(N{start_row + 36}+N{start_row + 14}=0,0,(N{start_row + 14}-N{start_row + 36})/N{start_row + 36}),1)",
            f"O{start_row + 28}": f"=IFERROR(IF(O{start_row + 36}+O{start_row + 14}=0,0,(O{start_row + 14}-O{start_row + 36})/O{start_row + 36}),1)",
            f"P{start_row + 28}": f"=IFERROR(IF(P{start_row + 36}+P{start_row + 14}=0,0,(P{start_row + 14}-P{start_row + 36})/P{start_row + 36}),1)",
            f"Q{start_row + 28}": f"=IFERROR(IF(Q{start_row + 36}+Q{start_row + 14}=0,0,(Q{start_row + 14}-Q{start_row + 36})/Q{start_row + 36}),1)",
            f"R{start_row + 28}": f"=IFERROR(IF(R{start_row + 36}+R{start_row + 14}=0,0,(R{start_row + 14}-R{start_row + 36})/R{start_row + 36}),1)",
            f"S{start_row + 28}": f"=IFERROR(IF(S{start_row + 36}+S{start_row + 14}=0,0,(S{start_row + 14}-S{start_row + 36})/S{start_row + 36}),1)",
            f"T{start_row + 28}": f"=IFERROR(IF(T{start_row + 36}+T{start_row + 14}=0,0,(T{start_row + 14}-T{start_row + 36})/T{start_row + 36}),1)",
            f"U{start_row + 28}": f"=IFERROR(IF(U{start_row + 36}+U{start_row + 14}=0,0,(U{start_row + 14}-U{start_row + 36})/U{start_row + 36}),1)",
            f"V{start_row + 28}": f"=IFERROR(IF(V{start_row + 36}+V{start_row + 14}=0,0,(V{start_row + 14}-V{start_row + 36})/V{start_row + 36}),1)",
            f"W{start_row + 28}": f"=IFERROR(IF(W{start_row + 36}+W{start_row + 14}=0,0,(W{start_row + 14}-W{start_row + 36})/W{start_row + 36}),1)",    
            f"I{start_row + 40}":loader.ly_com_eom_oh['FEB'] ,
            f"J{start_row + 40}":loader.ly_com_eom_oh['MAR'],
            f"K{start_row + 40}":loader.ly_com_eom_oh['APR'],
            f"L{start_row + 40}":loader.ly_com_eom_oh['MAY'],
            f"M{start_row + 40}":loader.ly_com_eom_oh['JUN'] ,
            f"N{start_row + 40}":loader.ly_com_eom_oh['JUL'],
            f"O{start_row + 40}":loader.ly_com_eom_oh['AUG'],
            f"P{start_row + 40}":loader.ly_com_eom_oh['SEP'] ,
            f"Q{start_row + 40}":loader.ly_com_eom_oh['OCT'] ,
            f"R{start_row + 40}":loader.ly_com_eom_oh['NOV'],
            f"S{start_row + 40}":loader.ly_com_eom_oh['DEC'],
            f"T{start_row + 40}":loader.ly_com_eom_oh['JAN'] ,

            f"U{start_row + 40}": f"=AVERAGEIFS(I{start_row + 40}:T{start_row + 40},$I$4:$T$4,\"<=\"&$T$4)",
            f"V{start_row + 40}": f"=AVERAGEIFS(I{start_row + 40}:N{start_row + 40},$I$4:$N$4,\"<=\"&$N$4)",
            f"W{start_row + 40}": f"=AVERAGEIFS(O{start_row + 40}:T{start_row + 40},$O$4:$T$4,\"<=\"&$T$4)",  
            f"I{start_row + 39}": f"=I{start_row + 38}-I{start_row + 40}",
            f"J{start_row + 39}": f"=J{start_row + 38}-J{start_row + 40}",
            f"K{start_row + 39}": f"=K{start_row + 38}-K{start_row + 40}",
            f"L{start_row + 39}": f"=L{start_row + 38}-L{start_row + 40}",
            f"M{start_row + 39}": f"=M{start_row + 38}-M{start_row + 40}",
            f"N{start_row + 39}": f"=N{start_row + 38}-N{start_row + 40}",
            f"O{start_row + 39}": f"=O{start_row + 38}-O{start_row + 40}",
            f"P{start_row + 39}": f"=P{start_row + 38}-P{start_row + 40}",
            f"Q{start_row + 39}": f"=Q{start_row + 38}-Q{start_row + 40}",
            f"R{start_row + 39}": f"=R{start_row + 38}-R{start_row + 40}",
            f"S{start_row + 39}": f"=S{start_row + 38}-S{start_row + 40}",
            f"T{start_row + 39}": f"=T{start_row + 38}-T{start_row + 40}",

            f"U{start_row + 39}": f"=AVERAGEIFS(I{start_row + 39}:T{start_row + 39},$I$4:$T$4,\"<=\"&$T$4)",
            f"V{start_row + 39}": f"=AVERAGEIFS(I{start_row + 39}:N{start_row + 39},$I$4:$N$4,\"<=\"&$N$4)",
            f"W{start_row + 39}": f"=AVERAGEIFS(O{start_row + 39}:T{start_row + 39},$O$4:$T$4,\"<=\"&$T$4)",
            f"I{start_row + 29}": f"=IFERROR(IF(I{start_row + 17}+I{start_row + 39}=0,0,(I{start_row + 17}-I{start_row + 39})/I{start_row + 39}),1)",
            f"J{start_row + 29}": f"=IFERROR(IF(J{start_row + 17}+J{start_row + 39}=0,0,(J{start_row + 17}-J{start_row + 39})/J{start_row + 39}),1)",
            f"K{start_row + 29}": f"=IFERROR(IF(K{start_row + 17}+K{start_row + 39}=0,0,(K{start_row + 17}-K{start_row + 39})/K{start_row + 39}),1)",
            f"L{start_row + 29}": f"=IFERROR(IF(L{start_row + 17}+L{start_row + 39}=0,0,(L{start_row + 17}-L{start_row + 39})/L{start_row + 39}),1)",
            f"M{start_row + 29}": f"=IFERROR(IF(M{start_row + 17}+M{start_row + 39}=0,0,(M{start_row + 17}-M{start_row + 39})/M{start_row + 39}),1)",
            f"N{start_row + 29}": f"=IFERROR(IF(N{start_row + 17}+N{start_row + 39}=0,0,(N{start_row + 17}-N{start_row + 39})/N{start_row + 39}),1)",
            f"O{start_row + 29}": f"=IFERROR(IF(O{start_row + 17}+O{start_row + 39}=0,0,(O{start_row + 17}-O{start_row + 39})/O{start_row + 39}),1)",
            f"P{start_row + 29}": f"=IFERROR(IF(P{start_row + 17}+P{start_row + 39}=0,0,(P{start_row + 17}-P{start_row + 39})/P{start_row + 39}),1)",
            f"Q{start_row + 29}": f"=IFERROR(IF(Q{start_row + 17}+Q{start_row + 39}=0,0,(Q{start_row + 17}-Q{start_row + 39})/Q{start_row + 39}),1)",
            f"R{start_row + 29}": f"=IFERROR(IF(R{start_row + 17}+R{start_row + 39}=0,0,(R{start_row + 17}-R{start_row + 39})/R{start_row + 39}),1)",
            f"S{start_row + 29}": f"=IFERROR(IF(S{start_row + 17}+S{start_row + 39}=0,0,(S{start_row + 17}-S{start_row + 39})/S{start_row + 39}),1)",
            f"T{start_row + 29}": f"=IFERROR(IF(T{start_row + 17}+T{start_row + 39}=0,0,(T{start_row + 17}-T{start_row + 39})/T{start_row + 39}),1)",
            f"U{start_row + 29}": f"=IFERROR(IF(U{start_row + 17}+U{start_row + 39}=0,0,(U{start_row + 17}-U{start_row + 39})/U{start_row + 39}),1)",
            f"V{start_row + 29}": f"=IFERROR(IF(V{start_row + 17}+V{start_row + 39}=0,0,(V{start_row + 17}-V{start_row + 39})/V{start_row + 39}),1)",
            f"W{start_row + 29}": f"=IFERROR(IF(W{start_row + 17}+W{start_row + 39}=0,0,(W{start_row + 17}-W{start_row + 39})/W{start_row + 39}),1)",
            f"I{start_row + 30}": loader.ty_omni_oo_units['FEB'],
            f"J{start_row + 30}": loader.ty_omni_oo_units['MAR'],
            f"K{start_row + 30}": loader.ty_omni_oo_units['APR'],
            f"L{start_row + 30}": loader.ty_omni_oo_units['MAY'],
            f"M{start_row + 30}":loader.ty_omni_oo_units['JUN'],
            f"N{start_row + 30}":loader.ty_omni_oo_units['JUL'],
            f"O{start_row + 30}": loader.ty_omni_oo_units['AUG'],
            f"P{start_row + 30}": loader.ty_omni_oo_units['SEP'],
            f"Q{start_row + 30}": loader.ty_omni_oo_units['OCT'],
            f"R{start_row + 30}": loader.ty_omni_oo_units['NOV'],
            f"S{start_row + 30}": loader.ty_omni_oo_units['DEC'],
            f"T{start_row + 30}": loader.ty_omni_oo_units['JAN'],

            f"U{start_row + 30}": f"=IFERROR(LOOKUP(2,1/($I${start_row}:$T${start_row}=$J$1),I{start_row + 30}:T{start_row + 30}),0)",
            f"V{start_row + 30}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($I${start_row}:$N${start_row}=$J$2),I{start_row + 30}:N{start_row + 30}),0),0),0)",
            f"W{start_row + 30}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($O${start_row}:$T${start_row}=$J$1),O{start_row + 30}:T{start_row + 30}),0),0),0)",
            f"I{start_row + 31}": loader.ty_com_oo_units['FEB'],
            f"J{start_row + 31}": loader.ty_com_oo_units['MAR'],
            f"K{start_row + 31}": loader.ty_com_oo_units['APR'],
            f"L{start_row + 31}": loader.ty_com_oo_units['MAY'],
            f"M{start_row + 31}": loader.ty_com_oo_units['JUN'],
            f"N{start_row + 31}": loader.ty_com_oo_units['JUL'],
            f"O{start_row + 31}": loader.ty_com_oo_units['AUG'],
            f"P{start_row + 31}":loader.ty_com_oo_units['SEP'],
            f"Q{start_row + 31}": loader.ty_com_oo_units['OCT'],
            f"R{start_row + 31}": loader.ty_com_oo_units['NOV'],
            f"S{start_row + 31}": loader.ty_com_oo_units['DEC'],
            f"T{start_row + 31}":loader.ty_com_oo_units['JAN'],

            f"U{start_row + 31}": f"=IFERROR(LOOKUP(2,1/($I${start_row}:$T${start_row}=$J$1),I{start_row + 31}:T{start_row + 31}),0)",
            f"V{start_row + 31}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($I${start_row}:$N${start_row}=$J$2),I{start_row + 31}:N{start_row + 31}),0),0),0)",
            f"W{start_row + 31}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($O${start_row}:$T${start_row}=$J$1),O{start_row + 31}:T{start_row + 31}),0),0),0)",
            f"I{start_row + 37}": f"=IFERROR(I{start_row + 36}/I{start_row + 34},0)",
            f"J{start_row + 37}": f"=IFERROR(J{start_row + 36}/J{start_row + 34},0)",
            f"K{start_row + 37}": f"=IFERROR(K{start_row + 36}/K{start_row + 34},0)",
            f"L{start_row + 37}": f"=IFERROR(L{start_row + 36}/L{start_row + 34},0)",
            f"M{start_row + 37}": f"=IFERROR(M{start_row + 36}/M{start_row + 34},0)",
            f"N{start_row + 37}": f"=IFERROR(N{start_row + 36}/N{start_row + 34},0)",
            f"O{start_row + 37}": f"=IFERROR(O{start_row + 36}/O{start_row + 34},0)",
            f"P{start_row + 37}": f"=IFERROR(P{start_row + 36}/P{start_row + 34},0)",
            f"Q{start_row + 37}": f"=IFERROR(Q{start_row + 36}/Q{start_row + 34},0)",
            f"R{start_row + 37}": f"=IFERROR(R{start_row + 36}/R{start_row + 34},0)",
            f"S{start_row + 37}": f"=IFERROR(S{start_row + 36}/S{start_row + 34},0)",
            f"T{start_row + 37}": f"=IFERROR(T{start_row + 36}/T{start_row + 34},0)",
            f"U{start_row + 37}": f"=IFERROR(U{start_row + 36}/U{start_row + 34},0)",
            f"V{start_row + 37}": f"=IFERROR(V{start_row + 36}/V{start_row + 34},0)",
            f"W{start_row + 37}": f"=IFERROR(W{start_row + 36}/W{start_row + 34},0)",
            f"I{start_row + 41}": f"=IFERROR(I{start_row + 40}/I{start_row + 38},0)",
            f"J{start_row + 41}": f"=IFERROR(J{start_row + 40}/J{start_row + 38},0)",
            f"K{start_row + 41}": f"=IFERROR(K{start_row + 40}/K{start_row + 38},0)",
            f"L{start_row + 41}": f"=IFERROR(L{start_row + 40}/L{start_row + 38},0)",
            f"M{start_row + 41}": f"=IFERROR(M{start_row + 40}/M{start_row + 38},0)",
            f"N{start_row + 41}": f"=IFERROR(N{start_row + 40}/N{start_row + 38},0)",
            f"O{start_row + 41}": f"=IFERROR(O{start_row + 40}/O{start_row + 38},0)",
            f"P{start_row + 41}": f"=IFERROR(P{start_row + 40}/P{start_row + 38},0)",
            f"Q{start_row + 41}": f"=IFERROR(Q{start_row + 40}/Q{start_row + 38},0)",
            f"R{start_row + 41}": f"=IFERROR(R{start_row + 40}/R{start_row + 38},0)",
            f"S{start_row + 41}": f"=IFERROR(S{start_row + 40}/S{start_row + 38},0)",
            f"T{start_row + 41}": f"=IFERROR(T{start_row + 40}/T{start_row + 38},0)",
            f"U{start_row + 41}": f"=IFERROR(U{start_row + 40}/U{start_row + 38},0)",
            f"V{start_row + 41}": f"=IFERROR(V{start_row + 40}/V{start_row + 38},0)",
            f"W{start_row + 41}": f"=IFERROR(W{start_row + 40}/W{start_row + 38},0)",
            f"I{start_row + 42}":loader.ly_omni_receipts['FEB'] ,
            f"J{start_row + 42}":loader.ly_omni_receipts['MAR'],
            f"K{start_row + 42}":loader.ly_omni_receipts['APR'],
            f"L{start_row + 42}": loader.ly_omni_receipts['MAY'],
            f"M{start_row + 42}":loader.ly_omni_receipts['JUN'],
            f"N{start_row + 42}": loader.ly_omni_receipts['JUL'],
            f"O{start_row + 42}": loader.ly_omni_receipts['AUG'],
            f"P{start_row + 42}": loader.ly_omni_receipts['SEP'],
            f"Q{start_row + 42}": loader.ly_omni_receipts['OCT'],
            f"R{start_row + 42}":loader.ly_omni_receipts['NOV'],
            f"S{start_row + 42}":loader.ly_omni_receipts['DEC'],
            f"T{start_row + 42}":loader.ly_omni_receipts['JAN'],

            f"U{start_row + 42}": f"=SUM(I{start_row + 42}:T{start_row + 42})",
            f"V{start_row + 42}": f"=SUM(I{start_row + 42}:N{start_row + 42})",
            f"W{start_row + 42}": f"=SUM(O{start_row + 42}:T{start_row + 42})",
            f"I{start_row + 43}": f"=IFERROR(IFERROR(I{start_row + 34}/(I{start_row + 34}+I{start_row + 38}),0),0)",
            f"J{start_row + 43}": f"=IFERROR(IFERROR(J{start_row + 34}/(J{start_row + 34}+J{start_row + 38}),0),0)",
            f"K{start_row + 43}": f"=IFERROR(IFERROR(K{start_row + 34}/(K{start_row + 34}+K{start_row + 38}),0),0)",
            f"L{start_row + 43}": f"=IFERROR(IFERROR(L{start_row + 34}/(L{start_row + 34}+L{start_row + 38}),0),0)",
            f"M{start_row + 43}": f"=IFERROR(IFERROR(M{start_row + 34}/(M{start_row + 34}+M{start_row + 38}),0),0)",
            f"N{start_row + 43}": f"=IFERROR(IFERROR(N{start_row + 34}/(N{start_row + 34}+N{start_row + 38}),0),0)",
            f"O{start_row + 43}": f"=IFERROR(IFERROR(O{start_row + 34}/(O{start_row + 34}+O{start_row + 38}),0),0)",
            f"P{start_row + 43}": f"=IFERROR(IFERROR(P{start_row + 34}/(P{start_row + 34}+P{start_row + 38}),0),0)",
            f"Q{start_row + 43}": f"=IFERROR(IFERROR(Q{start_row + 34}/(Q{start_row + 34}+Q{start_row + 38}),0),0)",
            f"R{start_row + 43}": f"=IFERROR(IFERROR(R{start_row + 34}/(R{start_row + 34}+R{start_row + 38}),0),0)",
            f"S{start_row + 43}": f"=IFERROR(IFERROR(S{start_row + 34}/(S{start_row + 34}+S{start_row + 38}),0),0)",
            f"T{start_row + 43}": f"=IFERROR(IFERROR(T{start_row + 34}/(T{start_row + 34}+T{start_row + 38}),0),0)",
            f"U{start_row + 43}": f"=IFERROR(IFERROR(U{start_row + 34}/(U{start_row + 34}+U{start_row + 38}),0),0)",
            f"V{start_row + 43}": f"=IFERROR(IFERROR(V{start_row + 34}/(V{start_row + 34}+V{start_row + 38}),0),0)",
            f"W{start_row + 43}": f"=IFERROR(IFERROR(W{start_row + 34}/(W{start_row + 34}+W{start_row + 38}),0),0)",
            f"I{start_row + 44}": f"=IFERROR((I{start_row + 34}-I{start_row + 36})/((I{start_row + 34}-I{start_row + 36})+(I{start_row + 38}-I{start_row + 40})),0)",
            f"J{start_row + 44}": f"=IFERROR((J{start_row + 34}-J{start_row + 36})/((J{start_row + 34}-J{start_row + 36})+(J{start_row + 38}-J{start_row + 40})),0)",
            f"K{start_row + 44}": f"=IFERROR((K{start_row + 34}-K{start_row + 36})/((K{start_row + 34}-K{start_row + 36})+(K{start_row + 38}-K{start_row + 40})),0)",
            f"L{start_row + 44}": f"=IFERROR((L{start_row + 34}-L{start_row + 36})/((L{start_row + 34}-L{start_row + 36})+(L{start_row + 38}-L{start_row + 40})),0)",
            f"M{start_row + 44}": f"=IFERROR((M{start_row + 34}-M{start_row + 36})/((M{start_row + 34}-M{start_row + 36})+(M{start_row + 38}-M{start_row + 40})),0)",
            f"N{start_row + 44}": f"=IFERROR((N{start_row + 34}-N{start_row + 36})/((N{start_row + 34}-N{start_row + 36})+(N{start_row + 38}-N{start_row + 40})),0)",
            f"O{start_row + 44}": f"=IFERROR((O{start_row + 34}-O{start_row + 36})/((O{start_row + 34}-O{start_row + 36})+(O{start_row + 38}-O{start_row + 40})),0)",
            f"P{start_row + 44}": f"=IFERROR((P{start_row + 34}-P{start_row + 36})/((P{start_row + 34}-P{start_row + 36})+(P{start_row + 38}-P{start_row + 40})),0)",
            f"Q{start_row + 44}": f"=IFERROR((Q{start_row + 34}-Q{start_row + 36})/((Q{start_row + 34}-Q{start_row + 36})+(Q{start_row + 38}-Q{start_row + 40})),0)",
            f"R{start_row + 44}": f"=IFERROR((R{start_row + 34}-R{start_row + 36})/((R{start_row + 34}-R{start_row + 36})+(R{start_row + 38}-R{start_row + 40})),0)",
            f"S{start_row + 44}": f"=IFERROR((S{start_row + 34}-S{start_row + 36})/((S{start_row + 34}-S{start_row + 36})+(S{start_row + 38}-S{start_row + 40})),0)",
            f"T{start_row + 44}": f"=IFERROR((T{start_row + 34}-T{start_row + 36})/((T{start_row + 34}-T{start_row + 36})+(T{start_row + 38}-T{start_row + 40})),0)",
            f"U{start_row + 44}": f"=IFERROR((U{start_row + 34}-U{start_row + 36})/((U{start_row + 34}-U{start_row + 36})+(U{start_row + 38}-U{start_row + 40})),0)",
            f"V{start_row + 44}": f"=IFERROR((V{start_row + 34}-V{start_row + 36})/((V{start_row + 34}-V{start_row + 36})+(V{start_row + 38}-V{start_row + 40})),0)",
            f"W{start_row + 44}": f"=IFERROR((W{start_row + 34}-W{start_row + 36})/((W{start_row + 34}-W{start_row + 36})+(W{start_row + 38}-W{start_row + 40})),0)",
            f"I{start_row + 45}": f"=IFERROR(IFERROR(I{start_row + 34}/I{start_row + 38},0),0)",
            f"J{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:J{start_row + 34})/AVERAGE(I{start_row + 38}:J{start_row + 38}),0),0)",
            f"K{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:K{start_row + 34})/AVERAGE(I{start_row + 38}:K{start_row + 38}),0),0)",
            f"L{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:L{start_row + 34})/AVERAGE(I{start_row + 38}:L{start_row + 38}),0),0)",
            f"M{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:M{start_row + 34})/AVERAGE(I{start_row + 38}:M{start_row + 38}),0),0)",
            f"N{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:N{start_row + 34})/AVERAGE(I{start_row + 38}:N{start_row + 38}),0),0)",
            f"O{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:O{start_row + 34})/AVERAGE(I{start_row + 38}:O{start_row + 38}),0),0)",
            f"P{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:P{start_row + 34})/AVERAGE(I{start_row + 38}:P{start_row + 38}),0),0)",
            f"Q{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:Q{start_row + 34})/AVERAGE(I{start_row + 38}:Q{start_row + 38}),0),0)",
            f"R{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:R{start_row + 34})/AVERAGE(I{start_row + 38}:R{start_row + 38}),0),0)",
            f"S{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:S{start_row + 34})/AVERAGE(I{start_row + 38}:S{start_row + 38}),0),0)",
            f"T{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:T{start_row + 34})/AVERAGE(I{start_row + 38}:T{start_row + 38}),0),0)",
            f"U{start_row + 45}": f"=IFERROR(U{start_row + 34}/U{start_row + 38},0)",
            f"V{start_row + 45}": f"=IFERROR(V{start_row + 34}/V{start_row + 38},0)",
            f"W{start_row + 45}": f"=IFERROR(W{start_row + 34}/W{start_row + 38},0)",
            f"I{start_row + 46}": f"=IFERROR(IFERROR(I{start_row + 35}/I{start_row + 39},0),0)",
            f"J{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:J{start_row + 35})/AVERAGE(I{start_row + 39}:J{start_row + 39}),0),0)",
            f"K{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:K{start_row + 35})/AVERAGE(I{start_row + 39}:K{start_row + 39}),0),0)",
            f"L{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:L{start_row + 35})/AVERAGE(I{start_row + 39}:L{start_row + 39}),0),0)",
            f"M{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:M{start_row + 35})/AVERAGE(I{start_row + 39}:M{start_row + 39}),0),0)",
            f"N{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:N{start_row + 35})/AVERAGE(I{start_row + 39}:N{start_row + 39}),0),0)",
            f"O{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:O{start_row + 35})/AVERAGE(I{start_row + 39}:O{start_row + 39}),0),0)",
            f"P{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:P{start_row + 35})/AVERAGE(I{start_row + 39}:P{start_row + 39}),0),0)",
            f"Q{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:Q{start_row + 35})/AVERAGE(I{start_row + 39}:Q{start_row + 39}),0),0)",
            f"R{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:R{start_row + 35})/AVERAGE(I{start_row + 39}:R{start_row + 39}),0),0)",
            f"S{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:S{start_row + 35})/AVERAGE(I{start_row + 39}:S{start_row + 39}),0),0)",
            f"T{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:T{start_row + 35})/AVERAGE(I{start_row + 39}:T{start_row + 39}),0),0)",
            f"U{start_row + 46}": f"=IFERROR(U{start_row + 35}/U{start_row + 39},0)",
            f"V{start_row + 46}": f"=IFERROR(V{start_row + 35}/V{start_row + 39},0)",
            f"W{start_row + 46}": f"=IFERROR(W{start_row + 35}/W{start_row + 39},0)",
            f"I{start_row + 47}":loader.ly_omni_sales_usd['FEB']  ,
            f"J{start_row + 47}": loader.ly_omni_sales_usd['MAR'] ,
            f"K{start_row + 47}": loader.ly_omni_sales_usd['APR'] ,
            f"L{start_row + 47}":loader.ly_omni_sales_usd['MAY'] ,
            f"M{start_row + 47}":loader.ly_omni_sales_usd['JUN'] ,
            f"N{start_row + 47}":loader.ly_omni_sales_usd['JUL'] ,
            f"O{start_row + 47}": loader.ly_omni_sales_usd['AUG'] ,
            f"P{start_row + 47}": loader.ly_omni_sales_usd['SEP'] ,
            f"Q{start_row + 47}":loader.ly_omni_sales_usd['OCT'] ,
            f"R{start_row + 47}": loader.ly_omni_sales_usd['NOV'] ,
            f"S{start_row + 47}": loader.ly_omni_sales_usd['DEC'] ,
            f"T{start_row + 47}": loader.ly_omni_sales_usd['JAN'] ,

            f"U{start_row + 47}": f"=SUM(I{start_row + 47}:T{start_row + 47})",
            f"V{start_row + 47}": f"=SUM(I{start_row + 47}:N{start_row + 47})",
            f"W{start_row + 47}": f"=SUM(O{start_row + 47}:T{start_row + 47})",
            f"I{start_row + 48}":loader.ly_com_sales_usd['FEB'],
            f"J{start_row + 48}":loader.ly_com_sales_usd['MAR'],
            f"K{start_row + 48}": loader.ly_com_sales_usd['APR'],
            f"L{start_row + 48}":loader.ly_com_sales_usd['MAY'],
            f"M{start_row + 48}": loader.ly_com_sales_usd['JUN'],
            f"N{start_row + 48}": loader.ly_com_sales_usd['JUL'],
            f"O{start_row + 48}":loader.ly_com_sales_usd['AUG'],
            f"P{start_row + 48}": loader.ly_com_sales_usd['SEP'],
            f"Q{start_row + 48}":loader.ly_com_sales_usd['OCT'],
            f"R{start_row + 48}":loader.ly_com_sales_usd['NOV'],
            f"S{start_row + 48}":loader.ly_com_sales_usd['DEC'],
            f"T{start_row + 48}":loader.ly_com_sales_usd['JAN'],

            f"U{start_row + 48}": f"=SUM(I{start_row + 48}:T{start_row + 48})",
            f"V{start_row + 48}": f"=SUM(I{start_row + 48}:N{start_row + 48})",
            f"W{start_row + 48}": f"=SUM(O{start_row + 48}:T{start_row + 48})",
            f"I{start_row + 49}": f"=IFERROR(TEXT(IFERROR(I{start_row + 47}/I{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((I{start_row + 47}/I{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"J{start_row + 49}": f"=IFERROR(TEXT(IFERROR(J{start_row + 47}/J{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((J{start_row + 47}/J{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"K{start_row + 49}": f"=IFERROR(TEXT(IFERROR(K{start_row + 47}/K{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((K{start_row + 47}/K{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"L{start_row + 49}": f"=IFERROR(TEXT(IFERROR(L{start_row + 47}/L{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((L{start_row + 47}/L{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"M{start_row + 49}": f"=IFERROR(TEXT(IFERROR(M{start_row + 47}/M{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((M{start_row + 47}/M{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"N{start_row + 49}": f"=IFERROR(TEXT(IFERROR(N{start_row + 47}/N{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((N{start_row + 47}/N{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"O{start_row + 49}": f"=IFERROR(TEXT(IFERROR(O{start_row + 47}/O{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((O{start_row + 47}/O{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"P{start_row + 49}": f"=IFERROR(TEXT(IFERROR(P{start_row + 47}/P{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((P{start_row + 47}/P{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"Q{start_row + 49}": f"=IFERROR(TEXT(IFERROR(Q{start_row + 47}/Q{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((Q{start_row + 47}/Q{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"R{start_row + 49}": f"=IFERROR(TEXT(IFERROR(R{start_row + 47}/R{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((R{start_row + 47}/R{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"S{start_row + 49}": f"=IFERROR(TEXT(IFERROR(S{start_row + 47}/S{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((S{start_row + 47}/S{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"T{start_row + 49}": f"=IFERROR(TEXT(IFERROR(T{start_row + 47}/T{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((T{start_row + 47}/T{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"U{start_row + 49}": f"=IFERROR(TEXT(IFERROR(U{start_row + 47}/U{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((U{start_row + 47}/U{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"V{start_row + 49}": f"=IFERROR(TEXT(IFERROR(V{start_row + 47}/V{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((V{start_row + 47}/V{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"W{start_row + 49}": f"=IFERROR(TEXT(IFERROR(W{start_row + 47}/W{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((W{start_row + 47}/W{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"G{start_row + 3}": f"=IFERROR(U{start_row + 34}+U{start_row + 34}*F{start_row + 3},0)",  # Dynamic references from row 5 onward
            f"C{start_row + 7}": f"=IF(G1=\"SP\",V{start_row + 9},W{start_row + 9})",
            f"D{start_row + 7}": f"=IF($G$1=\"SP\",C{start_row + 7}-V{start_row + 32},C{start_row + 7}-W{start_row + 32})",
            f"E{start_row + 7}": f"=IF($G$1=\"SP\",C{start_row + 7}-V{start_row + 32},C{start_row + 7}-W{start_row + 32})",
            f"E{start_row + 9}": f"=SUMIFS(I{start_row + 5}:T{start_row + 5},$I$4:$T$4,\">=\"&$F$4,$I$4:$T$4,\"<=\"&$G$4)/SUMIFS($I$3:$T$3,$I$4:$T$4,\">=\"&$F$4,$I$4:$T$4,\"<=\"&$G$4)",  # Static references in rows 1 to 4
            f"C{start_row + 10}": f"=IFERROR(C{start_row + 15}/E{start_row + 9},0)",
            f"D{start_row + 10}": f"=IFERROR(C{start_row + 14}/E{start_row + 9},0)",
            f"C{start_row + 11}": f"=IFERROR((D{start_row + 10}-D{start_row + 9})*E{start_row + 9},0)",  # Dynamic D and E references from row 5 onward
            f"D{start_row + 11}": f"=C{start_row + 11}*C{start_row + 19}",  # C24 kept static because it was a specific reference
            f"F{start_row + 11}": f"=IFERROR(IF(U{start_row + 6}>C{start_row + 14},(U{start_row + 6}-C{start_row + 14}+C{start_row + 14}-F{start_row + 10}-C{start_row + 12})/E{start_row + 9},(C{start_row + 14}-F{start_row + 10}-C{start_row + 12})/E{start_row + 9}),0)",
            f"G{start_row + 11}": f"=F{start_row + 11}*E{start_row + 9}",
            f"G{start_row + 19}": f"=((U{start_row + 20}/U{start_row + 12})-C{start_row + 19})/(U{start_row + 20}/U{start_row + 12})",
            f"E{start_row + 33}": f"=U{start_row + 6}-U{start_row + 8}-E{start_row + 8}",
            f"A{start_row + 1}": f"=$C{start_row}&H{start_row + 1}",
            f"A{start_row + 2}": f"=$C{start_row}&H{start_row + 2}",
            f"A{start_row + 3}": f"=$C{start_row}&H{start_row + 3}",
            f"A{start_row + 4}": f"=$C{start_row}&H{start_row + 4}",
            f"A{start_row + 5}": f"=$C{start_row}&H{start_row + 5}",
            f"A{start_row + 6}": f"=$C{start_row}&H{start_row + 6}",
            f"A{start_row + 7}": f"=$C{start_row}&H{start_row + 7}",
            f"A{start_row + 8}": f"=$C{start_row}&H{start_row + 8}",
            f"A{start_row + 9}": f"=$C{start_row}&H{start_row + 9}",
            f"A{start_row + 10}": f"=$C{start_row}&H{start_row +10}",
            f"A{start_row + 11}": f"=$C{start_row + 11}&\"Excess Proj\"",
            f"A{start_row + 12}": f"=$C{start_row + 12}&\"Qty to Release\"",
            f"A{start_row + 13}": f"=$C{start_row}&H{start_row +12}",
            f"A{start_row + 14}": f"=$C{start_row}&H{start_row +13}",
            f"A{start_row + 15}": f"=$C{start_row}&H{start_row +14}",
            f"A{start_row + 16}": f"=$C{start_row}&H{start_row +16}",
            f"A{start_row + 17}": f"=$C{start_row}&H{start_row +17}",
            f"A{start_row + 18}": f"=$C{start_row}&H{start_row +18}",

            f"A{start_row + 20}": f"=$C{start_row}&H{start_row +20}",
            f"A{start_row + 21}": f"=$C{start_row}&H{start_row +21}",
            f"A{start_row + 22}": f"=$C{start_row}&H{start_row +22}",
            f"A{start_row + 23}": f"=$C{start_row}&H{start_row +23}",
            f"A{start_row + 24}": f"=$C{start_row}&H{start_row +24}",
            f"A{start_row + 25}": f"=$C{start_row}&H{start_row +25}",
            f"A{start_row + 26}": f"=$C{start_row}&H{start_row +26}",
            f"A{start_row + 27}": f"=$C{start_row}&H{start_row +27}",
            f"A{start_row + 28}": f"=$C{start_row}&H{start_row +28}",
            f"A{start_row + 29}": f"=$C{start_row}&H{start_row +29}",
            f"A{start_row + 33}": f"=$C{start_row + 33}&\"Qty to Enter\"",
            f"A{start_row + 35}": f"=$C{start_row}&B{start_row +34}",
            f"A{start_row + 40}": f"=$C{start_row}&B{start_row +39}",
            f"I{start_row + 5}":planned_fc["FEB"],
            f"J{start_row + 5}":planned_fc["MAR"],
            f"K{start_row + 5}":planned_fc["APR"],
            f"L{start_row + 5}":planned_fc["MAY"],
            f"M{start_row + 5}":planned_fc["JUN"],
            f"N{start_row + 5}":planned_fc["JUL"],
            f"O{start_row + 5}":planned_fc["AUG"],
            f"P{start_row + 5}":planned_fc["SEP"],
            f"Q{start_row + 5}":planned_fc["OCT"],
            f"R{start_row + 5}":planned_fc["NOV"],
            f"S{start_row + 5}":planned_fc["DEC"],
            f"T{start_row + 5}":planned_fc["JAN"],

                }
                # Add dropdown validation to the specific cell (e.g., F column) for each loop

        # Apply formulas
        percentage_ranges = [
        f"I{start_row + 1}:W{start_row + 1}",
        f"I{start_row + 10}:W{start_row + 10}",
        f"I{start_row + 15}:W{start_row + 15}",
        f"I{start_row + 19}:W{start_row + 19}",
        f"I{start_row + 23}:W{start_row + 23}",
        f"I{start_row + 24}:W{start_row + 24}",
        f"I{start_row + 27}:W{start_row + 27}",
        f"I{start_row +28}:W{start_row + 28}",
        f"I{start_row +29}:W{start_row + 29}",
        f"I{start_row + 37}:W{start_row + 37}",
        f"I{start_row + 41}:W{start_row + 41}",
        f"I{start_row + 43}:W{start_row + 43}",
        f"I{start_row + 44}:W{start_row + 44}",
        f"E{start_row + 2}",
        f"C{start_row + 3}",
        f"F{start_row + 19}",
        f"G{start_row + 19}",
        f"F{start_row + 3}",
    ]
        rounded_ranges = [
            f"F{start_row + 2}",
            f"G{start_row + 3}",
            f"E{start_row + 9}",
            f"C{start_row + 10}",
            f"D{start_row + 10}",
            f"C{start_row + 11}",
            f"F{start_row + 11}",
            f"F{start_row + 18}",
            f"U{start_row + 7}:W{start_row + 7}",
            f"U{start_row + 16}:W{start_row + 16}",
            f"U{start_row + 17}:W{start_row + 17}",
            f"U{start_row + 18}:W{start_row + 18}",
            f"I{start_row + 20}:W{start_row + 20}",
            f"I{start_row + 21}:W{start_row + 21}",
            f"U{start_row + 38}:W{start_row + 38}",
            f"U{start_row + 39}:W{start_row + 39}",
            f"U{start_row + 40}:W{start_row + 40}",


        ]
        rounded_ranges_one = [

            f"I{start_row + 25}:T{start_row + 25}",
            f"I{start_row + 26}:T{start_row + 26}",
            f"I{start_row + 45}:T{start_row + 45}",
            f"I{start_row + 46}:T{start_row + 46}",

        ]
        rounded_ranges_two = [
            f"U{start_row + 25}:W{start_row + 25}",
            f"U{start_row + 26}:W{start_row + 26}",
            f"U{start_row + 45}:W{start_row + 45}",
            f"U{start_row + 46}:W{start_row + 46}",




        ]
        currency_ranges = [
        f"D{start_row + 11}",
        f"C{start_row + 19}",
        f"D{start_row + 19}",
        f"C{start_row + 20}",
        f"D{start_row + 12}",
        f"E{start_row + 19}",
        f"I{start_row + 20}:W{start_row + 20}",
        f"I{start_row + 21}:W{start_row + 21}",
        f"I{start_row + 47}:W{start_row + 47}",
        f"I{start_row + 48}:W{start_row + 48}",
        
        # Add more ranges as needed
    ]
        if pid_type == 'com_pid' and not pid_omni_status:
            key = f"{MONTH_COLUMN_MAP[current_month]}{start_row + 7}"
            value = loader.ty_com_eom_oh[current_month]+planned_shp[current_month]-planned_fc[current_month]
            dynamic_formulas[key] = value
        sheet = ws
                # Add values to column B from ALL_VALUES and apply alignment
        for i, value in enumerate(ALL_VALUES, start=start_row):
            cell = ws.cell(row=i, column=2, value=value)  # Column B is the 2nd column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add values to column H from H_VALUES and apply alignment
        for i, value in enumerate(H_VALUES, start=start_row):
            cell = ws.cell(row=i, column=8, value=value)  # Column H is the 8th column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add monthly values starting from column I (9th column) in the specified row
        for col, value in enumerate(MONTHLY_VALUES, start=9):
            ws.cell(row=start_row, column=col, value=value)
        for cell, formula in dynamic_formulas.items():
            try:
                ws[cell] = formula  # Set the formula directly in the specified cell
            except Exception as e:
                print(f"Error setting formula in {cell}: {e}")

        # Apply ROUND formatting with different decimal places
        apply_round_format(ws, rounded_ranges, decimal_places=0)
        apply_round_format(ws, rounded_ranges_one, decimal_places=1)
        apply_round_format(ws, rounded_ranges_two, decimal_places=2)
    # Apply formats
        apply_format(ws, percentage_ranges, "0%")        # Apply percentage format
        apply_format(ws, currency_ranges, "$#,##0") 
        # add_dropdown(ws, f"F{start_row + 1}", category_options)
        add_dropdown(ws, f"F{start_row + 4}", forecast_method_options)




    #######loop formatinf
        red_font= Font(color="FF0000") 
        blue_font=Font(color="0000FF") 
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        dark_pink = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        pink_font = Font(color="9C0006") 
        for row in ws[f"C{start_row}:G{start_row + 8}"]:
            for cell in row:
                cell.fill = gray

        for row in ws[f"C{start_row + 13}:G{start_row + 20}"]:  # Use the dynamically constructed range
            for cell in row:
                cell.fill = gray
        for row in ws[f"C{start_row + 27}:G{start_row + 32}"]:  # Offset the range dynamically
            for cell in row:
                cell.fill = gray  # Apply the gray fill 

        ws[f"D{start_row + 1}"].fill = white  # D6 becomes D(start_row + 1)
        ws[f"E{start_row + 3}"].fill = white  # E8 becomes E(start_row + 3)

        red_font_list = [f"B{start_row + 40}", f"B{start_row}", f"D{start_row + 16}", f"F{start_row + 7}"]
        for cell_address in red_font_list:
            ws[cell_address].font = red_font

        blue_font_list = [
            f"B{start_row + 9}",  # B14 -> start_row + 9
            f"B{start_row + 10}",  # B15 -> start_row + 10
            f"B{start_row + 33}",  # B38 -> start_row + 33
            f"D{start_row + 9}",  # D14 -> start_row + 9
            f"E{start_row + 9}",  # E14 -> start_row + 9
            f"D{start_row + 35}",  # D40 -> start_row + 35
            f"C{start_row + 10}",  # C15 -> start_row + 10
            f"D{start_row + 10}"   # D15 -> start_row + 10
        ]

        pink_fill_font = [
            f"C{start_row + 3}",  # C8 -> start_row + 3
            f"F{start_row + 3}",  # F8 -> start_row + 3
            f"F{start_row + 14}",  # F19 -> start_row + 14
            f"G{start_row + 14}",  # G19 -> start_row + 14
            f"C{start_row + 16}",  # C21 -> start_row + 16
            f"C{start_row + 24}"   # C29 -> start_row + 24
        ]
        for cell_address in pink_fill_font:
            ws[cell_address].fill = dark_pink
            ws[cell_address].font = pink_font


        # For the range H28:W31
        for row in ws[f"H{start_row + 23}:W{start_row + 26}"]:  # H28:W31 -> start_row + 23 to start_row + 26
            for cell in row:
                cell.font = blue_font

        # For the range H48:W51
        for row in ws[f"H{start_row + 43}:W{start_row + 46}"]:  # H48:W51 -> start_row + 43 to start_row + 46
            for cell in row:
                cell.font = blue_font

        # For the range I14:W14
        for row in ws[f"I{start_row + 9}:W{start_row + 9}"]:  # I14:W14 -> start_row + 9
            for cell in row:
                cell.font = red_font  
        yelow_fill_font_list=['E13','F13','Q13','R13','S13','G39',]
        for i in yelow_fill_font_list:
            ws[i].fill = light_yellow
            ws[i].font = yellow_font
        #Hyperlink
        ws[f"B{start_row + 26}"].font = Font(color="0563C1", underline="single", bold=False)  # B31 -> start_row + 26

        # Dynamically merge cells
        ws.merge_cells(f"D{start_row + 21}:G{start_row + 26}")  # D26:G31
        ws.merge_cells(f"D{start_row + 11}:E{start_row + 11}")  # D16:E16
        ws.merge_cells(f"D{start_row + 1}:E{start_row + 1}")    # D6:E6
        ws.merge_cells(f"D{start_row}:E{start_row}")            # D5:E5
        ws.merge_cells(f"D{start_row + 20}:F{start_row + 20}")  # D25:F25
        ws.merge_cells(f"B{start_row + 35}:C{start_row + 38}")  # B40:C43
        ws.merge_cells(f"D{start_row + 35}:G{start_row + 38}")  # D40:G43
        ws.merge_cells(f"C{start_row + 39}:G{start_row + 39}")  # C44:G44
        ws.merge_cells(f"B{start_row + 40}:G{start_row + 49}")  # B45:G54
        ws.merge_cells(f"C{start_row + 13}:F{start_row + 13}")  # C18:F18
        ws.merge_cells(f"D{start_row + 16}:G{start_row + 16}")  # D21:G21
        ws.merge_cells(f"C{start_row + 30}:G{start_row + 30}")  # C35:G35
        # Dynamically set alignment
        ws[f"B{start_row + 40}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)  # B45
        ws[f"D{start_row + 21}"].alignment = Alignment(horizontal='center', vertical='center')  # D26
        ws[f"D{start_row + 11}"].alignment = Alignment(horizontal='center', vertical='center')  # D16
        ws[f"D{start_row}"].alignment = Alignment(horizontal='center', vertical='center')       # D5
        ws[f"D{start_row + 1}"].alignment = Alignment(horizontal='right', vertical='center')    # D6
        ws[f"C{start_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')   # C8
        ws[f"D{start_row + 16}"].alignment = Alignment(horizontal='left', vertical='center')    # D21
        ws[f"D{start_row + 20}"].alignment = Alignment(horizontal='center', vertical='center')  # D25
        ws[f"C{start_row + 30}"].alignment = Alignment(horizontal='center', vertical='top')     # C35
        ws[f"D{start_row + 35}"].alignment = Alignment(horizontal='left', vertical='top')       # D40

        #gradient bg
        two_yellow_gradient_fill = GradientFill(type="linear",degree=90,stop=("FFFFFF", "FFFF99", "FFFFFF")) 
        top_green_gradient_fill = GradientFill(type="linear", degree=90, stop=("E2EFDA", "FFFFFF"))
        full_green_gradient_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        bott_green_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","E2EFDA"))
        bott_light_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","EDEDED"))
        top_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=("D9D9D9", "FFFFFF"))
        dark_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","DBDBDB"))
        top_orange_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFE699", "FFFFFF"))
        bott_orange_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","FFE699"))
        top_yellow_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFCC", "FFFFFF"))
        full_yellow_gradient_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        bott_yellow_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","FFFFCC"))
        bott_blue_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","DFECF7"))
        full_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        top_pink_gradient_fill = GradientFill(type="linear", degree=90, stop=("FCE4D6", "FFFFFF"))
        full_pink_gradient_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        bott_pink_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","FCE4D6"))
        # Pink
        for row in ws[f"C{start_row + 9}:G{start_row + 9}"]:  # C14:G14 -> start_row + 9
            for cell in row:
                cell.fill = bott_pink_gradient_fill  

        for row in ws[f"C{start_row + 10}:G{start_row + 11}"]:  # C15:G16 -> start_row + 10 to start_row + 11
            for cell in row:
                cell.fill = full_pink_gradient_fill  

        for row in ws[f"C{start_row + 12}:G{start_row + 12}"]:  # C17:G17 -> start_row + 12
            for cell in row:
                cell.fill = top_pink_gradient_fill 

        # Yellow
        for row in ws[f"I{start_row + 4}:W{start_row + 4}"]:  # I9:W9 -> start_row + 4
            for cell in row:
                cell.fill = two_yellow_gradient_fill  

        # Green
        for row in ws[f"I{start_row + 5}:W{start_row + 5}"]:  # I10:W10 -> start_row + 5
            for cell in row:
                cell.fill = bott_green_gradient_fill  

        for row in ws[f"I{start_row + 6}:W{start_row + 6}"]:  # I11:W11 -> start_row + 6
            for cell in row:
                cell.fill = full_green_gradient_fill  

        for row in ws[f"I{start_row + 7}:W{start_row + 7}"]:  # I12:W12 -> start_row + 7
            for cell in row:
                cell.fill = top_green_gradient_fill  

        # Blue
        bott_blue_fill_list = [
            f"I{start_row + 12}:W{start_row + 12}",  # I17:W17 -> start_row + 12
            f"I{start_row + 20}:W{start_row + 20}",  # I25:W25 -> start_row + 20
            f"I{start_row + 34}:W{start_row + 34}",  # I39:W39 -> start_row + 34
            f"I{start_row + 47}:W{start_row + 47}"   # I52:W52 -> start_row + 47
        ]
        for i in bott_blue_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = bott_blue_gradient_fill  
                    cell.font = Font(bold=True)  

        full_blue_fill_list = [
            f"I{start_row + 13}:W{start_row + 14}",  # I18:W19 -> start_row + 13 to start_row + 14
            f"I{start_row + 21}:W{start_row + 22}",  # I26:W27 -> start_row + 21 to start_row + 22
            f"I{start_row + 35}:W{start_row + 36}",  # I40:W41 -> start_row + 35 to start_row + 36
            f"I{start_row + 48}:W{start_row + 49}"   # I53:W54 -> start_row + 48 to start_row + 49
        ]
        for i in full_blue_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = full_blue_fill  
        # Gray
        for row in ws[f"I{start_row + 8}:W{start_row + 8}"]:  # I13:W13 -> start_row + 8
            for cell in row:
                cell.fill = bott_light_gray_gradient_fill 

        for row in ws[f"I{start_row + 9}:W{start_row + 9}"]:  # I14:W14 -> start_row + 9
            for cell in row:
                cell.fill = top_gray_gradient_fill 

        dark_gray_fill_list = [
            f"H{start_row}:W{start_row}",       # H5:W5 -> start_row
            f"H{start_row + 11}:W{start_row + 11}",  # H16:W16 -> start_row + 11
            f"H{start_row + 33}:W{start_row + 33}"   # H38:W38 -> start_row + 33
        ]
        for i in dark_gray_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = dark_gray_gradient_fill 
                    cell.font=Font(bold=True)

        # Individual cells
        ws[f"F{start_row + 1}"].fill = dark_gray_gradient_fill  # F6 -> start_row + 1
        ws[f"F{start_row + 4}"].fill = dark_gray_gradient_fill  # F9 -> start_row + 4
        ws[f"B{start_row + 26}"].fill = dark_gray_gradient_fill  # B31 -> start_row + 26

        # Yellow
        for row in ws[f"I{start_row + 27}:W{start_row + 27}"]:  # I32:W32 -> start_row + 27
            for cell in row:
                cell.fill = bott_yellow_gradient_fill  

        for row in ws[f"I{start_row + 28}:W{start_row + 28}"]:  # I33:W33 -> start_row + 28
            for cell in row:
                cell.fill = full_yellow_gradient_fill  

        for row in ws[f"I{start_row + 29}:W{start_row + 29}"]:  # I34:W34 -> start_row + 29
            for cell in row:
                cell.fill = top_yellow_gradient_fill  

        # Orange
        top_orange_fill_list = [
            f"I{start_row + 15}:W{start_row + 15}",  # I20:W20 -> start_row + 15
            f"I{start_row + 37}:W{start_row + 37}"  # I42:W42 -> start_row + 37
        ]
        bott_orange_fill_list = [
            f"I{start_row + 19}:W{start_row + 19}",  # I24:W24 -> start_row + 19
            f"I{start_row + 41}:W{start_row + 41}"  # I46:W46 -> start_row + 41
        ]

        # Apply top orange gradient fill
        for i in top_orange_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = top_orange_gradient_fill 

        # Apply bottom orange gradient fill
        for i in bott_orange_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = bott_orange_gradient_fill
        bold_list=['C5','D5','F5','C6','B26','B39','B44','C38','E38','G39','H20','H21','H24','H32','H33','H34','H46','H43','H42']
        for i in bold_list:
            ws[i].font = Font(bold=True)


        for row in ws[f"C{start_row}:W{start_row + 49}"]:  # Use the dynamically calculated range
            for cell in row:
                cell.border = gridline1  # Apply the border
                # Hide column A
        # Define border style for top and bottom only

        gridline_top_bottom = Border(
        left=Side(style="thin", color='FFFFFF'),
        right=Side(style="thin", color='FFFFFF'),
        top=Side(style="thin", color='D9D9D9'),
        bottom=Side(style="thin", color='D9D9D9')
    )

        for row in ws[f"B{start_row+50}:W{start_row+50}"]:
            for cell in row:
                cell.border = gridline_top_bottom
    # Define border style for left and right only

        
    ws.column_dimensions.group("A", "A", outline_level=1, hidden=True)

    wb.save(output_file)  # Save the workbook as an .xlsx file
    print(f"Workbook '{output_file}' saved successfully.")


    # At the end of this function, return a tuple of (output_file, final_data, store, coms, omni)
    return  store, coms, omni
    