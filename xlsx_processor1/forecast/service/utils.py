# from .createDataframe 
import re
import math
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from forecast.service.staticVariable import *
from forecast.service.config import sheets
from statistics import mean
from decimal import Decimal, ROUND_HALF_UP
 
def round_half_up(value, digits):
    return float(Decimal(str(value)).quantize(Decimal('1.' + '0' * digits), rounding=ROUND_HALF_UP))

def count_ttl_com_sale(LY_Unit_Sales, LY_MCOM_Unit_Sales):
    All_LY_Unit_Sales_list = [LY_Unit_Sales[month] for month in MONTHS]
    ALL_LY_MCOM_Unit_Sales = [LY_MCOM_Unit_Sales[month] for month in MONTHS]
    
    result_list = []
    for ly_mcom, ty_unit in zip(ALL_LY_MCOM_Unit_Sales, All_LY_Unit_Sales_list):
        if ly_mcom > 0 and ty_unit > 0:
            result_list.append(ly_mcom / ty_unit)

    average = sum(result_list) / len(result_list) if result_list else 0
    average_com_to_ttl_sales = average * 100
    return average_com_to_ttl_sales


def find_pid_type(Safe_Non_Safe,pid_value,LY_Unit_Sales,LY_MCOM_Unit_Sales,Door_count):
    pid_type=None
    if Safe_Non_Safe  in not_forecast_status and Door_count in [0,1,2]:
        pid_type='Not forecast'
    elif ((Safe_Non_Safe in ['FB','COM ONLY','COM REPLEN','VDF REPLEN'] or pid_value in VDF_item) and  Door_count <3) or (Safe_Non_Safe in ['OMNI'] and count_ttl_com_sale(LY_Unit_Sales,LY_MCOM_Unit_Sales)>=65):
        pid_type='com_pid'
        if Safe_Non_Safe=='OMNI' and count_ttl_com_sale(LY_Unit_Sales,LY_MCOM_Unit_Sales)<=65:
            print('com sale',count_ttl_com_sale(LY_Unit_Sales,LY_MCOM_Unit_Sales))
            pid_type='store_pid'
    elif Door_count >10 and not (Safe_Non_Safe in ['OMNI'] and count_ttl_com_sale(LY_Unit_Sales,LY_MCOM_Unit_Sales)>=65):
        pid_type='store_pid'
    else:
        pid_type='Not forecast'
    return pid_type

def get_vendor_by_pid(pid_value, master_sheet):
    """Retrieve vendor name for a given PID."""
    vendor_row = master_sheet.loc[master_sheet['PID'] == pid_value, 'Vendor Name']
    if vendor_row.empty:
        raise ValueError(f"PID {pid_value} not found in master sheet.")
    return vendor_row.values[0]
 
def get_vendor_details(vendor, vendor_sheet):
    """Retrieve country and lead time for a given vendor."""
    country_row = vendor_sheet.loc[vendor_sheet['Vendor Name'] == vendor, 'Country of Origin']
    country = country_row.values[0] if not country_row.empty else None
   
    lead_time_row = vendor_sheet.loc[vendor_sheet['Vendor Name'] == vendor, 'Lead Time(weeks)']
    lead_time = lead_time_row.values[0] if not lead_time_row.empty else 8
 
    if country == "South Africa" or country == "DRL":
        lead_time = lead_time + 2
 
    if np.isnan(lead_time):
        lead_time=8
   
    return country, lead_time
 
def calculate_forecast_date_basic(current_date: datetime, lead_time: int, country) -> datetime:
    """
    Calculate forecast date based on lead time (in weeks).
    If country is Italy and August is between current_date and forecast_date, skip August.
    """
    raw_forecast_date = current_date + timedelta(weeks=lead_time)
    return raw_forecast_date

def calculate_forecast_date(current_date: datetime, lead_time: int, country) -> datetime:
    """
    Calculate forecast date based on lead time (in weeks).
    If country is Italy and August is between current_date and forecast_date, skip August.
    """
    raw_forecast_date = current_date + timedelta(weeks=lead_time)

    # Only apply special case if country is a string and Italy
    if isinstance(country, str) and country.lower() == "italy":
        # Check if August is in the range
        dt = current_date
        while dt <= raw_forecast_date:
            if dt.month == 8:  # August
                raw_forecast_date += timedelta(days=31)
                break
            dt += timedelta(days=1)

    return raw_forecast_date

def adjust_lead_time(country, current_date, forecast_date, lead_time):
    """Adjust lead time based on holiday periods."""
    holiday_periods = {
        "China": ("2025-01-22", "2025-02-05", 11),
        "Italy": ("2025-08-01", "2025-08-31", 14)
    }
   
    currentdate = current_date.strftime("%Y-%m-%d")
    forecast_lead_time_date = forecast_date.strftime("%Y-%m-%d")
    leadtime_holiday=False
    if country in holiday_periods:
        start_date, end_date, adjusted_lead_time = holiday_periods[country]
        if currentdate <= end_date and forecast_lead_time_date >= start_date:
            leadtime_holiday=True
            return adjusted_lead_time,leadtime_holiday
   
    return lead_time,leadtime_holiday
from datetime import datetime

def extend_forecast_if_italy(forecast_date: datetime, country: str) -> datetime:
    """
    If forecast_date is in August and country is Italy, set to 15 September of that year.
    Otherwise, return original forecast_date.
    """
    if forecast_date.month == 8 and country.lower() == "italy":
        return datetime(forecast_date.year, 9, 15)
    return forecast_date

def convert_month_to_abbr(month_name):
    """Convert full month name to 3-letter abbreviation."""
    return month_name[:3].upper()
 
def get_week_of_month(date):
    """Get the week number of the given date within its month."""
    first_day_of_month = date.replace(day=1)
    days_diff = (date - first_day_of_month).days
    return math.ceil((days_diff + 1) / 7)
 
def is_late_forecast_week(forecast_date):
    """Return True if the forecast week is greater than week 2."""
    return get_week_of_month(forecast_date) >= 2
 
def get_forecast_info(forecast_date):
    """Return forecast date, forecast month (full and abbreviated), and year."""
    forecast_month = forecast_date.strftime("%B")
    forecast_month = convert_month_to_abbr(forecast_month)
    return forecast_month
 
def calculate_std_index_value(index_value_dict,STD_PERIOD):
    """
    Compute rounded index values and total for given months.
    """
    # Round index values for each STD month
    std_period_index_value_list = [index_value_dict[month] for month in STD_PERIOD]
    # Sum the index values
    total_std_period_index_value = round(sum(std_period_index_value_list),2)
 
    return total_std_period_index_value
 
def calculate_12th_month_forecast(std_ty_unit_sales_list, total_std_period_index_value):
    """
    Forecast 12th month using STD unit sales and index value.
    """
    # Get unit sales for STD months
    total_unit_sales = sum(std_ty_unit_sales_list)
 
    # Calculate forecast if index sum is non-zero
    if total_std_period_index_value:
        month_12_fc_index = round(total_unit_sales / total_std_period_index_value, 0)
    else:
        month_12_fc_index = 0
 
    return month_12_fc_index
 
def calculate_std_trend(std_ty_unit_sales_list, std_ly_unit_sales_list):
    """
    Calculate STD trend from LY and TY unit sales.
    """
    # Sum LY and TY unit sales
    ly_total = sum(std_ly_unit_sales_list)
    ty_total = sum(std_ty_unit_sales_list)
 
    # Compute trend if both totals are non-zero
    if ly_total and ty_total:
        std_trend = round((ty_total - ly_total) / ly_total, 2)
    else:
        std_trend = 0
 
    return std_trend
 
def calculate_fc_by_trend(s1, k1,f8, row4_values, row17_values, row39_values):
    fc_by_trend = {}
    for month,row4, row17, row39 in zip(MONTHS,row4_values, row17_values, row39_values):
        if s1 == 12 and k1 == 12:
            result = round(row17 + row17 * f8, 0)
        elif s1 == 12 and row4 < 7:
            result = round(row39 + row39 * f8, 0)
        elif s1 > 6 and row4 < 7:
            result = round(row17 + row17 * f8, 0)
        elif s1 < 7 and row4 > 6:
            result = round(row39 + row39 * f8, 0)
        elif s1 < 7 and row4 < 7:
            result = round(row39 + row39 * f8, 0)
        elif s1 > 6 and row4 > 6:
            result = round(row39 + row39 * f8, 0)
        else:
            result = None  # Fallback case
 
        fc_by_trend[month] = result
 
    return fc_by_trend

def calculate_fc_by_index(index_value, month_12_fc_index):
    """
    Generate forecast for each month using index values.
    """
    fc_by_index = {}
    for month in index_value:   # loop directly on 'FEB', 'MAR', etc.
        fc_by_index[month] = round(index_value[month] * month_12_fc_index, 0)
    return fc_by_index
 
def calculate_fc_by_average(fc_by_index, fc_by_trend):
    """
    Compute average forecast from index and trend forecasts.
    """
    fc_by_average= {
        key: round((fc_by_index[key] + fc_by_trend[key]) / 2)
        for key in fc_by_trend
    }
    return fc_by_average
 
def get_forecast_month_season(forecast_month):
    """
    Return season based on forecast month.
    """
    # Check which season the month belongs to
    if forecast_month in SPRING_MONTHS:
        return "SPRING"
    elif forecast_month in FALL_MONTHS:
        return "FALL"
    else:
        return "UNKNOWN"
def find_season_list(season):
    if season =='SPRING':
        season_month=SPRING_MONTHS
    else:
        season_month=FALL_MONTHS
    return season_month

def last_year_eom_oh_season(LY_OH_Units,LY_MCOM_OH_Units,season_month):
        LY_OH_Units_list_for_inventory_check = [LY_OH_Units[month] for month in season_month]
        LY_OH_MCOM_Units_list_for_inventory_check=[LY_MCOM_OH_Units[month] for month in season_month]
        
        last_year_store_eom_oh_for_inventory_check = [
            (ly_oh_unit - ly_oh_mcom )
            for ly_oh_unit,ly_oh_mcom  in zip(LY_OH_Units_list_for_inventory_check, LY_OH_MCOM_Units_list_for_inventory_check)
        ]
        return last_year_store_eom_oh_for_inventory_check
def is_maintained(eom_oh_list, threshold, door_count):
    """
    Check if average EOM OH is within threshold or above door count.
    """
    # Calculate average EOM OH
    average_eom_oh = sum(eom_oh_list) / len(eom_oh_list) if eom_oh_list else 0
 
    # Check if it's maintained
    return (average_eom_oh >= threshold * door_count) or (average_eom_oh > door_count)
 
def is_maintained_for_com(units, threshold=2, ratio=0.2):
    near_zero_count = sum(1 for x in units if x <= threshold)
    return near_zero_count / len(units) < ratio 
import logging
 
def decide_forecasting_method(inventory_maintained):
    """
    Decide forecasting method based on last year's inventory maintenance.
    """
 
    # Choose forecasting method
    if inventory_maintained:
        forecasting_method = "FC By Trend"
    else:
        forecasting_method = "FC By Index"
 
    return forecasting_method
 
def calculate_loss(door_count, average_value):
    """
    Calculate loss percentage from door count and average store EOM OH.
    """
    return (door_count / average_value) - 1
 
def determine_loss_percent(loss, rank, own_retail):
    """
    Determine the final loss percent based on rank and own retail price.
    """
    if own_retail < 1000:
        if rank in ['A', 'B']:
            return min(loss, 0.45)
        elif rank == 'C':
            return min(loss, 0.15)
        else:
            return min(loss, 0.10)
    else:
        return min(loss, 0.15)
 
def update_12_month_forecast_by_loss(month_12_fc_index, loss_percent):
    """
    Update 12-month forecast using loss percentage.
    """
    month_12_fc_index = round(month_12_fc_index * (1 + loss_percent))
    return month_12_fc_index
def new_trend_month_selection(current_month_weeks,current_month_upper,season_month):
    start_index = 0  # Always start from the first month in spring_months_upper
    if current_month_weeks > 2:
        end_index = season_month.index(current_month_upper) + 1 # Include the FCM month
    else:
        end_index = season_month.index(current_month_upper)
    
    selected_months_for_trend = season_month[start_index:end_index]
    return selected_months_for_trend

def is_same_sales(this_year_sales, last_year_sales, tolerance=0.2):
    """    
    Check if sales between two years are within ±tolerance of each other (bidirectional).
    """
    this_year_sales = np.array(this_year_sales, dtype=np.float64)
    last_year_sales = np.array(last_year_sales, dtype=np.float64)
 
    if len(this_year_sales) != len(last_year_sales):
        raise ValueError("Both sales lists must have the same length.")
 
    ratio = this_year_sales / last_year_sales
    inverse_ratio = last_year_sales / this_year_sales
 
    is_within_tolerance = (ratio >= 1 - tolerance) & (ratio <= 1 + tolerance) | \
                        (inverse_ratio >= 1 - tolerance) & (inverse_ratio <= 1 + tolerance)
 
    return is_within_tolerance.all()
 
def compare_seasonal_forecasts_by_method(fc_by_index, fc_by_trend,forcecast_month_season):
    """
    Compare seasonal forecasts and decide if average method can be used.
    """
    spring_fc_by_index_all = sum(fc_by_index[month] for month in SPRING_MONTHS)
    fall_fc_by_index_all = sum(fc_by_index[month] for month in FALL_MONTHS)
   
    spring_fc_by_trend_all = sum(fc_by_trend[month] for month in SPRING_MONTHS)
    fall_fc_by_trend_all = sum(fc_by_trend[month] for month in FALL_MONTHS)
 
    seasonal_total_fc_by_index = spring_fc_by_index_all if forcecast_month_season == "SPRING" else fall_fc_by_index_all
    seasonal_total_fc_by_trend = spring_fc_by_trend_all if forcecast_month_season == "SPRING" else fall_fc_by_trend_all
 
    # Compare difference
    difference = None
    if seasonal_total_fc_by_index is not None and seasonal_total_fc_by_trend is not None:
        difference = (abs(seasonal_total_fc_by_trend - seasonal_total_fc_by_index) / max(seasonal_total_fc_by_index, seasonal_total_fc_by_trend)) * 100
 
    return difference


def adjust_std_trend_minimum(std_trend_main, std_trend_new):
    """
    Adjusts the standard trend value based on comparison between main and new trend values.
    
    Args:
        std_trend_main (float): The original standard trend.
        std_trend_new (float): The newly calculated standard trend.
        
    Returns:
        float: The adjusted standard trend (f8).
    """
    if abs(std_trend_new) <= abs(std_trend_main):
        std_trend = std_trend_new
    else:
        std_trend=std_trend_main
    if std_trend > 0.65:
        std_trend = 0.40
    elif std_trend < -0.60:
        std_trend = -0.30
    else:
        std_trend = std_trend
    return std_trend
def handle_large_trend(std_trend_main):
    if std_trend_main > 0.65:
        std_trend = 0.40
    elif std_trend_main < -0.60:
        std_trend = -0.30
    else:
        std_trend = std_trend_main  # keep original if not exceeding thresholds

    return std_trend
def contains_no_longer_red_box(text):
    # Normalize to lowercase for case-insensitive matching
    text = str(text).lower()
    # Check for both possible phrases
    patterns = [r'no longer red box', r'no longer rb']
    return any(re.search(pattern, text) for pattern in patterns)
 
 
def get_recommended_forecast(forecasting_method, fc_by_index, fc_by_trend, fc_by_average=None):
    """
    Return recommended forecast based on selected forecasting method.
    """
    fc_by_average = calculate_fc_by_average(fc_by_index, fc_by_trend)
    print(fc_by_average)
    # Select forecast based on method
    if forecasting_method == "FC By Index":
        recommended_fc = fc_by_index
    elif forecasting_method == "FC By Trend":
        recommended_fc = fc_by_trend
    else:
        recommended_fc = fc_by_average
    print(recommended_fc)
  
    return recommended_fc
 
 
def calculate_planned_fc(row_4, row_9, row_17, row_43,V1, K1):
    """
    Calculate planned forecast based on selection type and rules.
    """
    planned_fc = {}
    print(row_43)
    print("V1",V1)
    print("Calculating planned forecast...",planned_fc)
 
    # Iterate through each retail month
    for idx, col in enumerate(MONTHS):
        print(f"Processing month: {col}")
       
        J4 = row_4[idx]
        print(f"J4 value for {col}: {J4}")
        J17 = row_17[col]
        print(f"J17 value for {col}: {J17}")
        J43 = row_43[col]
        print(f"J43 value for {col}: {J43}")
        J9 = row_9[col]
        print(f"Values for {col} - J4: {J4}, J17: {J17}, J43: {J43}, J9: {J9}")
        # Apply logic based on V1 value
        if V1 == "YTD" and J4 <= K1:
            planned_fc[col] = J17
        elif V1 == "CURRENT MTH" and J4 == K1:
            planned_fc[col] = J17
        elif V1 == "SPRING" and J4 < 7:
            planned_fc[col] = J17
        elif V1 == "FALL" and J4 > 6:
            planned_fc[col] = J17
        elif V1 == "LY FALL" and J4 > 6:
            planned_fc[col] = J43
        else:
            planned_fc[col] = J9
    print("Planned forecast calculated:", planned_fc)
 
    return planned_fc
 
def calculate_current_month_fc(current_month, ty_unit_sales):
    """
    Calculate current month forecast using TY sales and sales percentage.
    """
    # Get TY sales and percentage for current month
    ty_sales = ty_unit_sales[current_month]
    percentage = CURRENT_MONTH_SALES_PERCENTAGES
 
    # Compute forecast
    current_month_fc = round(ty_sales / (percentage / 100))
    return current_month_fc
def update_planned_fc_for_current_month(
    LY_Unit_Sales,
    recommended_fc,
    fc_by_trend,
    planned_fc,
    current_month,
    current_month_fc_by_percentage,
    current_month_weeks,
    previous_week_number,
    inventory_maintained_last_year,
    std_trend,
    check_no_red_box,
    actual_sale_unit
):
 
    # Recommended forecast
    current_month_recommended_fc = recommended_fc[current_month]
 
    # Decide week_to_compare
    week_to_compare = 4 if current_month_weeks == 5 else 3
 
    # Decide initial current_month_fc
    if previous_week_number < week_to_compare:
        current_month_fc = max(current_month_recommended_fc, current_month_fc_by_percentage)
    else:
        current_month_fc = current_month_fc_by_percentage
 
    # Adjust current_month_fc if inventory was maintained last year
    if inventory_maintained_last_year and not check_no_red_box :
        if std_trend > 0:
            if current_month_fc <= LY_Unit_Sales[current_month]:
                current_month_fc = fc_by_trend[current_month]
        elif std_trend < 0:
            if current_month_fc >= LY_Unit_Sales[current_month]:
                current_month_fc = fc_by_trend[current_month]
    if current_month_fc<actual_sale_unit :
        current_month_fc=current_month_fc_by_percentage
    if previous_week_number==4:
        current_month_fc=actual_sale_unit
    # Update planned_fc
    planned_fc[current_month] = current_month_fc if current_month_fc > 0 else 0
    return planned_fc
def calculate_in_transit_qty(D13, E19):
    """
    Calculate in-transit quantity from D13 and E19.
    """
    # Calculate in-transit value
    in_transit = 0 if (D13 - E19) < 0 else D13 - E19
 
    # Handle NaN cases
    if np.isnan(in_transit):
        in_transit = 0
 
    return in_transit
def extract_month(value):
    try:
        # Try parsing date if it's in full date format
        value =  pd.to_datetime(value).strftime('%b').upper()
        return str(value)
    except:
        # If already like 'Feb-25' or 'Apr-25'
        return str(value)[:3]
def get_return_quantity_dict(pid, df):
    # Filter rows for the given PID
    filtered = df[df['PID'] == pid]
   
    # Ensure Month is uppercase
    filtered['Month'] = filtered['Month'].str.upper()
    # Group and sum quantities per month
    monthly_qty = filtered.groupby('Month')['Quantity'].sum().to_dict()
 
    # Create ordered dict with all months, filling 0 for missing
    return_quantity_dict = {month: monthly_qty.get(month, 0) for month in MONTHS}
 
    required_quantity_dict_80_percent = {month: round(value * 0.8) for month, value in return_quantity_dict.items()}
 
    return return_quantity_dict, required_quantity_dict_80_percent
def clean_return_dict(return_quantity_dict,current_month):
    # Set months before current to 0
    current_month=find_next_month_after_forecast_month(current_month)
    for month in MONTHS:
        if month == current_month:
            break
        if month in return_quantity_dict:
            return_quantity_dict[month] = 0
    return return_quantity_dict
def calculate_planned_oh(v1, k1, row_10, row_11, row_21, row_37, row_43, row_17, current_month):
    # Define the fixed column-to-month mapping
    month_to_index = {month: idx + 1 for idx, month in enumerate(MONTHS)}  # FEB=1, ..., JAN=12
 
    # Rearrange month_order to start with the current_month
    start_idx = MONTHS.index(current_month)
    reordered_months = MONTHS[start_idx:] + MONTHS[:start_idx]
    # Initialize the output dictionary for row_12
    planned_oh = {}
 
    for idx, month in enumerate(reordered_months):
        # Determine the previous month (wrapping around if necessary)
        prev_month = reordered_months[idx - 1] if idx > 0 else reordered_months[-1]
 
        # Get the corresponding month number
        col4 = month_to_index[month]  # Month number based on the fixed mapping
        col10 = row_10.get(month, 0)
        col11 = row_11.get(month, 0)
        col21 = row_21.get(month, 0)
        col37 = row_37.get(month, 0)
        col43 = row_43.get(month, 0)
        col17 = row_17.get(month, 0)
 
        # Get the previous column's value
        prev_col12 = planned_oh.get(prev_month, 0)
        prev_col21 = row_21.get(prev_month, 0)

        if v1 == "Current MTH" and k1 == col4 and col4 == 1:
            val = row_43['JAN'] + col11 + col37 - col10
        elif v1 == "Current MTH" and k1 == col4:

            val = col21 + col11 - (col10 - col17)
        elif v1 == "Current MTH" and col4 == 1 and k1 > 1:
            val = planned_oh['JAN'] + col11 - col10
        elif v1 == "YTD" and col4 < k1:
            val = col21
        elif v1 == "Spring" and col4 < 7:
            val = col21
        elif v1 == "Fall" and col4 > 6 and col4 < k1:
            val = col21
        elif v1 == "Fall" and col4 == 1 and k1 > 1:
            val = planned_oh['JAN'] + col11 - col10
        elif v1 == "Fall" and col4 > 6 and col4 == k1:
            val = prev_col21 + col37 + col11 - col10
        elif v1 == "LY Fall" and col4 > 6:
            val = col43
        elif col4 == 1 and v1 == "LY FALL":
            val = row_43['JAN'] + col11 + col11 - col10
        elif col4 == k1 and col4 > 1:

            val = prev_col21 + col37 + col11 - col10
        else:

            val = prev_col12 + col11 - col10
        planned_oh[month] = val
 
    return planned_oh  
def get_birthstone_month(birthstone, birthstone_sheet):
    birthstone_sheet = birthstone_sheet.dropna(subset=['Birthstone'])
    birthstone_sheet['Birthstone'] = birthstone_sheet['Birthstone'].astype(str)
    birthstone_row = birthstone_sheet[birthstone_sheet['Birthstone'].str.upper() == birthstone.upper()]
    if not birthstone_row.empty:
        return birthstone_row['Month Name'].iloc[0]
    else:
        return None
 
def calculate_previous_month(birthstone_month):
    birthstone_month_abb = convert_month_to_abbr(birthstone_month)
    index = MONTHS.index(birthstone_month_abb)
    return MONTHS[index - 1]
 
def get_required_quantity_for_bsp(forecast_month, previous_month, category, KPI_Door_count):
    logging.info(f"Processing forecast month: {forecast_month}")
    if forecast_month == previous_month or forecast_month == 'Nov':
        if category == 'Studs':
            return 3 * KPI_Door_count
        elif category == 'Pendants':
            return 2 * KPI_Door_count
        else:
            return KPI_Door_count
    else:
        return KPI_Door_count

import logging

def get_required_quantity_for_nonbsp(forecast_month, previous_month, KPI_Door_count):
    if forecast_month == previous_month or forecast_month == 'Nov':
        quantity = round(1.8 * KPI_Door_count, 0)
        logging.info(f'Non-BSP: Forecast month matches previous or is Nov, required_quantity: {quantity}')
        return quantity
    else:
        logging.info(f'Non-BSP: Forecast month is different, required_quantity: {KPI_Door_count}')
        return KPI_Door_count

def process_bsp_or_nonbsp_product(bsp_row, birthstone_sheet, forecast_month, KPI_Door_count, pid_value):
    bsp_or_not = bsp_row['BSP_or_not'].iloc[0]
    bsp_or_not = str(bsp_or_not).strip().upper() if bsp_or_not else None
    logging.info(f'Processing PID: {pid_value}, BSP_or_not: {bsp_or_not}')
    birthstone_status=False
    required_quantity = None

    if bsp_or_not == 'BSP':
        logging.info(f'Added {pid_value} to all_birthstone_products [BSP]')
        
        category_bsp = bsp_row['category'].iloc[0]
        birthstone = bsp_row['Birthstone'].iloc[0]
        birthstone = str(birthstone).strip()
        logging.info(f'BSP Category: {category_bsp}, Birthstone: {birthstone}')
        
        birthstone_month = get_birthstone_month(birthstone, birthstone_sheet)
        logging.info(f'Birthstone month for {birthstone}: {birthstone_month}')
        
        if birthstone_month:
            previous_month = calculate_previous_month(birthstone_month)
            logging.info(f'Previous month of birthstone: {previous_month}')
            
            required_quantity = get_required_quantity_for_bsp(
                forecast_month, previous_month, category_bsp, KPI_Door_count
            )
            logging.info(f'Calculated BSP required_quantity: {required_quantity}')
            
            if forecast_month == previous_month or forecast_month == 'Nov':
                birthstone_status=True
                logging.info(f'Added {pid_value} to upcoming_birthstone_products [BSP]')
    else:
        logging.info(f'Added {pid_value} to all_birthstone_products [NON BSP]')

        category_bsp = bsp_row['category'].iloc[0]
        birthstone = bsp_row['Birthstone'].iloc[0]
        birthstone = str(birthstone).strip()
        logging.info(f'Non-BSP Category: {category_bsp}, Birthstone: {birthstone}')

        birthstone_month = get_birthstone_month(birthstone, birthstone_sheet)
        logging.info(f'Birthstone month for {birthstone}: {birthstone_month}')

        if birthstone_month:
            previous_month = calculate_previous_month(birthstone_month)
            logging.info(f'Previous month of birthstone: {previous_month}')

            required_quantity = get_required_quantity_for_nonbsp(
                forecast_month, previous_month, KPI_Door_count
            )
            logging.info(f'Calculated NON-BSP required_quantity: {required_quantity}')
            
            if forecast_month == previous_month or forecast_month == 'Nov':
                birthstone_status=True
                logging.info(f'Added {pid_value} to upcoming_birthstone_products [NON BSP]')

    return required_quantity,birthstone_status,birthstone,birthstone_month

def calculate_required_quantity(master_sheet, pid_value, birthstone_sheet, forecast_month, KPI_Door_count):
    required_quantity_month_dict = {}

    logging.info(f'Calculating required_quantity for PID: {pid_value}, Forecast Month: {forecast_month}')
    bsp_row = master_sheet[master_sheet['PID'] == pid_value]

    if not bsp_row.empty:
        required_quantity,birthstone_status,birthstone,birthstone_month = process_bsp_or_nonbsp_product(
            bsp_row, birthstone_sheet, forecast_month, KPI_Door_count, pid_value)
        if required_quantity is not None:
            required_quantity_month_dict[forecast_month] = required_quantity
            logging.info(f'Final required_quantity for {forecast_month}: {required_quantity}')
        else:
            logging.info(f'No required_quantity calculated for {pid_value}')
    else:
        logging.warning(f'No BSP row found in master_sheet for PID: {pid_value}')

    if not required_quantity_month_dict:
        required_quantity_month_dict[forecast_month] = KPI_Door_count
        logging.info(f'Defaulted required_quantity to KPI_Door_count: {KPI_Door_count}')

    return required_quantity_month_dict,birthstone_status,birthstone,birthstone_month 

 
 
def calculate_average_com_eom_oh(TY_OH_MCOM_Units):
    ty_com_eom_oh=[TY_OH_MCOM_Units[month] for month in ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL','AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']]
    non_zero_values = [x for x in ty_com_eom_oh if x > 0]
    if non_zero_values:
        return round(sum(non_zero_values) / len(non_zero_values), 0)
    else:
        return 2
 
def find_next_month_after_forecast_month(forecast_month):
    forecast_month_index = MONTHS.index(forecast_month)
 
    if forecast_month_index == len(MONTHS) - 1:  # If December (last month)
        return MONTHS[0] # February
    else:
        return MONTHS[(forecast_month_index + 1) % len(MONTHS)]
 
def update_required_quantity_for_forecast_month(forecast_month, planned_fc, required_quantity_month_dict, average_com_eom_oh,KPI_Door_count,country,week_of_forecast_month):
    next_month_after_forecast_month = find_next_month_after_forecast_month(forecast_month)
    # Calculate FLDC for the next month
    Calculate_FLDC = round((planned_fc[next_month_after_forecast_month]) / 2)
 
    # Update the required quantity for the current month
    required_quantity_month_dict[forecast_month] = (
        required_quantity_month_dict.get(forecast_month, 0) + Calculate_FLDC + average_com_eom_oh
    )
    required_quantity_month_dict[next_month_after_forecast_month] = KPI_Door_count + average_com_eom_oh
    if country=='Italy' and forecast_month=='JUL' and week_of_forecast_month > 2 :
        forecast_month_next_next_month=find_next_month_after_forecast_month(next_month_after_forecast_month)
        FLDC_next_month = round((planned_fc[forecast_month_next_next_month]) / 2)
        required_quantity_month_dict[next_month_after_forecast_month]+=FLDC_next_month

        required_quantity_month_dict[forecast_month_next_next_month] = KPI_Door_count + average_com_eom_oh
    return required_quantity_month_dict , Calculate_FLDC
 
 
def update_projection_for_month(month,required_quantity_month_dict,planned_oh,planned_shp,pid_value):
 
    required_quantity = required_quantity_month_dict.get(month, 0)
 
    if planned_oh[month] < required_quantity:
        # Calculate the difference between required quantity and planned OH
        difference = required_quantity - planned_oh[month]
       
        # Update gross_projection and plan_oh
        planned_shp[month] += difference
 
        # Add PID to alert list
 
    return planned_shp
 
 
 
def calculate_week_and_month(start_month_abbr, start_week, year, weeks_to_add):
    # Map abbreviated month names to their respective numbers
    month_map = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
        "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
        "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }
    # Convert the abbreviated month to its numeric equivalent
    if start_month_abbr not in month_map:
        return None,None
 
    start_month = month_map[start_month_abbr]
 
    # Define the start date as the first day of the given month and year
    start_date = datetime(year, start_month, 1)
 
    # Calculate the date corresponding to the given week in the start month
    days_to_start_week = (start_week - 1) * 7
    start_week_date = start_date + timedelta(days=days_to_start_week)
 
    # Add the specified number of weeks (convert weeks to days)
    target_date = start_week_date + timedelta(weeks=weeks_to_add)
 
    # Determine the target month and week
    target_month = target_date.month
    target_year = target_date.year
 
    # Find the week number relative to the month (using Sunday to Saturday week structure)
    first_day_of_target_month = datetime(target_year, target_month, 1)
    days_difference = (target_date - first_day_of_target_month).days
    target_week = days_difference // 7 + 1
 
    # Convert the month number to its abbreviation
    target_month_abbr = target_date.strftime("%b")
   
    return target_month_abbr,target_week
 
def check_holiday(target_month_abbr, target_week ,df_holidays):  
    # Filter the dataframe based on the target month and week
    result = df_holidays[(df_holidays['Month'] == target_month_abbr) & (df_holidays['Week'] == target_week)]
   
    if not result.empty:
        return result.iloc[0]['Holiday'],True
    else:
        return None,False
               
def adjust_planned_shp_for_holiday(check_is_holiday,category,holiday_name,forecast_month,required_quantity_month_dict,planned_shp):
    if check_is_holiday:
        if (category == "Men's" and holiday_name in ["father_day", "men_day"]) or \
           (category != "Men's" and holiday_name not in ["father_day", "men_day"]):
            planned_shp[forecast_month] += (required_quantity_month_dict[forecast_month] * 1.15)
    return planned_shp
 
def calculate_store_sale_thru(LY_Unit_Sales, LY_MCOM_Unit_Sales, LY_OH_Units, LY_MCOM_OH_Units):
    store_sell_thru = {}
    for month in MONTHS:
        numerator = LY_Unit_Sales[month] - LY_MCOM_Unit_Sales[month]
        denominator = numerator + (LY_OH_Units[month] - LY_MCOM_OH_Units[month])
        store_sell_thru[month] =numerator / denominator if denominator != 0 else 0
        average_store_sale_thru = round((sum(store_sell_thru.values()) / len(store_sell_thru)),2)
    return average_store_sale_thru
 
def season_sum(data_dict, season_months):
    total = 0
    for month in season_months:
        value = data_dict.get(month, 0)
        
        # Handle invalid cases
        if value in (None, '', '-', 'NA', 'N/A'):
            value = 0
        
        try:
            total += float(value)
        except (ValueError, TypeError):
            print(f"Warning: Invalid data '{value}' for month {month}. Treated as 0.")
            total += 0

    return total

def safe_float(value):
    if value in (None, '', '-', 'NA', 'N/A'):
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def sum_planned_shipments(season_months, forecast_month, planned_shipments, macys_proj_receipt, omni_receipts):
    start_index = 0
    if forecast_month=='JUL' or forecast_month=='AUG':
        season_months=MONTHS
    else:
        season_months=season_months
    end_index = season_months.index(forecast_month) + 1  # Include forecast month
    selected_months = season_months[start_index:end_index]
    logging.info(f'selected_months',selected_months)
    omni_receipts_sum = sum(safe_float(omni_receipts.get(month, 0)) for month in selected_months)
    planned_shipment_sum = sum(safe_float(planned_shipments.get(month, 0)) for month in selected_months)
    macys_proj_receipt_sum = sum(safe_float(macys_proj_receipt.get(month, 0)) for month in selected_months)

    sum_of_omni_receipt_and_planned_shipment = omni_receipts_sum + planned_shipment_sum

    return sum_of_omni_receipt_and_planned_shipment, macys_proj_receipt_sum

# def sum_planned_shipments(season_months, forecast_month, planned_shipments,macys_proj_receipt,omni_receipts):
   
#     start_index = 0  # Always start from the first month in spring_months_upper
#     end_index = season_months.index(forecast_month) + 2  # Include the FCM month
   
#     selected_months = season_months[start_index:end_index]
#     omni_receipts = sum(omni_receipts.get(month, 0) for month in selected_months)
#     planned_shipment_upto_next_month_after_forecast_month = sum(planned_shipments.get(month, 0) for month in selected_months)
#     macys_proj_receipt_upto_next_month_after_forecast_month = sum(macys_proj_receipt.get(month, 0) for month in selected_months)
#     sum_of_omni_receipt_and_planned_shipment_upto_next_month_after_forecast_month = omni_receipts + planned_shipment_upto_next_month_after_forecast_month
#     return sum_of_omni_receipt_and_planned_shipment_upto_next_month_after_forecast_month ,macys_proj_receipt_upto_next_month_after_forecast_month
 
 
def calculate_seasonwise_projection(forecast_month,planned_shipments,macys_proj_receipt,omni_receipts,forecast_season_month):

    sum_of_omni_receipt_and_planned_shipment_upto_next_month_after_forecast_month ,macys_proj_receipt_upto_next_month_after_forecast_month = sum_planned_shipments(
        forecast_season_month, forecast_month, planned_shipments, macys_proj_receipt, omni_receipts
    )
 
    return (sum_of_omni_receipt_and_planned_shipment_upto_next_month_after_forecast_month ,macys_proj_receipt_upto_next_month_after_forecast_month)
 
 
def determine_percentage_to_add_quantity(average_store_sale_thru, Own_Retail):
    if average_store_sale_thru > 0.14:
        percentage = 1.0
    elif Own_Retail <= 1000 and average_store_sale_thru >= 0.10:
        percentage = 1.0
    elif Own_Retail > 1000 and average_store_sale_thru >= 0.10:
        percentage = 0.75
    elif Own_Retail <= 1000 and average_store_sale_thru >= 0.04:
        percentage = 0.75
    elif Own_Retail >= 2000 and average_store_sale_thru >= 0.04:
        percentage = 0.50
    elif Own_Retail > 1000 and average_store_sale_thru >= 0.04:
        percentage = 0.65
    else:
        percentage = 0.40
    return percentage
 
def adjust_planned_shipments_based_on_macys(
    forecast_month,
    macys_proj_receipt_upto_next_month_after_forecast_month,
    sum_of_omni_receipt_and_planned_shipment_upto_next_month_after_forecast_month,
    percentage,
    planned_shipments,
    planned_oh,
    KPI_Door_count,
    category

):
    # Calculate remaining units
    remaining_units = macys_proj_receipt_upto_next_month_after_forecast_month - sum_of_omni_receipt_and_planned_shipment_upto_next_month_after_forecast_month
    additional_units=0
    if remaining_units > 0:
        if planned_oh[forecast_month] < 3 * KPI_Door_count:
            if category != "Men's":

                additional_units = round((remaining_units * percentage), 0)
                planned_shipments[forecast_month] += additional_units
                logging.info(f'Macy additional_units: {additional_units}')
    return planned_shipments,additional_units

def handle_return_qty(planned_shipments,return_quantity_dict,forecast_month):
    # Subtract return quantities from planned shipments
    logging.info(f'return_quantity_dict: {return_quantity_dict}')
    total_sum = sum(return_quantity_dict.values())
    planned_shipments = {
        month: planned_shipments.get(month, 0) - return_quantity_dict.get(month, 0)
        for month in MONTHS

    }
    planned_shipments[forecast_month]=total_sum
    return planned_shipments

def round_to_nearest_five(value):
    return math.ceil(value / 5) * 5
 
def calculate_total_added_qty(
    total_gross_projection,
    in_transit,
    planned_shp,
):
 
    # Calculate total planned shipments
    total_planned_shipments = sum(planned_shp.get(month, 0) for month in MONTHS)
 
    # Calculate total added quantity
    total_added_quantity = total_planned_shipments - total_gross_projection - in_transit
 
    return total_added_quantity


#com
def find_average_com_oh(TY_average_COM_OH, LY_average_COM_OH, std_trend):
    """
    Determines the COM OH average value to maintain based on trend and past/current inventory levels.

    Parameters:
        TY_average_COM_OH (float): This year's average COM OH
        LY_average_COM_OH (float): Last year's average COM OH
        std_trend (float): Trend direction (negative = downtrend, positive = uptrend)

    Returns:
        float: Final COM OH value to maintain
    """
    # Case 1: Downtrend, LY low, TY high
    if std_trend < 0 and LY_average_COM_OH < TY_average_COM_OH:
        return LY_average_COM_OH  # maintain last year’s average

    # Case 2: Downtrend, LY high, TY low
    elif std_trend < 0 and LY_average_COM_OH > TY_average_COM_OH:
        return LY_average_COM_OH  # still maintain last year’s average

    # Case 3: Uptrend (or flat)
    else:
        return max(LY_average_COM_OH, TY_average_COM_OH)  # maintain the maximum
    

def calculate_planned_oh_partial(v1, k1, row_10, row_11, row_21, row_37, row_43, row_17, current_month, override_value=None):
    month_to_index = {month: idx + 1 for idx, month in enumerate(MONTHS)}
    start_idx = MONTHS.index(current_month)
    reordered_months = MONTHS[start_idx:] + MONTHS[:start_idx]

    planned_oh = {}

    for idx, month in enumerate(reordered_months):
        col4 = month_to_index[month]
        prev_month = reordered_months[idx - 1] if idx > 0 else reordered_months[-1]

        col10 = row_10.get(month, 0)
        col11 = row_11.get(month, 0)
        col21 = row_21.get(month, 0)
        col37 = row_37.get(month, 0)
        col43 = row_43.get(month, 0)
        col17 = row_17.get(month, 0)

        prev_col12 = planned_oh.get(prev_month, 0)
        prev_col21 = row_21.get(prev_month, 0)

        if idx == 0 and override_value is not None:
            val = override_value
        elif v1 == "Current MTH" and k1 == col4 and col4 == 1:
            val = row_43['JAN'] + col11 + col37 - col10
        elif v1 == "Current MTH" and k1 == col4:
            val = col21 + col11 - (col10 - col17)
        elif v1 == "Current MTH" and col4 == 1 and k1 > 1:
            val = planned_oh['JAN'] + col11 - col10
        elif v1 == "YTD" and col4 < k1:
            val = col21
        elif v1 == "Spring" and col4 < 7:
            val = col21
        elif v1 == "Fall" and col4 > 6 and col4 < k1:
            val = col21
        elif v1 == "Fall" and col4 == 1 and k1 > 1:
            val = planned_oh['JAN'] + col11 - col10
        elif v1 == "Fall" and col4 > 6 and col4 == k1:
            val = prev_col21 + col37 + col11 - col10
        elif v1 == "LY Fall" and col4 > 6:
            val = col43
        elif col4 == 1 and v1 == "LY FALL":
            val = row_43['JAN'] + col11 + col11 - col10
        elif col4 == k1 and col4 > 1:
            val = prev_col21 + col37 + col11 - col10
        else:
            val = prev_col12 + col11 - col10

        planned_oh[month] = val

    return planned_oh

 
def required_quantity_for_com(forecast_month, planned_fc, average_com_oh):
    required_quantity_month_dict = {}
    next_month_after_forecast_month = find_next_month_after_forecast_month(forecast_month)
    # Calculate FLDC for the next month
    Calculate_FLDC = round((planned_fc[next_month_after_forecast_month]) / 2)
 
    # Update the required quantity for the current month
    required_quantity_month_dict[forecast_month] =  Calculate_FLDC + average_com_oh
    required_quantity_month_dict[next_month_after_forecast_month] = average_com_oh
    return required_quantity_month_dict,Calculate_FLDC

    from datetime import datetime, timedelta

def is_day_after_23(dt: datetime) -> bool:
    """
    Returns True if the datetime object's day is 24 or more.
    """
    return dt.day >= 24

def calculate_store_unit_sales_and_OH(TY_Unit_Sales, TY_MCOM_Unit_Sales):
    result = {}
    for key in TY_Unit_Sales:
        if key in TY_MCOM_Unit_Sales:  # Ensure the key exists in both dictionaries
            sales = TY_Unit_Sales[key]
            mcom_sales = TY_MCOM_Unit_Sales[key]
            result[key] = sales - mcom_sales
        else:
            result[key] = None  # Handle missing keys in TY_MCOM_Unit_Sales
    return result
 
def calculate_com_to_ttl_sales_and_OH(TY_MCOM_Unit_Sales, TY_Unit_Sales):
    result = {}
    for key in TY_MCOM_Unit_Sales:
        if key in TY_Unit_Sales and TY_Unit_Sales[key] != 0:  # Check key exists and denominator is not zero
            mcom_sales = TY_MCOM_Unit_Sales[key]
            unit_sales = TY_Unit_Sales[key]
            result[key] = mcom_sales / unit_sales
        else:
            result[key] = 0  # Set result as 0 in case of error (e.g., division by zero or missing key)
    return result
 
def format_sales_data(PTD_TY_Sales, TY_OH_Units, own_retail):
    result = {}
    for key in PTD_TY_Sales:
        sales = PTD_TY_Sales[key]
        units = TY_OH_Units.get(key, 0)  # Using get to handle missing keys, defaulting to 0
 
        # Calculate the sales per unit
        if units != 0:
            sales_per_unit = sales / units
        else:
            sales_per_unit = 0
 
        # Format the sales per unit as currency
        formatted_sales_per_unit = f"${sales_per_unit:.2f}"
 
        # Calculate the percentage change
        if own_retail != 0 and units != 0:
            percentage_change = ((sales / units - own_retail) / own_retail) * 100
        else:
            percentage_change = 0
 
        # Format the percentage change
        formatted_percentage_change = f"{percentage_change:.0f}%"
 
        # Combine the two formatted strings
        result[key] = f"{formatted_sales_per_unit} / {formatted_percentage_change}"
 
    return result
def calculate_omni_sell_through(TY_Unit_Sales, TY_OH_Units):
    result = {}
    for key in TY_Unit_Sales:
        unit_sales = TY_Unit_Sales.get(key, 0)
        oh_units = TY_OH_Units.get(key, 0)
        total_units = unit_sales + oh_units
 
        if total_units != 0:  # Check to avoid division by zero
            result[key] = unit_sales / total_units
        else:
            result[key] = 0  # Return 0 if division by zero would occur
 
    return result
 
def calculate_store_sell_through(TY_Unit_Sales, TY_MCOM_Unit_Sales, TY_OH_Units, TY_OH_MCOM_Units):
    result = {}
    for key in set(TY_Unit_Sales.keys()).union(TY_OH_Units.keys()):
        unit_sales = TY_Unit_Sales.get(key, 0)
        mcom_unit_sales = TY_MCOM_Unit_Sales.get(key, 0)
        oh_units = TY_OH_Units.get(key, 0)
        oh_mcom_units = TY_OH_MCOM_Units.get(key, 0)
 
        # Calculate differences
        net_units_sales = unit_sales - mcom_unit_sales
        net_oh_units = oh_units - oh_mcom_units
 
        # Sum of differences
        total_diff = net_units_sales + net_oh_units
 
        # Avoid division by zero
        if total_diff != 0:
            result[key] = net_units_sales / total_diff
        else:
            result[key] = 0  # If division by zero or no net change, result is 0
 
    return result
def calculate_turn(TY_Unit_Sales, TY_OH_Units):
 
 
    months = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']
    result = {}
    cumulative_sales = []
    cumulative_oh = []
 
    for month in months:
        sales = TY_Unit_Sales.get(month, 0)
        oh_units = TY_OH_Units.get(month, 0)
 
        cumulative_sales.append(sales)
        cumulative_oh.append(oh_units)
        print(month,sales,oh_units)
        if month == 'FEB':
            turn = sales / oh_units if oh_units != 0 else 0
        else:
            non_zero_oh = [val for val in cumulative_oh if val != 0]
            avg_oh = mean(non_zero_oh) if non_zero_oh else 0
            total_sales = sum(cumulative_sales)
            turn = total_sales / avg_oh if avg_oh != 0 else 0
 
        result[month] = round_half_up(turn, 1)
 
    return result
def calculate_diff(TY_store_unit_sales, LY_store_unit_sales):
    result = {}
    for key in TY_store_unit_sales:
        ty_sales = TY_store_unit_sales.get(key, 0)
        ly_sales = LY_store_unit_sales.get(key, 0)
 
        # Check if the sum of this year's and last year's sales is zero
        if (ty_sales + ly_sales) == 0:
            result[key] = 0
        else:
            # Compute the ratio only if LY sales are not zero to avoid division by zero
            if ly_sales != 0:
                result[key] = (ty_sales - ly_sales) / ly_sales
            else:
                # Handle potential division by zero by setting to 1 (as per IFERROR)
                result[key] = 1
    return result

def calculate_planned_sell_through(planned_fc, plan_oh):
    result = {}
    for key in planned_fc:
        if key in plan_oh:  # Ensuring the key exists in both dictionaries
            fc = planned_fc[key]
            oh = plan_oh[key]
            if (fc + oh) != 0:  # Avoid division by zero
                result[key] = round(fc / (fc + oh), 2)
            else:
                result[key] = None  # Handle division by zero, if any
        else:
            result[key] = None  # Handle missing keys in plan_oh
    return result

def find_STD(month_map,Month1,Month2):
    abbr1 = month_map[Month1]
    abbr2 = month_map[Month2]
 
    # Find indices
    start_index = MONTHS.index(abbr1)
    end_index = MONTHS.index(abbr2)
 
    # Slice the list
    Std_PID = MONTHS[start_index:end_index + 1]
    return Std_PID

def calculate_index_value(Current_FC_Index):

    # print("Current_FC_index",Current_FC_Index)
    # if sheets:
    #     index_df_raw = sheets["Index"] 
    # print("Current_FC_index",Current_FC_Index)
    # index_df = index_df_raw.iloc[2:43, :16]

    data = [
        ["Amy", 0.23, 0.03, 0.06, 0.08, 0.05, 0.03, 0.03, 0.04, 0.04, 0.14, 0.24, 0.02, 0.32, 0.53, 0.15],
        ["Anklet", 0.06, 0.07, 0.11, 0.12, 0.10, 0.10, 0.06, 0.07, 0.05, 0.10, 0.14, 0.01, 0.24, 0.52, 0.28],
        ["Aqua", 0.08, 0.23, 0.07, 0.07, 0.04, 0.03, 0.04, 0.04, 0.03, 0.10, 0.25, 0.01, 0.38, 0.50, 0.14],
        ["Bridal", 0.16, 0.06, 0.08, 0.05, 0.05, 0.03, 0.12, 0.07, 0.09, 0.14, 0.14, 0.01, 0.30, 0.59, 0.31],
        ["BT", 0.06, 0.05, 0.05, 0.07, 0.05, 0.05, 0.04, 0.04, 0.04, 0.18, 0.37, 0.00, 0.15, 0.72, 0.17],
        ["Citrine", 0.06, 0.03, 0.04, 0.07, 0.04, 0.03, 0.05, 0.05, 0.07, 0.30, 0.24, 0.01, 0.14, 0.74, 0.20],
        ["Cross", 0.07, 0.07, 0.10, 0.08, 0.07, 0.05, 0.06, 0.06, 0.06, 0.11, 0.24, 0.01, 0.25, 0.59, 0.23],
        ["CZ", 0.08, 0.07, 0.07, 0.08, 0.07, 0.06, 0.06, 0.07, 0.06, 0.11, 0.24, 0.02, 0.22, 0.60, 0.25],
        ["Dia", 0.13, 0.03, 0.07, 0.08, 0.05, 0.04, 0.08, 0.06, 0.08, 0.13, 0.23, 0.02, 0.23, 0.62, 0.25],
        ["Ear", 0.04, 0.11, 0.13, 0.11, 0.09, 0.08, 0.06, 0.06, 0.03, 0.12, 0.15, 0.02, 0.29, 0.49, 0.23],
        ["EMER", 0.07, 0.05, 0.10, 0.16, 0.06, 0.05, 0.05, 0.04, 0.04, 0.13, 0.24, 0.01, 0.22, 0.54, 0.17],
        ["Garnet", 0.09, 0.04, 0.05, 0.08, 0.05, 0.04, 0.03, 0.04, 0.04, 0.12, 0.30, 0.12, 0.18, 0.56, 0.15],
        ["Gem", 0.08, 0.06, 0.10, 0.09, 0.10, 0.04, 0.05, 0.06, 0.06, 0.12, 0.22, 0.03, 0.24, 0.54, 0.21],
        ["GEM EAR", 0.09, 0.05, 0.07, 0.07, 0.06, 0.04, 0.06, 0.08, 0.08, 0.14, 0.24, 0.02, 0.21, 0.64, 0.27],
        ["Gold Chain", 0.07, 0.05, 0.10, 0.07, 0.08, 0.04, 0.06, 0.06, 0.09, 0.17, 0.22, 0.00, 0.22, 0.63, 0.25],
        ["GOLD EAR", 0.09, 0.05, 0.06, 0.11, 0.07, 0.04, 0.06, 0.09, 0.08, 0.11, 0.23, 0.01, 0.20, 0.61, 0.27],
        ["Heart", 0.11, 0.05, 0.06, 0.08, 0.06, 0.05, 0.05, 0.05, 0.05, 0.13, 0.29, 0.02, 0.22, 0.63, 0.20],
        ["Heavy Gold Chain", 0.05, 0.07, 0.11, 0.07, 0.12, 0.04, 0.07, 0.07, 0.09, 0.13, 0.19, 0.00, 0.22, 0.59, 0.27],
        ["Jade", 0.08, 0.07, 0.07, 0.06, 0.05, 0.05, 0.05, 0.08, 0.09, 0.10, 0.25, 0.03, 0.22, 0.63, 0.27],
        ["KIDS", 0.07, 0.08, 0.08, 0.06, 0.07, 0.06, 0.07, 0.07, 0.07, 0.15, 0.20, 0.02, 0.23, 0.62, 0.27],
        ["Locket", 0.06, 0.06, 0.09, 0.10, 0.05, 0.04, 0.04, 0.06, 0.05, 0.12, 0.33, 0.00, 0.21, 0.64, 0.19],
        ["Mens Gold Bracelet", 0.10, 0.06, 0.11, 0.06, 0.11, 0.04, 0.10, 0.04, 0.10, 0.09, 0.18, 0.00, 0.27, 0.55, 0.29],
        ["Mens Misc", 0.13, 0.07, 0.16, 0.07, 0.07, 0.06, 0.08, 0.08, 0.06, 0.09, 0.12, 0.01, 0.36, 0.48, 0.28],
        ["Mens Silver chain", 0.06, 0.06, 0.07, 0.07, 0.10, 0.05, 0.06, 0.06, 0.07, 0.13, 0.28, 0.00, 0.20, 0.63, 0.23],
        ["Mom", 0.05, 0.03, 0.12, 0.25, 0.03, 0.03, 0.03, 0.03, 0.03, 0.09, 0.31, 0.01, 0.19, 0.52, 0.12],
        ["MOP", 0.13, 0.07, 0.06, 0.04, 0.05, 0.07, 0.05, 0.08, 0.06, 0.13, 0.25, 0.02, 0.25, 0.64, 0.26],
        ["Neck", 0.09, 0.05, 0.08, 0.07, 0.06, 0.06, 0.05, 0.07, 0.08, 0.19, 0.19, 0.01, 0.22, 0.64, 0.26],
        ["Onyx", 0.10, 0.08, 0.09, 0.07, 0.06, 0.04, 0.05, 0.07, 0.07, 0.12, 0.23, 0.02, 0.28, 0.58, 0.23],
        ["Opal", 0.05, 0.03, 0.05, 0.07, 0.05, 0.04, 0.04, 0.10, 0.17, 0.12, 0.26, 0.01, 0.13, 0.73, 0.35],
        ["Pearl", 0.07, 0.06, 0.08, 0.08, 0.06, 0.04, 0.06, 0.07, 0.08, 0.12, 0.27, 0.02, 0.21, 0.64, 0.25],
        ["Peridot", 0.05, 0.03, 0.05, 0.07, 0.05, 0.07, 0.29, 0.05, 0.03, 0.07, 0.24, 0.01, 0.12, 0.74, 0.43],
        ["Religious", 0.07, 0.07, 0.10, 0.08, 0.07, 0.05, 0.06, 0.06, 0.06, 0.11, 0.24, 0.01, 0.25, 0.59, 0.23],
        ["Ring", 0.11, 0.07, 0.11, 0.07, 0.07, 0.05, 0.06, 0.07, 0.07, 0.13, 0.18, 0.00, 0.29, 0.57, 0.26],
        ["Ruby", 0.07, 0.05, 0.06, 0.08, 0.08, 0.13, 0.05, 0.05, 0.04, 0.14, 0.23, 0.02, 0.18, 0.64, 0.27],
        ["Saph", 0.06, 0.04, 0.06, 0.07, 0.05, 0.03, 0.06, 0.17, 0.05, 0.14, 0.26, 0.01, 0.16, 0.71, 0.31],
        ["Womens Silver Chain", 0.06, 0.06, 0.07, 0.06, 0.07, 0.06, 0.06, 0.07, 0.06, 0.17, 0.23, 0.02, 0.19, 0.66, 0.26],
        ["Wrist", 0.06, 0.06, 0.08, 0.07, 0.07, 0.05, 0.06, 0.07, 0.07, 0.14, 0.25, 0.02, 0.20, 0.64, 0.25],
        ["Grand Total", 0.07, 0.06, 0.08, 0.08, 0.07, 0.05, 0.06, 0.07, 0.07, 0.14, 0.23, 0.01, 0.21, 0.62, 0.25]
    ]
    columns= [
        "Index", "FEB", 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT','NOV', 'DEC', 'JAN',
        "STD Index", "F6MTH", "F4MTH"
    ]
    # Append remaining data
    df_full = pd.DataFrame(data, columns=columns)
    index_row_data = df_full.loc[df_full['Index'].astype(str).str.lower() == Current_FC_Index.lower()]
    print("Index row data :",index_row_data)
    
    index_value = {}
    # Loop through each month and fetch its value
    for month in MONTHS:
        index_value[month] = index_row_data[month].iloc[0] if not index_row_data.empty else 0
    print("Index value :",index_value)

    return index_value

import os 
from openpyxl import load_workbook


def get_c2_value(category,pid,std_trend,STD_index_value,currect_fc_index,month_12_fc_index,forecasting_method,planned_shp,planned_fc,path):
    import forecast.service.staticVariable as st
    base_dir = os.path.join("media/processed_files", path) 
    
    filename = f"{category}.xlsx"
    filepath = os.path.join(base_dir, filename)
 
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel file not found: {filepath}")
   
    # Load workbook and get the active sheet
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_col=3):  # Only search in Column C (3rd col)
        if row[2].value == pid:  # Column C → index 2 (0-based)
            start_row= row[2].row
 
    ws[f"F{start_row + 3}"]= std_trend
    ws[f"F{start_row + 1}"]= currect_fc_index
    ws[f"F{start_row + 2}"]= month_12_fc_index  # Added IFERROR to prevent divide-by-zero issues
    ws[f"F{start_row + 4}"]= forecasting_method
    ws[f"I{start_row + 5}"]=planned_fc["FEB"]
    ws[f"J{start_row + 5}"]=planned_fc["MAR"]
    ws[f"K{start_row + 5}"]=planned_fc["APR"]
    ws[f"L{start_row + 5}"]=planned_fc["MAY"]
    ws[f"M{start_row + 5}"]=planned_fc["JUN"]
    ws[f"N{start_row + 5}"]= planned_fc["JUL"]
    ws[f"O{start_row + 5}"]=planned_fc["AUG"]
    ws[f"P{start_row + 5}"]=planned_fc["SEP"]
    ws[f"Q{start_row + 5}"]=planned_fc["OCT"]
    ws[f"R{start_row + 5}"]=planned_fc["NOV"]
    ws[f"S{start_row + 5}"]=planned_fc["DEC"]
    ws[f"T{start_row + 5}"]=planned_fc["JAN"]
 
    ws[f"I{start_row + 6}"]= planned_shp["FEB"]
    ws[f"J{start_row + 6}"]= planned_shp["MAR"]
    ws[f"K{start_row + 6}"]=planned_shp["APR"]
    ws[f"L{start_row + 6}"]= planned_shp["MAY"]
    ws[f"M{start_row + 6}"]= planned_shp["JUN"]
    ws[f"N{start_row + 6}"]= planned_shp["JUL"]
    ws[f"O{start_row + 6}"]=planned_shp["AUG"]
    ws[f"P{start_row + 6}"]=planned_shp["SEP"]
    ws[f"Q{start_row + 6}"]=planned_shp["OCT"]
    ws[f"R{start_row + 6}"]=planned_shp["NOV"]
    ws[f"S{start_row + 6}"]=planned_shp["DEC"]
    ws[f"T{start_row + 6}"]=planned_shp["JAN"]
 
    wb.save(filepath)